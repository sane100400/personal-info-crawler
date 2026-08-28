"""Create a compact, restricted workbook for human URL review."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def write_labeling_workbook(
    path: Path,
    rows: list[dict[str, str]],
    source_urls: dict[str, str],
) -> None:
    """Write one compact row per URL; reviewers inspect the linked source."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "라벨링"
    headers = [
        "번호",
        "sample_id",
        "원문 링크",
        "도메인",
        "제목",
        "판정",
        "메모",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, row in enumerate(rows, start=1):
        sheet.append(
            [
                index,
                row["sample_id"],
                "원문 열기",
                row.get("registrable_domain", ""),
                row.get("masked_title", ""),
                "",
                "",
            ]
        )
        excel_row = index + 1
        link_cell = sheet.cell(excel_row, 3)
        link_cell.hyperlink = source_urls[row["sample_id"]]
        link_cell.style = "Hyperlink"
        link_cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(1, 8):
            sheet.cell(excel_row, column).alignment = Alignment(
                vertical="top",
                wrap_text=column in {5, 7},
            )
        sheet.row_dimensions[excel_row].height = 42

    widths = {
        "A": 7,
        "B": 14,
        "C": 12,
        "D": 22,
        "E": 55,
        "F": 12,
        "G": 55,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 28
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = f"A1:G{len(rows) + 1}"

    validation = DataValidation(
        type="list",
        formula1='"정탐,오탐,보류"',
        allow_blank=True,
    )
    validation.promptTitle = "판정 선택"
    validation.prompt = "원문을 확인한 뒤 정탐, 오탐, 보류 중 하나를 선택하세요."
    validation.errorTitle = "입력값 확인"
    validation.error = "정탐, 오탐, 보류 중 하나만 입력할 수 있습니다."
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"F2:F{len(rows) + 1}")

    for value, color in (
        ("정탐", "C6EFCE"),
        ("오탐", "FFC7CE"),
        ("보류", "FFEB9C"),
    ):
        sheet.conditional_formatting.add(
            f"F2:F{len(rows) + 1}",
            FormulaRule(
                formula=[f'$F2="{value}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )

    guide = workbook.create_sheet("안내")
    guide_rows = [
        ("항목", "판단 기준"),
        (
            "정탐",
            "개인정보 DB·계정·신분증·통장 등이 거래 대상이고, "
            "작성자의 직접적인 판매·매입·제작·중개 의사가 확인되는 원게시물",
        ),
        (
            "오탐",
            "뉴스·피해 사례·신고 안내·정상 서비스 소개·검색결과처럼 "
            "작성자의 직접적인 거래 의사가 없는 글",
        ),
        (
            "보류",
            "링크에 접근할 수 없거나 거래 의사·원게시물 여부를 확정하기 어려운 글",
        ),
        (
            "확인 방법",
            "라벨링 시트의 원문 열기를 눌러 게시물을 직접 확인한 뒤 판정합니다. "
            "수집 본문이나 자동 판정 결과만으로 정답을 정하지 않습니다.",
        ),
        (
            "주의",
            "개인정보 원문이나 연락처가 없다는 이유만으로 오탐 처리하지 않습니다. "
            "원문에 연락하거나 첨부파일을 내려받지 않습니다.",
        ),
    ]
    for item in guide_rows:
        guide.append(item)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    guide.column_dimensions["A"].width = 14
    guide.column_dimensions["B"].width = 100
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(2, len(guide_rows) + 1):
        guide.row_dimensions[row_number].height = 45
    guide.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
