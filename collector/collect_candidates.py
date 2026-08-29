#!/usr/bin/env python3
"""Collect and immediately de-identify public-web research candidates.

The script intentionally does not support login, CAPTCHA handling, attachments,
or access-control bypass. Search result pages are used only to discover final
content URLs; they never become dataset samples.
"""

from __future__ import annotations

import argparse
import base64
import copy
import csv
import datetime as dt
import hashlib
import hmac
import ipaddress
import json
import math
import os
import random
import re
import secrets
import socket
import subprocess
import sys
import time
import unicodedata
from collections import Counter, defaultdict, deque
from collections.abc import Iterable
from dataclasses import asdict, dataclass, field
from pathlib import Path
from urllib.parse import (
    parse_qs,
    parse_qsl,
    quote_plus,
    urlencode,
    urljoin,
    urlsplit,
    urlunsplit,
)

import requests
import trafilatura
import yaml
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from collector.labeling_workbook import write_labeling_workbook

SCHEMA = [
    "sample_id",
    "collected_at",
    "source_type",
    "registrable_domain",
    "source_unit_kind",
    "source_unit_hmac",
    "url_hmac",
    "http_status",
    "final_url_hmac",
    "page_type",
    "live_status",
    "extraction_status",
    "masked_title",
    "masked_text",
    "language_mix",
    "obfuscation_type",
    "intent_label",
    "target_label",
    "contact_label",
    "final_label",
    "evidence_spans",
    "annotator_1",
    "annotator_2",
    "adjudicated_label",
    "near_duplicate_cluster",
    "near_duplicate_fingerprint",
    "campaign_group",
]

LOG_SCHEMA = [
    "url_hmac",
    "query_group",
    "outcome",
    "http_status",
    "reason",
    "attempted_at",
    "text_chars",
    "extraction_method",
]

EXTRACTION_FAILURE_SCHEMA = [
    "url_hmac",
    "query_group",
    "attempted_at",
    "http_status",
    "reason",
    "extraction_status",
    "text_chars",
    "extraction_method",
]

KEYWORD_EXPANSION_SCHEMA = [
    "round_number",
    "query_group",
    "detection_type",
    "query",
    "document_frequency",
    "domain_frequency",
]

RESTRICTED_REVIEW_SCHEMA = [
    "sample_id",
    "collected_at",
    "source_url",
    "registrable_domain",
    "title",
    "text",
]

SEARCH_HOSTS = {
    "google.com",
    "www.google.com",
    "google.co.kr",
    "www.google.co.kr",
    "bing.com",
    "www.bing.com",
    "search.naver.com",
    "naver.com",
    "www.naver.com",
    "duckduckgo.com",
    "html.duckduckgo.com",
    "search.daum.net",
}

BLOCKED_EXTENSIONS = {
    ".7z",
    ".apk",
    ".avi",
    ".bin",
    ".csv",
    ".doc",
    ".docx",
    ".dmg",
    ".exe",
    ".gif",
    ".gz",
    ".hwp",
    ".hwpx",
    ".iso",
    ".jpeg",
    ".jpg",
    ".json",
    ".mkv",
    ".mov",
    ".mp3",
    ".mp4",
    ".msi",
    ".pdf",
    ".png",
    ".ppt",
    ".pptx",
    ".rar",
    ".tar",
    ".tgz",
    ".torrent",
    ".wav",
    ".webp",
    ".xls",
    ".xlsx",
    ".xml",
    ".xz",
    ".zip",
}

USER_AGENT = (
    "PersonalInfoIllicitPostCrawler/0.1 (public-web; no-login; contact: research-team)"
)
MAX_HTML_BYTES = 1_000_000
MAX_TEXT_CHARS = 20_000
DEFAULT_MIN_TEXT_CHARS = 80
CONNECT_TIMEOUT_SECONDS = 4
READ_TIMEOUT_SECONDS = 10
MAX_PROVIDER_NAVIGATION_ERRORS = 3

COLLECTION_TYPES = (
    "개인정보DB",
    "계정·아이디·가입인증",
    "통장·계좌",
    "신분증·여권 위조/제작",
)

SOCIAL_ACCOUNT_HOSTS = {
    "facebook.com",
    "instagram.com",
    "m.facebook.com",
    "mobile.twitter.com",
    "t.me",
    "telegram.me",
    "threads.net",
    "tiktok.com",
    "twitter.com",
    "www.facebook.com",
    "www.instagram.com",
    "www.threads.net",
    "www.tiktok.com",
    "www.youtube.com",
    "www.x.com",
    "x.com",
    "youtube.com",
}


@dataclass
class Candidate:
    url: str
    query_group: str
    detection_type: str
    source_type: str = "search"
    discovery_text: str = ""
    search_provider: str = ""


@dataclass(frozen=True)
class QuerySpec:
    group: str
    detection_type: str
    query: str


@dataclass(frozen=True)
class KeywordExpansion:
    round_number: int
    query_group: str
    detection_type: str
    query: str
    document_frequency: int
    domain_frequency: int


@dataclass
class CollectionLog:
    url_hmac: str
    query_group: str
    outcome: str
    http_status: str = ""
    reason: str = ""
    text_chars: int = 0
    extraction_method: str = ""
    attempted_at: str = field(
        default_factory=lambda: dt.datetime.now(
            dt.timezone(dt.timedelta(hours=9))
        ).isoformat(timespec="seconds")
    )


@dataclass
class DetectionEntry:
    detected_on: dt.date
    url: str
    detection_type: str
    registrant: str
    note: str = "자동 수집 후보·사람 검토 필요"


class RateLimiter:
    def __init__(self, delay_seconds: float) -> None:
        self.delay = delay_seconds
        self.last_request: dict[str, float] = defaultdict(lambda: 0.0)

    def wait(self, host: str) -> None:
        elapsed = time.monotonic() - self.last_request[host]
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)
        self.last_request[host] = time.monotonic()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=int, default=200)
    parser.add_argument("--cdp", default="127.0.0.1:9222")
    parser.add_argument("--out", type=Path, default=Path("output"))
    parser.add_argument("--template", type=Path, default=Path("(양식) 탐지내역.xlsx"))
    parser.add_argument("--registrant", default="", help="탐지내역 양식의 등록자 이름")
    parser.add_argument(
        "--skip-detection-workbook",
        action="store_true",
        help="원 URL 제출용 Excel은 만들지 않고 마스킹된 연구용 CSV만 저장",
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument(
        "--queries",
        type=Path,
        help="비공개 검색어 YAML 파일(예: config/queries.local.yaml)",
    )
    source.add_argument(
        "--seed-file",
        action="append",
        type=Path,
        help="비공개 URL 시드 CSV 또는 줄 단위 텍스트 파일(반복 지정 가능)",
    )
    parser.add_argument(
        "--seed-offset",
        type=int,
        default=0,
        help="시드 후보 앞부분을 건너뛸 개수(독립 후속 파일럿용)",
    )
    parser.add_argument(
        "--exclude-csv",
        action="append",
        type=Path,
        default=[],
        help="이전 표본 CSV의 source_url을 후보에서 제외(반복 지정 가능)",
    )
    parser.add_argument("--search-delay", type=float, default=3.0)
    parser.add_argument("--domain-delay", type=float, default=2.0)
    parser.add_argument("--search-pages", type=int, default=2)
    parser.add_argument(
        "--search-page-offset",
        type=int,
        default=0,
        help="검색 공급자별로 건너뛸 결과 페이지 수(0은 첫 페이지부터)",
    )
    parser.add_argument(
        "--search-provider",
        action="append",
        choices=(
            "naver",
            "naver_blog",
            "naver_cafe",
            "naver_kin",
            "naver_news",
            "daum",
            "daum_blog",
            "bing",
            "duckduckgo",
            "google",
            "google_api",
        ),
        help="사용할 검색 공급자(반복 지정 가능, 기본은 전체)",
    )
    parser.add_argument(
        "--google-api-key-env",
        default="GOOGLE_CSE_API_KEY",
        help="Google 검색 API 키를 읽을 환경변수 이름",
    )
    parser.add_argument(
        "--google-cse-id-env",
        default="GOOGLE_CSE_ID",
        help="Google Programmable Search Engine ID를 읽을 환경변수 이름",
    )
    parser.add_argument(
        "--query-variants",
        type=int,
        default=1,
        help="검색어별 자동 변형 개수(1은 원본만 사용)",
    )
    parser.add_argument(
        "--strict-search",
        action="store_true",
        help="검색어에 구문 일치와 일반 문서 제외 조건을 적용",
    )
    parser.add_argument(
        "--relevance-gate",
        choices=("off", "labeling", "review", "intent", "strict"),
        default="off",
        help="목적지 본문에서 개인정보 대상·거래 의사 문맥을 선별",
    )
    parser.add_argument(
        "--discovery-relevance-gate",
        choices=("off", "labeling", "review", "intent", "strict"),
        default=None,
        help="검색 요약 예비 필터(생략하면 --relevance-gate와 같음)",
    )
    parser.add_argument(
        "--provider-stale-pages",
        type=int,
        default=12,
        help="새 후보가 없을 때 검색 공급자를 바꿀 연속 페이지 수(0은 비활성화)",
    )
    parser.add_argument(
        "--keyword-expansion-rounds",
        type=int,
        default=0,
        help="고관련 검색 요약에서 새 2어절 검색어를 만들어 반복할 횟수",
    )
    parser.add_argument(
        "--keyword-expansion-per-round",
        type=int,
        default=20,
        help="키워드 확장 라운드마다 추가할 최대 검색어 수",
    )
    parser.add_argument(
        "--keyword-expansion-min-domains",
        type=int,
        default=2,
        help="확장 검색어를 채택하는 데 필요한 서로 다른 출처 도메인 수",
    )
    parser.add_argument(
        "--keyword-expansion-require-contact",
        action="store_true",
        help="연락 수단이 함께 나온 대상어 조합만 확장 검색어로 채택",
    )
    parser.add_argument("--min-text-chars", type=int, default=DEFAULT_MIN_TEXT_CHARS)
    parser.add_argument(
        "--min-korean-chars",
        type=int,
        default=0,
        help="제목·본문에 필요한 최소 한글 음절 수(0은 비활성화)",
    )
    parser.add_argument("--checkpoint-every", type=int, default=25)
    parser.add_argument(
        "--follow-links-per-page",
        type=int,
        default=0,
        help="본문 수집에 성공한 페이지에서 추가할 관련 공개 내부 링크 수",
    )
    parser.add_argument(
        "--candidate-pool-limit",
        type=int,
        default=0,
        help="추가 발견을 포함한 최대 후보 수(0은 목표의 4배)",
    )
    parser.add_argument(
        "--max-candidates-per-domain",
        type=int,
        default=100,
        help="검색 발견·내부 링크 확장을 합친 도메인당 후보 수(0은 제한 없음)",
    )
    parser.add_argument(
        "--max-candidates-per-source-unit",
        type=int,
        default=10,
        help="같은 SNS 계정·채널 또는 게시판에서 확인할 후보 수 상한",
    )
    parser.add_argument(
        "--max-records-per-domain",
        type=int,
        default=0,
        help="최종 표본에 보존할 도메인별 절대 상한(0은 제한 없음)",
    )
    parser.add_argument(
        "--max-domain-share",
        type=float,
        default=0.0,
        help="최종 표본에서 단일 도메인의 최대 비율(기본 비활성화)",
    )
    parser.add_argument(
        "--max-records-per-source-unit",
        type=int,
        default=1,
        help="같은 SNS 계정·채널 또는 게시판에서 보존할 최대 대표 게시물 수",
    )
    parser.add_argument(
        "--min-domains",
        type=int,
        default=0,
        help="완료 판정에 필요한 최소 도메인 수(0은 도메인 상한에서 자동 계산)",
    )
    parser.add_argument(
        "--min-type-share",
        type=float,
        default=0.05,
        help="네 가지 수집 유형별 최소 비율(기본 0.05, 0은 비활성화)",
    )
    parser.add_argument(
        "--max-records-per-campaign",
        type=int,
        default=1,
        help="동일 연락처 캠페인에서 보존할 최대 건수(0은 제한 없음)",
    )
    parser.add_argument(
        "--refresh-discovery",
        action="store_true",
        help="재개 시 저장된 후보 큐 대신 검색을 다시 실행해 새 후보를 추가",
    )
    parser.add_argument(
        "--expand-existing-links",
        action="store_true",
        help="재개 시 기존 성공 페이지를 표본에 중복 저장하지 않고 내부 링크 탐색에 사용",
    )
    parser.add_argument(
        "--revalidate-existing",
        action="store_true",
        help="재개 전 기존 공유 데이터를 현재 본문 게이트·출처·캠페인 기준으로 다시 검증",
    )
    parser.add_argument("--resume", action="store_true")
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if args.target < 1:
        raise ValueError("--target must be at least 1")
    if args.search_pages < 1 or args.search_pages > 20:
        raise ValueError("--search-pages must be between 1 and 20")
    if args.query_variants < 1 or args.query_variants > 10:
        raise ValueError("--query-variants must be between 1 and 10")
    if args.seed_offset < 0:
        raise ValueError("--seed-offset cannot be negative")
    if args.seed_offset and not args.seed_file:
        raise ValueError("--seed-offset requires --seed-file")
    missing_exclusions = [path for path in args.exclude_csv if not path.is_file()]
    if missing_exclusions:
        raise FileNotFoundError(
            f"Excluded CSV does not exist: {missing_exclusions[0]}"
        )
    if args.search_delay < 0 or args.domain_delay < 0:
        raise ValueError("Request delays cannot be negative")
    if args.search_page_offset < 0 or args.search_page_offset > 100:
        raise ValueError("--search-page-offset must be between 0 and 100")
    if args.min_text_chars < 40 or args.min_text_chars > MAX_TEXT_CHARS:
        raise ValueError(f"--min-text-chars must be between 40 and {MAX_TEXT_CHARS}")
    if args.min_korean_chars < 0 or args.min_korean_chars > MAX_TEXT_CHARS:
        raise ValueError(f"--min-korean-chars must be between 0 and {MAX_TEXT_CHARS}")
    if args.checkpoint_every < 1 or args.checkpoint_every > 500:
        raise ValueError("--checkpoint-every must be between 1 and 500")
    if args.follow_links_per_page < 0 or args.follow_links_per_page > 20:
        raise ValueError("--follow-links-per-page must be between 0 and 20")
    if args.candidate_pool_limit < 0:
        raise ValueError("--candidate-pool-limit cannot be negative")
    if args.max_candidates_per_domain < 0:
        raise ValueError("--max-candidates-per-domain cannot be negative")
    if args.max_candidates_per_source_unit < 0:
        raise ValueError("--max-candidates-per-source-unit cannot be negative")
    if args.max_records_per_domain < 0:
        raise ValueError("--max-records-per-domain cannot be negative")
    if not 0 <= args.max_domain_share <= 1:
        raise ValueError("--max-domain-share must be between 0 and 1")
    if args.max_records_per_source_unit < 1:
        raise ValueError("--max-records-per-source-unit must be at least 1")
    if args.min_domains < 0:
        raise ValueError("--min-domains cannot be negative")
    if not 0 <= args.min_type_share <= 0.25:
        raise ValueError("--min-type-share must be between 0 and 0.25")
    if args.max_records_per_campaign < 0:
        raise ValueError("--max-records-per-campaign cannot be negative")
    if args.provider_stale_pages < 0 or args.provider_stale_pages > 1000:
        raise ValueError("--provider-stale-pages must be between 0 and 1000")
    if args.keyword_expansion_rounds < 0 or args.keyword_expansion_rounds > 5:
        raise ValueError("--keyword-expansion-rounds must be between 0 and 5")
    if (
        args.keyword_expansion_per_round < 1
        or args.keyword_expansion_per_round > 100
    ):
        raise ValueError("--keyword-expansion-per-round must be between 1 and 100")
    if (
        args.keyword_expansion_min_domains < 1
        or args.keyword_expansion_min_domains > 20
    ):
        raise ValueError("--keyword-expansion-min-domains must be between 1 and 20")
    if args.seed_file and args.keyword_expansion_rounds:
        raise ValueError("Keyword expansion requires --queries, not --seed-file")
    if args.refresh_discovery and (not args.resume or not args.queries):
        raise ValueError("--refresh-discovery requires --resume and --queries")
    if args.revalidate_existing and not args.resume:
        raise ValueError("--revalidate-existing requires --resume")
    if args.revalidate_existing and not args.skip_detection_workbook:
        raise ValueError(
            "--revalidate-existing currently requires --skip-detection-workbook"
        )
    if args.expand_existing_links and (
        not args.resume or not args.seed_file or not args.follow_links_per_page
    ):
        raise ValueError(
            "--expand-existing-links requires --resume, --seed-file, and "
            "--follow-links-per-page"
        )
    if args.search_provider and "google_api" in args.search_provider:
        if not os.environ.get(args.google_api_key_env, "").strip():
            raise ValueError(
                f"Google API key environment variable is empty: "
                f"{args.google_api_key_env}"
            )
        if not os.environ.get(args.google_cse_id_env, "").strip():
            raise ValueError(
                f"Google CSE ID environment variable is empty: "
                f"{args.google_cse_id_env}"
            )
    if not args.skip_detection_workbook and not args.registrant.strip():
        raise ValueError("--registrant cannot be empty")


def get_or_create_hmac_key(private_dir: Path) -> bytes:
    private_dir.mkdir(parents=True, exist_ok=True)
    os.chmod(private_dir, 0o700)
    key_path = private_dir / "url_hmac_key"
    if key_path.exists():
        key = key_path.read_bytes().strip()
        if len(key) < 32:
            raise RuntimeError("Existing HMAC key is unexpectedly short")
        return key
    key = secrets.token_hex(32).encode("ascii")
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    fd = os.open(key_path, flags, 0o600)
    with os.fdopen(fd, "wb") as handle:
        handle.write(key + b"\n")
    return key


def canonicalize_url(raw: str) -> str | None:
    try:
        parts = urlsplit(raw)
    except ValueError:
        return None
    if parts.scheme.lower() not in {"http", "https"} or not parts.hostname:
        return None
    if parts.username or parts.password:
        return None
    path_lower = parts.path.lower().rstrip("/") or "/"
    if any(path_lower.endswith(ext) for ext in BLOCKED_EXTENSIONS):
        return None
    # Fragments and common analytics parameters do not identify a separate document.
    tracking_names = {"fbclid", "gclid", "dclid", "msclkid", "ref_src"}
    kept = [
        (name, value)
        for name, value in parse_qsl(parts.query, keep_blank_values=True)
        if not name.lower().startswith("utm_") and name.lower() not in tracking_names
    ]
    host = (parts.hostname or "").lower()
    path = parts.path or "/"
    # Naver exposes the same blog document through desktop/mobile hosts and
    # sometimes appends navigation parameters. Treat those representations as
    # one document so a prior holdout cannot re-enter a later sample.
    segments = [segment for segment in path.split("/") if segment]
    if host in {"blog.naver.com", "m.blog.naver.com"}:
        if len(segments) == 2 and segments[1].isdigit():
            host = "m.blog.naver.com"
            path = f"/{segments[0]}/{segments[1]}"
            kept = []
    if host in {"creativebox.kr", "www.creativebox.kr"}:
        if (
            len(segments) == 2
            and segments[0]
            in {"igtrade", "ttmarket", "fbmarket", "etcmarket", "trade"}
            and segments[1].isdigit()
        ):
            host = "creativebox.kr"
            path = f"/{segments[0]}/{segments[1]}"
            kept = []
    normalized_netloc = f"[{host}]" if ":" in host else host
    if parts.port:
        normalized_netloc += f":{parts.port}"
    return urlunsplit(
        (
            parts.scheme.lower(),
            normalized_netloc,
            path,
            urlencode(kept),
            "",
        )
    )


def public_content_fallback_url(url: str) -> str | None:
    """Return an equivalent public content URL for known frame-only pages."""
    parts = urlsplit(url)
    host = (parts.hostname or "").lower()
    segments = [segment for segment in parts.path.split("/") if segment]
    if host == "blog.naver.com" and len(segments) == 2 and segments[1].isdigit():
        return urlunsplit(
            ("https", "m.blog.naver.com", f"/{segments[0]}/{segments[1]}", "", "")
        )
    return None


def unwrap_search_result_url(raw: str) -> str:
    """Resolve only transparent search redirect encodings, without requesting them."""
    try:
        parts = urlsplit(raw)
        host = (parts.hostname or "").lower()
        query = parse_qs(parts.query)
        if host.endswith("bing.com") and parts.path == "/ck/a":
            encoded = query.get("u", [""])[0]
            if encoded.startswith("a1"):
                payload = encoded[2:]
                payload += "=" * (-len(payload) % 4)
                decoded = base64.urlsafe_b64decode(payload).decode("utf-8")
                if decoded.startswith(("http://", "https://")):
                    return decoded
        if host.endswith("google.com") and parts.path == "/url":
            target = query.get("q", [""])[0]
            if target.startswith(("http://", "https://")):
                return target
        if host.endswith("duckduckgo.com") and parts.path.startswith("/l/"):
            target = query.get("uddg", [""])[0]
            if target.startswith(("http://", "https://")):
                return target
    except (ValueError, UnicodeDecodeError):
        pass
    return raw


def registrable_domain(host: str) -> str:
    host = host.lower().strip(".")
    labels = host.split(".")
    common_second_level = {
        "co.kr",
        "or.kr",
        "go.kr",
        "ac.kr",
        "ne.kr",
        "com.au",
        "co.jp",
        "co.uk",
    }
    if len(labels) >= 3 and ".".join(labels[-2:]) in common_second_level:
        return ".".join(labels[-3:])
    return ".".join(labels[-2:]) if len(labels) >= 2 else host


def source_unit_descriptor(
    url: str,
    title: str = "",
    text: str = "",
    allow_unresolved_social_post: bool = False,
) -> tuple[str, str]:
    """Return the counting unit: SNS account/channel, board, or standalone site."""
    parts = urlsplit(url)
    host = (parts.hostname or "").lower()
    domain = registrable_domain(host)
    segments = [segment for segment in parts.path.split("/") if segment]
    query = parse_qs(parts.query)

    if host in {"t.me", "telegram.me", "www.telegram.me"}:
        account_segments = segments[1:] if segments[:1] == ["s"] else segments
        if account_segments and account_segments[0] not in {"joinchat", "share"}:
            return "social_account", f"telegram:{account_segments[0].lower()}"
        return "social_account", "telegram:unresolved"

    if host in SOCIAL_ACCOUNT_HOSTS:
        reserved = {
            "accounts",
            "explore",
            "groups",
            "hashtag",
            "home",
            "i",
            "intent",
            "marketplace",
            "p",
            "reel",
            "search",
            "share",
            "stories",
            "watch",
        }
        account = ""
        if segments and segments[0].lower() not in reserved:
            account = segments[0].lstrip("@").lower()
        if not account and query.get("id"):
            account = query["id"][0].lower()
        if not account:
            handle = re.search(
                r"(?<!\w)@([A-Za-z0-9_.-]{3,})",
                title + "\n" + text[:2_000],
            )
            if handle:
                account = handle.group(1).lower()
        if account:
            # x.com and twitter.com are two URL forms for the same platform;
            # the same handle must not count twice merely because both forms
            # appeared in search results.
            platform = "x-twitter" if domain in {"x.com", "twitter.com"} else domain
            return "social_account", f"{platform}:{account}"
        if allow_unresolved_social_post and segments:
            return "social_candidate", f"{domain}:{'/'.join(segments[:2])}"
        return "social_account", f"{domain}:unresolved"

    if host in {"blog.naver.com", "m.blog.naver.com"}:
        blog_id = ""
        if segments and segments[0].lower() != "postview.naver":
            blog_id = segments[0]
        elif query.get("blogId"):
            blog_id = query["blogId"][0]
        return "social_account", f"naver-blog:{blog_id or 'unresolved'}"

    if host.endswith("cafe.naver.com"):
        lowered_segments = [segment.lower() for segment in segments]
        if (
            len(lowered_segments) >= 3
            and lowered_segments[:2] == ["ca-fe", "cafes"]
        ):
            return "board", f"naver-cafe:{lowered_segments[2]}"
        club_id = str(
            (query.get("clubid") or query.get("clubId") or [""])[0]
        ).strip().lower()
        if club_id:
            return "board", f"naver-cafe:{club_id}"
        if lowered_segments and lowered_segments[0] not in {
            "articleread.nhn",
            "ca-fe",
        }:
            return "board", f"naver-cafe:{lowered_segments[0]}"
        return "board", "naver-cafe:unresolved"
    if host.endswith("cafe.daum.net"):
        lowered_segments = [segment.lower() for segment in segments]
        cafe_id = str(
            (query.get("grpid") or query.get("grpId") or [""])[0]
        ).strip().lower()
        if not cafe_id and lowered_segments:
            cafe_id = lowered_segments[0]
        return "board", f"daum-cafe:{cafe_id or 'unresolved'}"
    if domain in {"reddit.com", "reddit.co.kr"} and len(segments) >= 2:
        if segments[0].lower() == "r":
            return "board", f"reddit:{segments[1].lower()}"

    board_parameters = (
        "bo_table",
        "board_no",
        "board",
        "bbs_id",
        "mid",
    )
    for name in board_parameters:
        value = str((query.get(name) or [""])[0]).strip().lower()
        if value:
            return "board", f"{domain}:{name}={value}"

    if domain == "dcinside.com" and query.get("id"):
        return "board", f"dcinside:{query['id'][0].lower()}"
    if domain == "creativebox.kr" and segments:
        return "board", f"creativebox:{segments[0].lower()}"
    if domain == "i-boss.co.kr" and segments:
        match = re.match(r"ab-(\d+)-", segments[0], re.IGNORECASE)
        if match:
            return "board", f"i-boss:ab-{match.group(1)}"

    known_market_sections = {
        "hellomarket.com": "item",
        "itemmania.com": "buy",
        "joongna.com": "product",
        "kmong.com": "gig",
        "teamblind.com": "post",
        "z2u.com": "accounts",
    }
    market_section = known_market_sections.get(domain)
    if market_section and market_section in {segment.lower() for segment in segments}:
        return "board", f"{domain}:{market_section}"

    path_lower = parts.path.lower().rstrip("/") or "/"
    if "bmode=view" in parts.query.lower() and (
        "t=board" in parts.query.lower() or "idx=" in parts.query.lower()
    ):
        page_id = str((query.get("page_id") or [""])[0]).lower()
        return "board", f"{domain}:{path_lower}:{page_id}"
    iboard = re.search(r"/(?:bbs/)?board(?:\.php)?/([^/?#]+)", path_lower)
    if iboard:
        return "board", f"{domain}:board:{iboard.group(1)}"
    if "/bbs/" in path_lower:
        return "board", f"{domain}:{path_lower.rsplit('/', 1)[0]}"
    if "/gallery/" in path_lower:
        return "board", f"{domain}:gallery"
    if re.search(r"community[^/]*-view\.html$", path_lower):
        board_path = re.sub(r"-view\.html$", "", path_lower)
        return "board", f"{domain}:{board_path}"
    if re.search(r"/(?:article|community|forum|qna|review)/", path_lower):
        return "board", f"{domain}:{path_lower.rsplit('/', 1)[0]}"

    return "site", domain


def source_unit_token(key: bytes, descriptor: tuple[str, str]) -> str:
    kind, identity = descriptor
    digest = hmac.new(
        key,
        f"{kind}\n{identity}".encode("utf-8", "ignore"),
        hashlib.sha256,
    ).hexdigest()[:24]
    return f"{kind}-hmac:{digest}"


def effective_domain_record_limit(
    target: int,
    absolute_limit: int,
    maximum_share: float,
) -> int:
    """Resolve the strictest active per-domain limit for a target sample."""
    limits: list[int] = []
    if absolute_limit:
        limits.append(absolute_limit)
    if maximum_share:
        limits.append(max(1, math.floor(target * maximum_share)))
    return min(limits, default=0)


def effective_minimum_domains(
    target: int,
    domain_limit: int,
    requested_minimum: int,
) -> int:
    required_by_cap = math.ceil(target / domain_limit) if domain_limit else 0
    return max(requested_minimum, required_by_cap)


def collection_type_minimums(target: int, minimum_share: float) -> dict[str, int]:
    minimum = math.floor(target * minimum_share) if minimum_share else 0
    return {name: minimum for name in COLLECTION_TYPES if minimum}


def infer_collection_type(
    configured_type: str,
    title: str,
    text: str,
) -> str:
    """Map a collected post to one mutually exclusive sampling stratum."""
    combined = normalize_extracted_text(title + "\n" + text[:6_000]).lower()
    identity_document = re.search(
        r"여권|passport|신분증|주민등록증|민증|운전면허증|"
        r"외국인등록증|외국인\s*등록증",
        combined,
    )
    bank_account = re.search(
        r"통장|대포\s*통장|법인\s*통장|계좌|체크\s*카드|otp",
        combined,
        re.IGNORECASE,
    )
    personal_database = re.search(
        r"개인정보|개인\s*정보|고객\s*(?:명단|리스트|디비|db)|"
        r"(?:대출|보험|주식|코인|부동산|회원|연락처|유흥|렌탈|휴대폰|"
        r"통신|병원|성형|치과|맘카페|자동차|배달|맛집)\s*(?:디비|db)|"
        r"(?:디비|db)\s*(?:판매|구매|매입|삽니다|팝니다)",
        combined,
        re.IGNORECASE,
    )
    account_or_verification = re.search(
        r"아이디|id\s*(?:판매|구매|매입)|계정|가입\s*인증|본인\s*인증|"
        r"실명\s*인증|비실명|대포\s*폰|유심|문자\s*인증|"
        r"네이버|카카오|카톡|구글|인스타|페이스북|트위터|틱톡|쿠팡|배민",
        combined,
        re.IGNORECASE,
    )
    if identity_document:
        return "신분증·여권 위조/제작"
    if bank_account:
        return "통장·계좌"
    if personal_database:
        return "개인정보DB"
    if account_or_verification:
        return "계정·아이디·가입인증"
    return {
        "개인정보DB": "개인정보DB",
        "포털ID": "계정·아이디·가입인증",
        "여권 및 통장": "통장·계좌",
    }.get(configured_type, "기타")


def interleave_candidates_by_domain(
    candidates: Iterable[Candidate],
    max_per_domain: int = 0,
    max_per_source_unit: int = 0,
) -> list[Candidate]:
    """Round-robin domains and source units while bounding repeated posts."""
    buckets: dict[str, dict[tuple[str, str], deque[Candidate]]] = {}
    for candidate in candidates:
        domain = registrable_domain(urlsplit(candidate.url).hostname or "")
        unit = source_unit_descriptor(
            candidate.url,
            candidate.discovery_text,
            "",
            allow_unresolved_social_post=True,
        )
        domain_buckets = buckets.setdefault(domain, {})
        total_for_domain = sum(len(items) for items in domain_buckets.values())
        if max_per_domain and total_for_domain >= max_per_domain:
            continue
        unit_bucket = domain_buckets.setdefault(unit, deque())
        if max_per_source_unit and len(unit_bucket) >= max_per_source_unit:
            continue
        unit_bucket.append(candidate)
    ordered: list[Candidate] = []
    active = deque(buckets)
    active_units = {domain: deque(units) for domain, units in buckets.items()}
    while active:
        domain = active.popleft()
        units = active_units[domain]
        unit = units.popleft()
        unit_bucket = buckets[domain][unit]
        ordered.append(unit_bucket.popleft())
        if unit_bucket:
            units.append(unit)
        if units:
            active.append(domain)
    return ordered


def candidate_domain_count(candidates: Iterable[Candidate]) -> int:
    return len(
        {
            registrable_domain(urlsplit(candidate.url).hostname or "")
            for candidate in candidates
        }
    )


def candidate_source_unit_count(candidates: Iterable[Candidate]) -> int:
    return len(
        {
            source_unit_descriptor(
                candidate.url,
                candidate.discovery_text,
                "",
                allow_unresolved_social_post=True,
            )
            for candidate in candidates
        }
    )


def is_known_source_unit_candidate(
    url: str,
    discovery_text: str,
    known_source_units: set[tuple[str, str]],
) -> bool:
    """Return whether a search hit belongs to an already retained source.

    Search discovery must optimize for new boards/accounts/sites, not merely
    unseen post URLs.  Use the conservative final counting descriptor here:
    unresolved social post URLs stay in one platform-level bucket instead of
    each post being treated as a new account.
    """
    if not known_source_units:
        return False
    return (
        source_unit_descriptor(url, discovery_text, "") in known_source_units
    )


def exclude_known_source_unit_candidates(
    candidates: Iterable[Candidate],
    known_source_units: set[tuple[str, str]],
) -> list[Candidate]:
    return [
        candidate
        for candidate in candidates
        if not is_known_source_unit_candidate(
            candidate.url,
            candidate.discovery_text,
            known_source_units,
        )
    ]


def prioritize_candidates_by_type_deficit(
    candidates: Iterable[Candidate],
    current_counts: Counter[str],
    minimum_counts: dict[str, int],
) -> list[Candidate]:
    """Try strata still below their minimum before already abundant strata."""
    deficits = {
        name
        for name, minimum in minimum_counts.items()
        if current_counts[name] < minimum
    }
    return sorted(
        candidates,
        key=lambda candidate: infer_collection_type(
            candidate.detection_type,
            candidate.discovery_text,
            "",
        )
        not in deficits,
    )


def url_digest(key: bytes, url: str) -> str:
    return hmac.new(key, url.encode("utf-8", "ignore"), hashlib.sha256).hexdigest()


def is_public_http_url(url: str) -> tuple[bool, str]:
    try:
        parts = urlsplit(url)
        if parts.scheme not in {"http", "https"} or not parts.hostname:
            return False, "unsupported_url"
        port = parts.port or (443 if parts.scheme == "https" else 80)
        infos = socket.getaddrinfo(parts.hostname, port, type=socket.SOCK_STREAM)
        if not infos:
            return False, "dns_empty"
        for info in infos:
            address = ipaddress.ip_address(info[4][0])
            if not address.is_global:
                return False, "non_public_address"
        return True, ""
    except (OSError, ValueError):
        return False, "dns_or_url_error"


def connect_browser(cdp_address: str) -> webdriver.Chrome:
    options = Options()
    options.add_experimental_option("debuggerAddress", cdp_address)
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(25)
    driver.set_script_timeout(10)
    return driver


def disconnect_browser(driver: webdriver.Chrome) -> None:
    """Detach ChromeDriver without closing the externally managed CDP browser."""
    try:
        driver.service.stop()
    except (AttributeError, OSError):
        pass


def load_query_specs(path: Path) -> list[QuerySpec]:
    if not path.exists():
        raise FileNotFoundError(f"Private query file not found: {path}")
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    groups = raw.get("groups")
    if not isinstance(groups, list) or not groups:
        raise ValueError("Query YAML must contain a non-empty 'groups' list")
    specs: list[QuerySpec] = []
    names: set[str] = set()
    for item in groups:
        if not isinstance(item, dict):
            raise TypeError("Each query group must be a mapping")
        name = str(item.get("name", "")).strip()
        detection_type = str(item.get("detection_type", "")).strip()
        queries = item.get("queries")
        if not name or not re.fullmatch(r"[a-z0-9_-]+", name):
            raise ValueError(
                "Query group names may contain only lowercase letters, digits, '_' and '-'"
            )
        if name in names:
            raise ValueError(f"Duplicate query group: {name}")
        if detection_type not in DETECTION_TYPES:
            raise ValueError(
                f"Unsupported detection type in group {name}: {detection_type}"
            )
        if not isinstance(queries, list) or not queries:
            raise ValueError(f"Query group {name} must contain at least one query")
        names.add(name)
        for query in queries:
            query = str(query).strip()
            if not query:
                raise ValueError(f"Query group {name} contains an empty query")
            if len(query) > 200:
                raise ValueError(f"Query in group {name} exceeds 200 characters")
            specs.append(QuerySpec(name, detection_type, query))
    return specs


def expand_query_specs(specs: Iterable[QuerySpec], variants: int) -> list[QuerySpec]:
    suffixes = (
        "",
        " 텔레그램",
        " 오픈채팅",
        " 문의",
        " 게시판",
        " 블로그",
        " 연락처",
        " 판매",
        " 구매",
        " 실시간",
    )
    expanded: list[QuerySpec] = []
    seen: set[tuple[str, str]] = set()
    for spec in specs:
        for suffix in suffixes[:variants]:
            query = (spec.query + suffix).strip()
            key = (spec.group, query)
            if key in seen or len(query) > 200:
                continue
            seen.add(key)
            expanded.append(QuerySpec(spec.group, spec.detection_type, query))
    return expanded


SEARCH_NEGATIVE_FILTERS = (
    "-개인정보처리방침 -이용약관 -위키 -로그인 -회원가입 -고객센터 "
    "-뉴스 -기사 -보도 -사건 -판결 -처벌 -법률상담 -예방 -주의 -경고"
)


def constrain_query_specs(specs: Iterable[QuerySpec]) -> list[QuerySpec]:
    """Favor exact illicit-trade phrases and suppress known generic documents."""
    constrained: list[QuerySpec] = []
    seen: set[tuple[str, str]] = set()
    for spec in specs:
        words = spec.query.split()
        candidates = [
            f"{spec.query} {SEARCH_NEGATIVE_FILTERS}",
            f'"{spec.query}" {SEARCH_NEGATIVE_FILTERS}',
        ]
        if len(words) >= 3:
            candidates.extend(
                (
                    f'"{" ".join(words[:2])}" {" ".join(words[2:])} '
                    f"{SEARCH_NEGATIVE_FILTERS}",
                    f'{" ".join(words[:-2])} "{" ".join(words[-2:])}" '
                    f"{SEARCH_NEGATIVE_FILTERS}",
                )
            )
        for query in candidates:
            key = (spec.group, query)
            if key in seen or len(query) > 400:
                continue
            seen.add(key)
            constrained.append(QuerySpec(spec.group, spec.detection_type, query))
    return constrained


def strip_negative_search_terms(query: str) -> str:
    """Remove web-search minus filters for vertical search surfaces."""
    return re.sub(r"\s+-\S+", "", query).strip()


def load_seed_candidates(path: Path) -> list[Candidate]:
    if not path.exists():
        raise FileNotFoundError(f"Private seed file not found: {path}")
    if path.suffix.lower() == ".jsonl":
        return load_candidate_queue(path)
    candidates: dict[str, Candidate] = {}
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames or not (
                {"url", "raw_url", "source_url"} & set(reader.fieldnames)
            ):
                raise ValueError(
                    "Seed CSV requires a 'url', 'raw_url', or 'source_url' column"
                )
            for row in reader:
                url = canonicalize_url(
                    str(
                        row.get("url")
                        or row.get("raw_url")
                        or row.get("source_url")
                        or ""
                    ).strip()
                )
                detection_type = str(row.get("detection_type") or "기타").strip()
                group = str(row.get("query_group") or "private_seed").strip()
                if not url:
                    continue
                if detection_type not in DETECTION_TYPES:
                    raise ValueError(
                        f"Unsupported seed detection type: {detection_type}"
                    )
                candidates.setdefault(
                    url,
                    Candidate(
                        url,
                        group,
                        detection_type,
                        source_type="seed",
                        discovery_text=normalize_extracted_text(
                            str(row.get("title") or row.get("masked_title") or "")
                            + "\n"
                            + str(row.get("text") or row.get("masked_text") or "")
                        )[:2_000],
                        search_provider=str(row.get("search_provider") or ""),
                    ),
                )
    else:
        for line in path.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            url = canonicalize_url(stripped)
            if url:
                candidates.setdefault(
                    url,
                    Candidate(url, "private_seed", "기타", source_type="seed"),
                )
    if not candidates:
        raise ValueError("Seed file did not contain any usable public HTTP(S) URLs")
    return list(candidates.values())


def load_excluded_urls(paths: Iterable[Path]) -> set[str]:
    """Load canonical source URLs from prior sample CSVs for holdout runs."""
    excluded: set[str] = set()
    for path in paths:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fields = set(reader.fieldnames or ())
            url_field = next(
                (name for name in ("source_url", "raw_url", "url") if name in fields),
                "",
            )
            if not url_field:
                raise ValueError(
                    f"Excluded CSV requires source_url, raw_url, or url: {path}"
                )
            for row in reader:
                url = canonicalize_url(str(row.get(url_field) or "").strip())
                if url:
                    excluded.add(url)
    return excluded


def load_excluded_fingerprints(paths: Iterable[Path]) -> set[str]:
    """Load document fingerprints from prior samples when the column exists."""
    excluded: set[str] = set()
    for path in paths:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fields = set(reader.fieldnames or ())
            fingerprint_field = next(
                (
                    name
                    for name in (
                        "near_duplicate_fingerprint",
                        "near_duplicate_cluster",
                    )
                    if name in fields
                ),
                "",
            )
            if not fingerprint_field:
                continue
            for row in reader:
                fingerprint = str(row.get(fingerprint_field) or "").strip()
                if fingerprint:
                    excluded.add(fingerprint)
    return excluded


def save_candidate_queue(path: Path, candidates: Iterable[Candidate]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    os.chmod(path.parent, 0o700)
    with path.open("w", encoding="utf-8") as handle:
        for candidate in candidates:
            handle.write(json.dumps(asdict(candidate), ensure_ascii=False) + "\n")
    os.chmod(path, 0o600)


def load_candidate_queue(path: Path) -> list[Candidate]:
    candidates: dict[str, Candidate] = {}
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                item = json.loads(line)
                candidate = Candidate(
                    url=str(item["url"]),
                    query_group=str(item["query_group"]),
                    detection_type=str(item["detection_type"]),
                    source_type=str(item.get("source_type") or "search"),
                    discovery_text=str(item.get("discovery_text") or ""),
                    search_provider=str(item.get("search_provider") or ""),
                )
            except (KeyError, TypeError, json.JSONDecodeError) as exc:
                raise ValueError(
                    f"Invalid candidate queue row {line_number}: {path}"
                ) from exc
            url = canonicalize_url(candidate.url)
            if url:
                candidate.url = url
                candidates.setdefault(url, candidate)
    if not candidates:
        raise ValueError(f"Candidate queue is empty: {path}")
    return list(candidates.values())


def ordered_provider_names(
    available: Iterable[str], requested: Iterable[str] | None
) -> list[str]:
    """Keep CLI provider order while removing duplicates and unknown names."""
    available_names = list(available)
    if requested is None:
        return available_names
    allowed = set(available_names)
    ordered: list[str] = []
    for name in requested:
        if name in allowed and name not in ordered:
            ordered.append(name)
    return ordered


def discover_google_api_candidates(
    session: requests.Session,
    query_specs: list[QuerySpec],
    desired: int,
    pages: int,
    delay: float,
    prefilter_mode: str,
    api_key: str,
    cse_id: str,
    soft_target_multiplier: int = 3,
    minimum_domains: int = 0,
    minimum_source_units: int = 0,
    max_candidates_per_domain: int = 0,
    known_source_units: set[tuple[str, str]] | None = None,
    page_offset: int = 0,
) -> list[Candidate]:
    """Discover candidates through Google's official JSON API without key logging."""
    found: dict[str, Candidate] = {}
    found_domain_counts: Counter[str] = Counter()
    soft_target = max(desired * soft_target_multiplier, desired + 30)
    for spec in query_specs:
        for page in range(page_offset, min(page_offset + pages, 10)):
            try:
                response = session.get(
                    "https://customsearch.googleapis.com/customsearch/v1",
                    params={
                        "key": api_key,
                        "cx": cse_id,
                        "q": spec.query,
                        "start": page * 10 + 1,
                        "num": 10,
                        "hl": "ko",
                        "filter": "0",
                        "fields": "items(link,title,snippet)",
                    },
                    timeout=(CONNECT_TIMEOUT_SECONDS, READ_TIMEOUT_SECONDS),
                )
            except requests.RequestException:
                print(
                    "google_api: request failed; stopping without retry bypass",
                    flush=True,
                )
                return [
                    item
                    for item in found.values()
                    if discovery_candidate_passes(item, prefilter_mode)
                ]
            try:
                if response.status_code in {401, 403, 429}:
                    print(
                        f"google_api: HTTP {response.status_code}; "
                        "check credentials, quota, and engine scope",
                        flush=True,
                    )
                    return [
                        item
                        for item in found.values()
                        if discovery_candidate_passes(item, prefilter_mode)
                    ]
                if response.status_code != 200:
                    print(
                        f"google_api: HTTP {response.status_code}; "
                        "stopping provider",
                        flush=True,
                    )
                    return [
                        item
                        for item in found.values()
                        if discovery_candidate_passes(item, prefilter_mode)
                    ]
                try:
                    payload = response.json()
                except requests.exceptions.JSONDecodeError:
                    print("google_api: invalid JSON response", flush=True)
                    return [
                        item
                        for item in found.values()
                        if discovery_candidate_passes(item, prefilter_mode)
                    ]
            finally:
                response.close()

            before_page = len(found)
            for item in payload.get("items") or []:
                url = canonicalize_url(str(item.get("link") or ""))
                if not url:
                    continue
                host = (urlsplit(url).hostname or "").lower()
                if host in SEARCH_HOSTS or host.endswith((".google.com", ".bing.com")):
                    continue
                discovery_text = normalize_extracted_text(
                    str(item.get("title") or "")
                    + "\n"
                    + str(item.get("snippet") or "")
                )[:2_000]
                if not discovery_text:
                    continue
                if is_known_source_unit_candidate(
                    url,
                    discovery_text,
                    known_source_units or set(),
                ):
                    continue
                domain = registrable_domain(urlsplit(url).hostname or "")
                if (
                    url not in found
                    and max_candidates_per_domain
                    and found_domain_counts[domain] >= max_candidates_per_domain
                ):
                    continue
                if url not in found:
                    found[url] = Candidate(
                        url=url,
                        query_group=spec.group,
                        detection_type=spec.detection_type,
                        discovery_text=discovery_text,
                        search_provider="google_api",
                    )
                    found_domain_counts[domain] += 1
                elif discovery_text not in found[url].discovery_text:
                    found[url].discovery_text = normalize_extracted_text(
                        found[url].discovery_text + "\n" + discovery_text
                    )[:2_000]
            qualified = [
                item
                for item in found.values()
                if discovery_candidate_passes(item, prefilter_mode)
            ]
            print(
                f"google_api {spec.group}: {len(found)} discovered, "
                f"{len(qualified)} qualified (+{len(found) - before_page})",
                flush=True,
            )
            if (
                len(qualified) >= soft_target
                and candidate_domain_count(qualified) >= minimum_domains
                and candidate_source_unit_count(qualified) >= minimum_source_units
            ):
                return qualified
            if len(found) == before_page:
                break
            time.sleep(delay)
    return [
        item
        for item in found.values()
        if discovery_candidate_passes(item, prefilter_mode)
    ]


def discover_candidates(
    driver: webdriver.Chrome,
    query_specs: list[QuerySpec],
    desired: int,
    pages: int,
    delay: float,
    soft_target_multiplier: int = 3,
    prefilter_mode: str = "off",
    providers_enabled: list[str] | None = None,
    provider_stale_pages_limit: int = 12,
    minimum_domains: int = 0,
    minimum_source_units: int = 0,
    max_candidates_per_domain: int = 0,
    known_source_units: set[tuple[str, str]] | None = None,
    page_offset: int = 0,
) -> list[Candidate]:
    found: dict[str, Candidate] = {}
    found_domain_counts: Counter[str] = Counter()
    broad_queries = [item for item in query_specs if '"' not in item.query]
    phrase_queries = [item for item in query_specs if '"' in item.query]
    # Balance research groups without allowing narrow phrase variants to trigger
    # the provider's stale-result cutoff before broader queries are attempted.
    rng = random.Random(20260817)
    rng.shuffle(broad_queries)
    rng.shuffle(phrase_queries)
    soft_target = max(desired * soft_target_multiplier, desired + 30)

    providers = (
        (
            "naver",
            lambda query, page: (
                "https://search.naver.com/search.naver?where=web&start="
                f"{page * 15 + 1}&query={quote_plus(query)}"
            ),
            ".fds-web-doc-root a[href], "
            ".fds-ugc-single-intention-item-list-rra a[href]",
        ),
        (
            "naver_blog",
            lambda query, page: (
                "https://search.naver.com/search.naver?ssc=tab.blog.all&"
                "where=blog&sm=tab_jum&start="
                f"{page * 7 + 1}&query={quote_plus(strip_negative_search_terms(query))}"
            ),
            "._fe_view_root a[href*='blog.naver.com/']",
        ),
        (
            "naver_cafe",
            lambda query, page: (
                "https://search.naver.com/search.naver?ssc=tab.cafe.all&"
                "where=cafe&sm=tab_jum&start="
                f"{page * 7 + 1}&query={quote_plus(strip_negative_search_terms(query))}"
            ),
            "li.bx._bx a.title_link[href*='cafe.naver.com/']",
        ),
        (
            "naver_kin",
            lambda query, page: (
                "https://search.naver.com/search.naver?ssc=tab.kin.all&"
                "where=kin&sm=tab_jum&start="
                f"{page * 10 + 1}&query={quote_plus(strip_negative_search_terms(query))}"
            ),
            "a[href]",
        ),
        (
            "naver_news",
            lambda query, page: (
                "https://search.naver.com/search.naver?ssc=tab.news.all&"
                "where=news&sm=tab_jum&start="
                f"{page * 10 + 1}&query={quote_plus(strip_negative_search_terms(query))}"
            ),
            ".fds-news-item-list-tab a[href]",
        ),
        (
            "daum",
            lambda query, page: (
                "https://search.daum.net/search?w="
                + ("tot" if page == 0 else "fusion")
                + f"&q={quote_plus(query)}"
                + ("" if page == 0 else f"&p={page + 1}&DA=PGD")
            ),
            "#twaColl c-card a[href], #twcColl c-card a[href]",
        ),
        (
            "daum_blog",
            lambda query, page: (
                "https://search.daum.net/search?w=fusion&col=blog&"
                f"q={quote_plus(query)}&p={page + 1}"
            ),
            "#twcColl c-card a[href]",
        ),
        (
            "bing",
            lambda query, page: (
                "https://www.bing.com/search?setlang=ko-kr&count=20&first="
                f"{page * 20 + 1}&q={quote_plus(query)}"
            ),
            "li.b_algo h2 a",
        ),
        (
            "duckduckgo",
            lambda query, page: (
                "https://html.duckduckgo.com/html/?kl=kr-kr&s="
                f"{page * 30}&q={quote_plus(query)}"
            ),
            "a.result__a",
        ),
        (
            "google",
            lambda query, page: (
                "https://www.google.com/search?filter=0&num=20&hl=ko&start="
                f"{page * 20}&q={quote_plus(query)}"
            ),
            "a:has(h3)",
        ),
    )
    provider_by_name = {item[0]: item for item in providers}
    selected_provider_names = ordered_provider_names(
        provider_by_name, providers_enabled
    )
    selected_providers = tuple(
        provider_by_name[name] for name in selected_provider_names
    )

    def rotate(items: list[QuerySpec], provider_index: int) -> list[QuerySpec]:
        if not items:
            return []
        provider_count = max(1, len(selected_providers))
        offset = (provider_index * max(1, len(items) // provider_count)) % len(items)
        return items[offset:] + items[:offset]

    for provider_index, (provider_name, make_url, selector) in enumerate(
        selected_providers
    ):
        provider_query_items = rotate(broad_queries, provider_index) + rotate(
            phrase_queries, provider_index
        )
        provider_blocked = False
        provider_stale_pages = 0
        provider_navigation_errors = 0
        for spec in provider_query_items:
            stale_pages = 0
            for page in range(page_offset, page_offset + pages):
                before_page = len(found)
                try:
                    driver.get(make_url(spec.query, page))
                    provider_navigation_errors = 0
                    # Daum hydrates integrated-web result cards after the main
                    # document load event. Reading the DOM immediately sees an
                    # empty container even though public results appear moments
                    # later in the same page.
                    if provider_name.startswith("daum"):
                        time.sleep(max(0.75, min(delay, 2.0)))
                        for _ in range(8):
                            ready = driver.execute_script(
                                "return Boolean(document.querySelector("
                                "'#twaColl c-card a[href], "
                                "#twcColl c-card a[href]'));"
                            )
                            if ready:
                                break
                            time.sleep(0.25)
                except TimeoutException:
                    try:
                        driver.execute_cdp_cmd("Page.stopLoading", {})
                    except WebDriverException:
                        pass
                except WebDriverException:
                    provider_navigation_errors += 1
                    if (
                        provider_navigation_errors
                        >= MAX_PROVIDER_NAVIGATION_ERRORS
                    ):
                        print(
                            f"{provider_name}: "
                            f"{MAX_PROVIDER_NAVIGATION_ERRORS} consecutive "
                            "navigation errors; switching provider",
                            flush=True,
                        )
                        provider_blocked = True
                        break
                    time.sleep(delay)
                    continue
                try:
                    page_lower = driver.execute_script(
                        "return document.body ? "
                        "document.body.innerText.toLowerCase() : '';"
                    )
                except (TimeoutException, WebDriverException):
                    print(
                        f"{provider_name}: result page unavailable; "
                        "switching provider without retry bypass",
                        flush=True,
                    )
                    provider_blocked = True
                    break
                challenge_markers = (
                    "captcha",
                    "unusual traffic",
                    "verify you are human",
                    "our systems have detected unusual traffic",
                )
                if "sorry" in driver.current_url or any(
                    marker in page_lower for marker in challenge_markers
                ):
                    print(
                        f"{provider_name}: challenge detected; switching provider without bypass",
                        flush=True,
                    )
                    provider_blocked = True
                    break
                try:
                    anchors = driver.execute_script(
                        "return Array.from(document.querySelectorAll(arguments[0]))"
                        ".map(a => ({href:a.href, "
                        "text:(a.innerText || a.textContent || '').trim(), "
                        "context:((a.closest('li.b_algo, li.bx._bx, .result, "
                        ".fds-web-doc-root, .fds-ugc-single-intention-item-list-rra, "
                        ".fds-ugc-single-intention-item-list-tab, ._fe_view_root, "
                        ".MjjYud, .g, c-card') || a).innerText || '').trim(), "
                        "ignored:Boolean(a.closest('.sds-comps-profile-source, "
                        ".api_ly_save'))}));",
                        selector,
                    )
                except (TimeoutException, WebDriverException):
                    print(
                        f"{provider_name}: result links unavailable; "
                        "switching provider without retry bypass",
                        flush=True,
                    )
                    provider_blocked = True
                    break
                for anchor in anchors or []:
                    if anchor.get("ignored"):
                        continue
                    raw = str(anchor.get("href") or "")
                    discovery_text = normalize_extracted_text(
                        str(anchor.get("context") or anchor.get("text") or "")
                    )[:2_000]
                    if not discovery_text or discovery_text in {
                        "새 창 열림",
                        "Keep에 저장",
                        "Keep에 바로가기새 창 열림",
                    }:
                        continue
                    raw = unwrap_search_result_url(raw)
                    url = canonicalize_url(raw)
                    if not url:
                        continue
                    host = (urlsplit(url).hostname or "").lower()
                    if host in SEARCH_HOSTS or host.endswith(
                        (".google.com", ".bing.com")
                    ):
                        continue
                    if is_known_source_unit_candidate(
                        url,
                        discovery_text,
                        known_source_units or set(),
                    ):
                        continue
                    domain = registrable_domain(urlsplit(url).hostname or "")
                    if (
                        url not in found
                        and max_candidates_per_domain
                        and found_domain_counts[domain]
                        >= max_candidates_per_domain
                    ):
                        continue
                    if url not in found:
                        found[url] = Candidate(
                            url=url,
                            query_group=spec.group,
                            detection_type=spec.detection_type,
                            discovery_text=discovery_text,
                            search_provider=provider_name,
                        )
                        found_domain_counts[domain] += 1
                    elif discovery_text and discovery_text not in found[url].discovery_text:
                        found[url].discovery_text = normalize_extracted_text(
                            found[url].discovery_text + "\n" + discovery_text
                        )[:2_000]
                qualified = (
                    [
                        item
                        for item in found.values()
                        if discovery_candidate_passes(item, prefilter_mode)
                    ]
                    if prefilter_mode != "off"
                    else list(found.values())
                )
                print(
                    f"{provider_name} {spec.group}: {len(found)} discovered, "
                    f"{len(qualified)} qualified "
                    f"(+{len(found) - before_page})",
                    flush=True,
                )
                if (
                    len(qualified) >= soft_target
                    and candidate_domain_count(qualified) >= minimum_domains
                    and candidate_source_unit_count(qualified)
                    >= minimum_source_units
                ):
                    return qualified
                if len(found) == before_page:
                    provider_stale_pages += 1
                    if (
                        provider_stale_pages_limit
                        and provider_stale_pages >= provider_stale_pages_limit
                    ):
                        print(
                            f"{provider_name}: {provider_stale_pages_limit} pages "
                            "without a new "
                            "candidate; switching provider",
                            flush=True,
                        )
                        provider_blocked = True
                        break
                else:
                    provider_stale_pages = 0
                if len(found) == before_page:
                    stale_pages += 1
                    if stale_pages >= 2:
                        break
                else:
                    stale_pages = 0
                time.sleep(delay)
            if provider_blocked:
                break
    if prefilter_mode != "off":
        return [
            item
            for item in found.values()
            if discovery_candidate_passes(item, prefilter_mode)
        ]
    return list(found.values())


def request_once(
    session: requests.Session,
    url: str,
    limiter: RateLimiter,
    max_redirects: int = 5,
) -> tuple[requests.Response | None, str, str]:
    current = url
    for _ in range(max_redirects + 1):
        safe, reason = is_public_http_url(current)
        if not safe:
            return None, current, reason
        host = urlsplit(current).hostname or ""
        limiter.wait(host)
        try:
            response = session.get(
                current,
                allow_redirects=False,
                timeout=(CONNECT_TIMEOUT_SECONDS, READ_TIMEOUT_SECONDS),
                stream=True,
                headers={"Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.1"},
            )
        except requests.RequestException as exc:
            return None, current, type(exc).__name__
        if response.is_redirect or response.is_permanent_redirect:
            location = response.headers.get("Location")
            response.close()
            if not location:
                return None, current, "redirect_without_location"
            next_url = canonicalize_url(urljoin(current, location))
            if not next_url:
                return None, current, "unsafe_redirect"
            current = next_url
            continue
        return response, current, ""
    return None, current, "too_many_redirects"


def robots_allowed(
    session: requests.Session,
    url: str,
    limiter: RateLimiter,
    cache: dict[str, tuple[bool, str]],
) -> tuple[bool, str]:
    parts = urlsplit(url)
    origin = f"{parts.scheme}://{parts.netloc}"
    if origin in cache:
        return cache[origin]
    robots_url = origin + "/robots.txt"
    response, _, reason = request_once(session, robots_url, limiter, max_redirects=2)
    if response is None:
        # A network failure is not interpreted as an explicit robots denial.
        result = (True, "robots_unavailable:" + reason)
    elif response.status_code in {401, 403}:
        response.close()
        result = (False, "robots_forbidden")
    elif response.status_code >= 400:
        response.close()
        result = (True, "robots_absent")
    else:
        raw = response.raw.read(256_000, decode_content=True).decode("utf-8", "replace")
        response.close()
        from urllib.robotparser import RobotFileParser

        parser = RobotFileParser()
        parser.set_url(robots_url)
        parser.parse(raw.splitlines())
        allowed = parser.can_fetch(USER_AGENT, url)
        result = (allowed, "robots_allowed" if allowed else "robots_disallowed")
    cache[origin] = result
    return result


def read_html(response: requests.Response) -> tuple[str | None, str]:
    content_type = response.headers.get("Content-Type", "").lower()
    if "text/html" not in content_type and "application/xhtml+xml" not in content_type:
        return None, "non_html_content"
    declared = response.headers.get("Content-Length")
    if declared and declared.isdigit() and int(declared) > MAX_HTML_BYTES:
        return None, "content_too_large"
    chunks: list[bytes] = []
    total = 0
    for chunk in response.iter_content(64 * 1024):
        total += len(chunk)
        if total > MAX_HTML_BYTES:
            return None, "content_too_large"
        chunks.append(chunk)
    response.encoding = response.encoding or response.apparent_encoding or "utf-8"
    try:
        return b"".join(chunks).decode(response.encoding, "replace"), ""
    except LookupError:
        return b"".join(chunks).decode("utf-8", "replace"), ""


def normalize_extracted_text(value: str | None) -> str:
    text = re.sub(r"[ \t]+", " ", value or "")
    text = re.sub(r"\n[ \t]+", "\n", text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()[:MAX_TEXT_CHARS]


def extract_title_text(html: str, url: str) -> tuple[str, str, str]:
    soup = BeautifulSoup(html, "lxml")
    title = ""
    if soup.title and soup.title.string:
        title = soup.title.string.strip()

    # Korean legacy boards often keep the post in a table cell while generic
    # article extraction selects a longer footer. Prefer only selectors that
    # strongly identify an individual post before trying trafilatura.
    strong_post_candidates = [
        normalize_extracted_text(node.get_text("\n", strip=True))
        for selector in (
            ".post-content",
            ".article-content",
            ".article_view",
            ".board_view",
            ".view_content",
            ".view_text",
            ".view_cont",
            ".board_contents",
            ".board-content",
            ".bo_v_con",
            "#bo_v_con",
            "td.con_f",
        )
        for node in soup.select(selector)
    ]
    strong_post_text = max(strong_post_candidates, key=len, default="")
    if len(strong_post_text) >= DEFAULT_MIN_TEXT_CHARS:
        return title[:2_000], strong_post_text, "strong_post_container"

    complex_dom = len(html) > 600_000 or len(soup.find_all(True, limit=5_001)) > 5_000
    precision_text = None
    if not complex_dom:
        precision_text = trafilatura.extract(
            html,
            url=url,
            include_comments=False,
            include_tables=False,
            include_links=False,
            favor_precision=True,
            output_format="txt",
        )
    precision_text = normalize_extracted_text(precision_text)
    if len(precision_text) >= DEFAULT_MIN_TEXT_CHARS:
        return title[:2_000], precision_text, "trafilatura_precision"

    # Short forum posts are often missed by article-focused extraction. Prefer a
    # visible main/article container before falling back to a broader recall pass.
    for node in soup(["script", "style", "noscript", "svg", "nav", "footer", "header"]):
        node.decompose()
    main_candidates = [
        normalize_extracted_text(node.get_text("\n", strip=True))
        for selector in (
            "article",
            "main",
            "[role='main']",
            ".post-content",
            ".post_view",
            ".article-content",
            ".article_view",
            ".board_view",
            "#content",
        )
        for node in soup.select(selector)
    ]
    main_text = max(main_candidates, key=len, default="")
    if len(main_text) >= DEFAULT_MIN_TEXT_CHARS:
        return title[:2_000], main_text, "main_container"

    recall_text = None
    if not complex_dom:
        recall_text = trafilatura.extract(
            html,
            url=url,
            include_comments=False,
            include_tables=False,
            include_links=False,
            favor_recall=True,
            output_format="txt",
        )
    recall_text = normalize_extracted_text(recall_text)
    if len(recall_text) >= DEFAULT_MIN_TEXT_CHARS:
        return title[:2_000], recall_text, "trafilatura_recall"

    body_text = normalize_extracted_text(soup.get_text("\n", strip=True))
    candidates = (
        (precision_text, "trafilatura_precision_short"),
        (main_text, "main_container_short"),
        (recall_text, "trafilatura_recall_short"),
        (body_text, "visible_body_fallback"),
    )
    text, method = max(candidates, key=lambda item: len(item[0]))
    return title[:2_000], text, method


def text_quality_reason(
    text: str, minimum_chars: int, minimum_korean_chars: int = 0, title: str = ""
) -> str:
    if len(text) < minimum_chars:
        return "insufficient_text"
    lowered = text.lower()
    challenge_markers = (
        "verify you are human",
        "checking your browser",
        "enable javascript and cookies to continue",
        "access denied",
        "비정상적인 접근",
        "자동입력 방지",
    )
    if any(marker in lowered for marker in challenge_markers):
        return "challenge_or_access_page"
    tokens = re.findall(r"[가-힣A-Za-z0-9_]{2,}", text)
    if len(tokens) < max(6, minimum_chars // 20):
        return "insufficient_meaningful_tokens"
    if minimum_korean_chars:
        korean_chars = len(re.findall(r"[가-힣]", title + "\n" + text))
        if korean_chars < minimum_korean_chars:
            return "insufficient_korean_text"
    return ""


RELEVANCE_TARGET = re.compile(
    r"(?:고객|회원|보험|대출|주식|코인|부동산|업체|사업자|마케팅|쇼핑몰|"
    r"성인|토토|유흥|렌탈|휴대폰|통신|병원|성형|치과|맘카페|자동차|배달|맛집)\s*"
    r"(?:DB|디비|명단|리스트)|(?:고객|회원)\s*정보|"
    r"(?:개인정보|연락처|전화번호|휴대폰번호|주민등록번호|주민번호|여권|통장|계좌|"
    r"신분증|주민등록증|운전면허증|면허증|외국인등록증)"
    r"(?:\s*(?:DB|디비|명단|리스트))?|"
    r"(?:네이버|다음|카카오|구글|쿠팡|배민|밴드|인스타|페이스북|포털)\s*"
    r"(?:계정|아이디|ID)|"
    r"(?:계정|아이디)\s*(?:판매|팝니다|매입|삽니다|거래)",
    re.IGNORECASE,
)
LABELING_TARGET = re.compile(
    RELEVANCE_TARGET.pattern
    + r"|(?<![A-Za-z0-9])DB(?![A-Za-z0-9])|디비|계정|아이디|"
    r"본인\s*인증|실명\s*인증|비실명|명의|대포\s*통장|리딩방",
    re.IGNORECASE,
)
RELEVANCE_SHORT_TARGET = re.compile(
    r"(?<![가-힣A-Za-z0-9])(?:"
    r"[가-힣]{0,10}(?:디비|아이디|계정)|"
    r"[가-힣]{0,10}(?:DB|ID)|"
    r"[가-힣]{0,10}(?:여권|통장|계좌|신분증|주민등록증|운전면허증|면허증|외국인등록증)"
    r")(?![가-힣A-Za-z0-9])",
    re.IGNORECASE,
)
RELEVANCE_STRICT_TARGET = re.compile(
    r"(?:고객|회원|보험|대출|주식|코인|부동산|업체|사업자|마케팅|쇼핑몰|"
    r"성인|토토|유흥|렌탈|휴대폰|통신|병원|성형|치과|맘카페|자동차|배달|맛집)\s*"
    r"(?:DB|디비|명단|리스트)|(?:고객|회원)\s*정보|"
    r"(?:개인정보|연락처|전화번호|휴대폰번호|휴대전화\s*번호|이메일|"
    r"주민등록번호|주민번호)\s*"
    r"(?:데이터(?:베이스)?|DB|디비|명단|목록|리스트|판매|팝니다|매입|"
    r"삽니다|거래|제공)|"
    r"(?:여권|통장|계좌|신분증|주민등록증|운전면허증|면허증|외국인등록증)|"
    r"(?:네이버|다음|카카오|구글|쿠팡|배민|밴드|인스타|인스타그램|페이스북|"
    r"트위터|엑스|틱톡|포털)\s*(?:계정|아이디|ID)|"
    r"(?:카카오톡|카톡|텔레그램|텔그|텔레)\s*(?:계정|아이디|ID)|"
    r"(?:대량|다중|실명|비실명|가입|본인|마케팅|광고|디엠|육성|신규)\s*"
    r"(?:계정|아이디|ID)|"
    r"(?:계정|아이디|ID).{0,15}"
    r"(?:대량|다중|실명|비실명|여러|개당|명의|마케팅|광고|디엠)|"
    r"(?:본인|실명|가입)\s*인증(?:\s*(?:계정|아이디|자료))?\s*"
    r"(?:판매|팝니다|매입|삽니다|거래)|"
    r"명의\s*(?:판매|팝니다|매입|삽니다|대여|거래)",
    re.IGNORECASE,
)
RELEVANCE_STRICT_SHORT_TARGET = re.compile(
    r"(?<![가-힣A-Za-z0-9])(?:"
    r"(?:디비|DB)|"
    r"(?:여권|통장|계좌|신분증|주민등록증|운전면허증|면허증|외국인등록증)"
    r")(?![가-힣A-Za-z0-9])",
    re.IGNORECASE,
)
RELEVANCE_TRADE = re.compile(
    r"판매|팝니다|매입|삽니다|구매|대량|건당|단가|거래|공급|보유|"
    r"위조|복제|제작|도용|"
    r"최신\s*(?:DB|디비|명단)",
    re.IGNORECASE,
)
RELEVANCE_TITLE_TRADE = re.compile(
    r"판매|팝니다|매입|삽니다|구매|대량|건당|단가|공급|보유|"
    r"위조|복제|제작|도용|"
    r"최신\s*(?:DB|디비|명단)",
    re.IGNORECASE,
)
EXPANSION_CONTACT_TERM = re.compile(
    r"(?<![가-힣A-Za-z0-9])(?:텔레그램|텔그|텔레|telegram|TG|"
    r"오픈채팅|오픈톡|카카오톡|카톡)(?![가-힣A-Za-z0-9])",
    re.IGNORECASE,
)
RELEVANCE_CONTACT = re.compile(
    r"\[(?:EMAIL|PHONE|MESSENGER_ID|ACCOUNT|URL|CONTACT_URL)\]|"
    r"https?://|www\.|(?<!\w)@[A-Za-z0-9_]{3,}|"
    + EXPANSION_CONTACT_TERM.pattern
    + r"|문의\s*[:：]?",
    re.IGNORECASE,
)
RELEVANCE_DISCOVERY_CONTACT = re.compile(
    r"\[(?:EMAIL|PHONE|MESSENGER_ID|ACCOUNT)\]|"
    r"(?<!\w)@[A-Za-z0-9_]{3,}|"
    + EXPANSION_CONTACT_TERM.pattern
    + r"|(?:카톡|텔레그램|텔그|텔레|오픈채팅)\s*(?:문의|연락|아이디|ID)|"
    r"문의\s*(?:주세요|바랍니다|가능|[:：])",
    re.IGNORECASE,
)
RELEVANCE_STRONG_CONTACT = re.compile(
    r"\[(?:EMAIL|PHONE|MESSENGER_ID|ACCOUNT)\]|"
    r"(?<!\w)@[A-Za-z0-9_]{3,}|"
    r"(?:텔레그램|telegram|텔그|텔레|카카오톡|카톡|오픈채팅|라인|line)\s*"
    r"(?:(?:아이디|id|주소|문의|연락|[:：])\s*)?[@:]?\s*"
    r"(?!계정(?:\s|$)|데이터(?:\s|$)|채널(?:\s|$)|그룹(?:\s|$)|봇(?:\s|$))"
    r"[A-Za-z0-9_.-]{3,}",
    re.IGNORECASE,
)
RELEVANCE_DIRECT_OFFER = re.compile(
    r"판매\s*(?:합니다|해요|중|가능)|팝니다|매입\s*(?:합니다|해요|중|가능)|"
    r"삽니다|구매\s*(?:합니다|해요|원합니다)|"
    r"구합니다|구해요|구하죠|찾습니다|찾고\s*있습니다|"
    r"제공\s*(?:합니다|해드립니다|드립니다|가능)|공급\s*(?:합니다|가능)|"
    r"납품\s*(?:합니다|가능)|취급\s*(?:합니다|중)|"
    r"대량\s*(?:보유|판매|매입|공급)|"
    r"(?:DB|디비|계좌|통장|계정|아이디|여권|신분증|민증)\s*"
    r"(?:대량\s*)?(?:판매|매입|구매|공급|납품|취급|보유|임대|대여)|"
    r"판매\s*(?:업체|전문)|(?:주문|구매|판매)\s*문의|"
    r"(?:임대|대여|중개)\s*(?:합니다|해요|중|가능|받습니다|원합니다)|"
    r"거래\s*(?:합니다|해요|중|가능합니다|받습니다|원합니다)|"
    r"위조\s*(?:가능|전문|의뢰|문의)|"
    r"(?:여권|신분증|민증|주민등록증|면허증|운전면허증)\s*"
    r"(?:위조|복제|제작)\s*(?:가능|전문|의뢰|문의)?|"
    r"의뢰\s*(?:받습니다|받아요|주세요|문의)",
    re.IGNORECASE,
)
RELEVANCE_WEAK_OFFER = re.compile(
    r"건당|단가|가격\s*[:：]?|대량\s*진행|"
    r"(?:판매|매입|구매|거래|주문)\s*(?:문의|상담)|"
    r"(?:DB|디비|데이터|계정|아이디).{0,40}(?:필요하신|필요하시면|필요하면)|"
    r"문의\s*(?:주세요|주시면|주시기|주십쇼|바랍니다|가능|부탁)",
    re.IGNORECASE,
)
RELEVANCE_NEGATED_OFFER = re.compile(
    r"(?:판매|매입|구매|거래|대여|제휴)(?:은|는|를)?\s*"
    r"(?:하지|받지)\s*않",
    re.IGNORECASE,
)
RELEVANCE_REPORTING_CONTEXT = re.compile(
    r"뉴스|기사(?:본문)?|보도(?:자료|입니다)?|적발|검거|체포|기소|송치|구속|"
    r"경찰|검찰|법원|판결|선고|혐의|피고인|사건\s*(?:요약|개요)|"
    r"\[(?:기고|취재파일|단독)\]|편집자\s*주|"
    r"(?:기자|특파원)\s*(?:=|·|:)|무단전재|재배포\s*금지|"
    r"취재를\s*종합하면|편집자\s*주|"
    r"(?:관련\s*업계|당국|업계).{0,30}(?:따르면|밝혔|전했)|"
    r"(?:밝혀졌|알려졌|보도했|전해졌|나타났|해석된다|지적이\s*나온다)|"
    r"상담사례|법률\s*상담|처벌|대응\s*방법|예방|주의(?:하세요|해야)|경고|"
    r"피해\s*(?:사례|경험담)|사기\s*(?:입니다|당했|피해)|(?:경찰|수사대)에?\s*신고|"
    r"개인정보보호법\s*위반|(?:전화|연락).{0,40}(?:폭주|차단|받으시나요|오나요|왔나요)|"
    r"(?:판매|매입|구매)하라는\s*(?:DM|디엠|연락|문자)|"
    r"(?:불법\s*)?(?:거래|판매).{0,20}(?:성행|활개|우려|논란)|"
    r"확인\s*방법|궁금(?:합니다|할)|"
    r"(?:할까요|인가요|되나요|있나요|없나요|아시나요|가능한가요)\s*[?？]?",
    re.IGNORECASE,
)
RELEVANCE_LEGAL_DECISION_CONTEXT = re.compile(
    r"조세심판원|심판청구|처분개요|처분청|청구인은|청구주장|"
    r"쟁점금액|귀속분|경정[·ㆍ]?고지|주문\s*심판청구를\s*기각|"
    r"이유\s*\d+\.?\s*처분개요|판례|사건번호",
    re.IGNORECASE,
)
RELEVANCE_SINGLE_ACCOUNT_CONTEXT = re.compile(
    r"게임\s*계정|FC\s*모바일|로드\s*모바일|피파(?:온라인)?|한게임|순비피|"
    r"롤\s*계정|롤계정|리그\s*오브\s*레전드|"
    r"배틀그라운드|카트라이더|쿠키런|아이온|메이플(?:스토리)?|리니지|"
    r"던전앤파이터|로스트아크|바람의나라(?:\s*연)?|"
    r"넥슨\s*계정|스팀\s*계정|게임머니|캐릭터\s*(?:판매|거래)|"
    r"(?:게임상|캐릭터|레벨|전투력|아이템).{0,30}계정|"
    r"계정.{0,30}(?:게임상|캐릭터|레벨|전투력|아이템)|"
    r"구글\s*연동|계정\s*하나|실사용(?:하던)?\s*계정|계정\s*급처|계정\s*스펙",
    re.IGNORECASE,
)
RELEVANCE_NORMAL_PRODUCT_CONTEXT = re.compile(
    r"여권\s*(?:케이스|커버|지갑)|(?:통장|카드)\s*(?:케이스|커버|지갑|비닐)|"
    r"(?:마이너스|청약|어린이|적금|예금|급여|입출금)\s*통장|"
    r"(?:은행|농협|금융사).{0,30}(?:통장|계좌)\s*(?:상품|출시|판매)|"
    r"(?:통장|계좌).{0,20}(?:금리|대출|상품|출시|가입)|"
    r"(?:모바일\s*신분증|정부24|PASS\s*앱).{0,30}(?:발급|재발급|등록|사용)|"
    r"(?:IRP|퇴직연금|연금저축).{0,30}(?:계좌|국채|매입|판매)|"
    r"(?:비즈니스|개인|기업)\s*(?:체킹|checking)\s*(?:어카운트|계좌)?|"
    r"(?:checking\s*account|체킹\s*계좌|FDIC|예금자\s*보호)|"
    r"(?:Steam\s*Deck|스팀\s*덱).{0,80}(?:제품|대량\s*구매)|"
    r"(?:저울|제품|상품|모델|사양|견적).{0,120}\bDB[-_]\d+[A-Za-z0-9-]*|"
    r"\bDB[-_]\d+[A-Za-z0-9-]*.{0,120}(?:저울|제품|상품|모델|사양|견적)|"
    r"(?:계좌간\s*환전|외화\s*통장|원화\s*통장|외환\s*(?:매입|매도|환전))|"
    r"(?:계정|아이디).{0,25}(?:정지\s*조건|복구\s*방법|해지\s*방법|만드는\s*법)",
    re.IGNORECASE,
)
RELEVANCE_DB_IT_SYSTEM_CONTEXT = re.compile(
    r"(?:DBMS|eXperDB|PostgreSQL|MariaDB|Oracle\s*DB).{0,180}"
    r"(?:구축|설치|소프트웨어|기술\s*지원|서버|입찰|프로젝트)|"
    r"(?:채널계|전산\s*시스템|정보\s*시스템|서버).{0,100}(?:DB|디비).{0,100}"
    r"(?:분리|구축|구매|납품|입찰)|"
    r"(?:DB|디비).{0,80}(?:분리\s*구축|DBMS\s*환경\s*구축|"
    r"소프트웨어\s*구매|기술\s*지원\s*및\s*교육)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_DB_COMPLIANCE_GUIDE_PHRASE = re.compile(
    r"개인정보.{0,80}(?:수집|이용|제\s*3\s*자\s*제공).{0,100}"
    r"(?:동의|절차|확인|중요)|"
    r"(?:정보의\s*출처|수집\s*목적|이용\s*목적).{0,100}"
    r"(?:동의|절차|확인|중요)|"
    r"(?:이름|전화번호|연락처).{0,100}개인정보.{0,80}"
    r"(?:수집|이용|제\s*3\s*자\s*제공)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_DB_COMPLIANCE_GUIDE_STRUCTURE = re.compile(
    r"확인할\s*내용|주의할\s*점|안전하게\s*운영|"
    r"운영하는\s*방법|자주\s*묻는\s*질문|대신\s*알아보는|"
    r"접근해서는\s*안\s*됩니다|신중하게\s*확인",
    re.IGNORECASE,
)
RELEVANCE_DERIVATIVES_TRADING_SERVICE = re.compile(
    r"해외\s*선물.{0,160}(?:대여\s*계좌|대여\s*업체|"
    r"실체결|증거금|나스닥|선물\s*지수)|"
    r"(?:대여\s*계좌|대여\s*업체).{0,160}"
    r"(?:해외\s*선물|나스닥|선물\s*지수|증거금)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_EVENT_TICKET_CONTEXT = re.compile(
    r"(?:공연|콘서트|팬클럽|관객\s*입장|공연\s*시작).{0,180}"
    r"(?:티켓|예매|좌석|추첨제)|"
    r"(?:티켓|예매|좌석|추첨제).{0,180}"
    r"(?:공연|콘서트|팬클럽|관객\s*입장|공연\s*시작)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_NEGATED_DB_TRADE_CONTEXT = re.compile(
    # Reluctant gaps keep two nearby negated statements as two matches.  A
    # greedy gap could swallow both sentences and make repeated sales copy
    # such as "DB를 구매하지 않습니다 ... DB 구매를 하지 않고" look
    # like only one weak negation.
    r"(?:DB|디비).{0,100}?(?:구매|매입).{0,30}?(?:하지\s*않|안\s*(?:하|합))|"
    r"(?:구매|매입).{0,30}?(?:하지\s*않|안\s*(?:하|합)).{0,100}?(?:DB|디비)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_NORMAL_ID_PRODUCT_CONTEXT = re.compile(
    r"(?:사원증|학생증|방문증|출입증|협회\s*신분증|종교\s*신분증|"
    r"미니\s*신분증|명찰|자격증).{0,120}"
    r"(?:상품명|상품목록|ITEMS?|장바구니|배송조회|결제)|"
    r"(?:상품명|상품목록|ITEMS?|장바구니|배송조회|결제).{0,120}"
    r"(?:사원증|학생증|방문증|출입증|협회\s*신분증|종교\s*신분증|"
    r"미니\s*신분증|명찰|자격증)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_DRIVER_LICENSE_PHOTO_GUIDE = re.compile(
    r"운전면허증\s*제작용\s*사진.{0,80}(?:표준\s*규격|제출)|"
    r"(?:표준\s*규격|사진\s*규격).{0,80}운전면허증",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_GIFT_CARD_ONLY = re.compile(
    r"(?:상품권|문화상품권|기프티콘).{0,80}(?:매입|판매|구매|현금화)|"
    r"(?:매입|판매|구매|현금화).{0,80}(?:상품권|문화상품권|기프티콘)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_CORE_PERSONAL_TARGET = re.compile(
    r"개인정보|고객\s*(?:DB|디비|명단|리스트|정보)|회원\s*(?:DB|디비|명단|리스트|정보)|"
    r"(?:대출|보험|주식|부동산|마케팅|업체|사업자)\s*(?:DB|디비|명단|리스트)|"
    r"연락처|전화번호|휴대폰번호|주민등록번호|주민번호|이메일\s*(?:DB|목록|리스트)|"
    r"(?:네이버|다음|구글|쿠팡|배민|밴드|인스타(?:그램)?|페이스북|트위터|틱톡)\s*"
    r"(?:계정|아이디)|(?:카카오톡|카톡|텔레그램|텔그)\s*계정|"
    r"(?:통장|계좌|여권|신분증|주민등록증|운전면허증|외국인등록증)",
    re.IGNORECASE,
)
RELEVANCE_PUBLIC_BUSINESS_DIRECTORY_CONTEXT = re.compile(
    r"(?:전국|국내).{0,40}(?:학원|PC방|피시방|미용실|인테리어|업체|사업자)"
    r".{0,40}(?:주소록|연락처)\s*(?:DB|디비)?|"
    r"(?:업장명|상호명).{0,120}(?:구주소|신주소|우편번호).{0,120}"
    r"(?:팩스|홈페이지|업종)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_NORMAL_TELECOM_SERVICE = re.compile(
    r"인터넷\s*3사\s*(?:비교|가입)\s*(?:상담|문의)|"
    r"(?:SK|KT|LG).{0,120}(?:요금제|결합|IPTV|설치\s*가능)|"
    r"(?:통신사별|인터넷\s*가입).{0,100}(?:요금|약정|결합상품|설치\s*가능)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_QUESTION_OR_GUIDE_TITLE = re.compile(
    r"(?:방법|사용법|가이드|문의사항|조건과|조건\s*총정리|총정리|차이|"
    r"주의보|성행|알고\s*보니|도구\s*\d+가지|부수입\s*만들기)|"
    r"(?:고소|신고|처벌|사기).{0,20}(?:되|돼|됨|하|당|인가|인가요)|"
    r"(?:판매|매입|구매|거래).{0,20}(?:질문|해도|되나|될까|사기칠)|"
    r"(?:어떻게|왜|무슨\s*일|이런\s*경우)|"
    r"구별법|믿어도\s*될까",
    re.IGNORECASE,
)
RELEVANCE_BUYING_INQUIRY = re.compile(
    r"(?:DB|디비|계정|아이디|통장|계좌|여권|신분증).{0,160}"
    r"(?:(?:가격|단가|얼마).{0,80}(?:구매|매입|삽니다|구합니다)|"
    r"(?:구매|매입|삽니다|구합니다).{0,80}(?:가격|단가|얼마|문의))|"
    r"(?:구매|매입|삽니다|구합니다).{0,80}"
    r"(?:DB|디비|계정|아이디|통장|계좌|여권|신분증).{0,80}"
    r"(?:가격|단가|얼마|문의)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_BUYING_INQUIRY_QUESTION = re.compile(
    r"(?:가격|단가|얼마).{0,100}"
    r"(?:인가요|하나요|할까요|되나요|맞는지|맞는\s*건지|모르겠|궁금|문의드)|"
    r"(?:구매|매입|삽니다|구합니다).{0,80}(?:가격|단가|얼마).{0,80}"
    r"(?:문의드|궁금|인가요|하나요|할까요)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_AGGREGATION_OR_COMMENTARY_CONTEXT = re.compile(
    r"박제\s*(?:채널|방)|피해자\s*제보|사건\s*내용|"
    r"이슈\s*/\s*유머|"
    r"(?:뭐|무엇)하는데|왜\s*(?:저런|이런)|이상하(?:네요|구요)|"
    r"피싱\s*범죄|보안.{0,20}점검|범죄인데",
    re.IGNORECASE,
)
RELEVANCE_WARNING_AGAINST_TRADE = re.compile(
    r"(?:판매자들?|구매자들?)\s*필독|절대\s*(?:팔지|사지|거래하지)\s*마|"
    r"(?:절대\s*)?(?:구매|매입|판매|거래)\s*하지\s*마|"
    r"(?:팔면|파는\s*순간).{0,40}(?:불법|처벌|법으로\s*엮)|"
    r"(?:되팔렘|사기꾼|업자).{0,80}(?:조심|주의|피해|악질)|"
    r"(?:계정|아이디|DB|디비|통장).{0,80}(?:팔지|사지)\s*마|"
    r"(?:명의|계정|아이디)를?\s*넘기는\s*(?:것|거).{0,60}"
    r"(?:잘못|불법|범죄)|"
    r"(?:계정|아이디)를?\s*판매하게\s*되면.{0,120}(?:범죄|사기|악용)|"
    r"이런\s*(?:계정|아이디)\s*판매.{0,80}무시하",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_TRADE_MOTIVE_QUESTION = re.compile(
    r"(?:계정|아이디|DB|디비|통장|계좌)\s*(?:사는|사가는|매입하는)\s*"
    r"(?:애들?|사람들?|업자들?).{0,50}(?:뭐야|왜|뭐\s*하려)|"
    r"왜.{0,50}(?:계정|아이디|DB|디비|통장|계좌).{0,30}"
    r"(?:사는|사가는|매입하는)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_EMPTY_CONTAINER_CONTEXT = re.compile(
    r"(?:등록된|작성된|표시할)\s*(?:게시물|게시글|글)이?\s*(?:없습니다|없음)|"
    r"(?:게시물|게시글|포스트)가?\s*(?:없습니다|없음)|"
    r"no\s+(?:posts?|entries)\s+(?:to\s+display|found)|nothing\s+here|"
    r"此处没有可显示的博文",
    re.IGNORECASE,
)
RELEVANCE_EMPTY_LISTING_TEMPLATE = re.compile(
    r"1\.\s*(?:페이지명|계정명|채널명)\s*:\s*\n\s*"
    r"2\.\s*팔로워\s*수\s*:\s*\n\s*"
    r"3\.\s*매매가\s*:\s*\n\s*"
    r"4\.\s*안전거래\s*가능\s*여부\s*:\s*\n\s*"
    r"5\.\s*거래\s*문의\s*연락처[^:]*:\s*\n",
    re.IGNORECASE,
)
RELEVANCE_CORPORATE_TRANSFER_CONTEXT = re.compile(
    r"(?:법인|회사)\s*(?:양도\s*양수|양도양수|매매|매입).{0,200}"
    r"(?:대표자\s*변경|자본금|법무사|세금\s*체납)|"
    r"(?:대표자\s*변경|자본금|법무사|세금\s*체납).{0,200}"
    r"(?:법인|회사)\s*(?:양도\s*양수|양도양수|매매|매입)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_SEO_EXPLAINER_PHRASE = re.compile(
    r"자주\s*언급됩니다|도움이\s*될\s*수\s*있습니다|"
    r"중요하게\s*(?:봅니다|언급됩니다)|살펴보는\s*것이\s*좋습니다|"
    r"확인하는\s*경우가\s*많습니다|구성\s*포인트|"
    r"함께\s*담으면\s*좋은\s*내용|페이지\s*완성도|"
    r"방향을\s*잡는\s*데\s*도움|구조를\s*말합니다|"
    r"개념으로\s*이해할\s*수\s*있습니다|검토하는\s*방식이\s*중요|"
    r"안내하는\s*것이\s*중요|확인해야\s*할\s*기준|"
    r"보는\s*것이\s*좋습니다|관리\s*기준\s*마련|"
    r"안내형\s*페이지|안내형\s*콘텐츠|문단별?\s*역할|페이지\s*성격|"
    r"설명형\s*(?:문장|흐름)|키워드.{0,30}(?:반복|나열)|"
    r"페이지.{0,30}(?:구성|흐름).{0,30}(?:자연스럽|안정적)|"
    r"이\s*페이지는.{0,180}(?:정리|구성)|페이지를\s*보는\s*사람|"
    r"(?:강한|자극적인)\s*문구|설득력\s*있는\s*흐름|"
    r"페이지.{0,40}신뢰를\s*보여|"
    r"열쇠는.{0,80}입니다|도움이\s*됩니다|유리합니다|"
    r"매력적인\s*옵션|가능하게\s*합니다|"
    r"중요할\s*수\s*있습니다|새로운\s*길을\s*제공",
    re.IGNORECASE,
)
RELEVANCE_UNAMBIGUOUS_OFFER = re.compile(
    r"판매\s*(?:합니다|해요|중)|팝니다|매입\s*(?:합니다|해요|중)|삽니다|"
    r"구매\s*(?:합니다|해요|원합니다)|제공\s*(?:합니다|해드립니다|드립니다)|"
    r"공급\s*(?:합니다|가능)|납품\s*(?:합니다|가능)|"
    r"(?:위조|제작)\s*(?:가능|전문|의뢰)|의뢰\s*(?:받습니다|받아요)",
    re.IGNORECASE,
)
RELEVANCE_ATTRIBUTABLE_OFFER_VERB = re.compile(
    r"판매\s*(?:합니다|해요|중|가능)|팝니다|"
    r"매입\s*(?:합니다|해요|중|가능)|삽니다|"
    r"구매\s*(?:합니다|해요|원합니다)|"
    r"구합니다|구해요|찾습니다|찾고\s*있습니다|"
    r"(?:임대|대여|중개)\s*(?:합니다|해요|중|가능|받습니다|원합니다)|"
    r"위조\s*(?:가능|전문|의뢰|문의)|의뢰\s*(?:받습니다|받아요)",
    re.IGNORECASE,
)
RELEVANCE_ACCOUNT_TOOL_CONTEXT = re.compile(
    r"(?:계정|아이디)\s*(?:자동\s*)?생성\s*프로그램|"
    r"(?:텔레그램|카카오톡|카톡)\s*(?:무한\s*)?생성\s*프로그램|"
    r"다중\s*(?:멀티\s*)?접속\s*(?:프로그램|시스템)|자동\s*(?:친구\s*추가|발송기)|"
    r"번호\s*인증\s*프로그램|"
    r"(?:텔레그램|업무\s*자동화|알림|예약|관리)\s*봇.{0,80}"
    r"(?:맞춤\s*제작|업무\s*자동화|API\s*연동)|"
    r"(?:맞춤\s*제작|업무\s*자동화|API\s*연동).{0,80}"
    r"(?:텔레그램|업무\s*자동화|알림|예약|관리)\s*봇",
    re.IGNORECASE,
)
RELEVANCE_VIRTUAL_NUMBER_SERVICE = re.compile(
    r"가상\s*(?:전화)?번호.{0,120}(?:SMS|문자)\s*인증|"
    r"(?:SMS|문자)\s*인증.{0,120}가상\s*(?:전화)?번호|"
    r"인증번호\s*(?:수신|발급).{0,80}(?:국가|자동화|번호)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_DB_MANAGEMENT_SOFTWARE = re.compile(
    r"(?:DB\s*Manager|디비\s*매니저)|"
    r"(?:DB|디비)\s*구매\s*(?:대신|하지\s*않)|"
    r"자체\s*수집.{0,100}(?:통합\s*)?관리|"
    r"고객\s*데이터.{0,100}(?:관리\s*시스템|자동\s*수집)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_INFORMATIONAL_ARTICLE_PHRASE = re.compile(
    r"이\s*글은.{0,80}(?:정리한|설명하는)\s*것입니다|"
    r"(?:관점|측면)에서\s*(?:살펴|정리|보면)|"
    r"(?:연구|안내서|법률|보고서).{0,80}(?:따르면|의하면)|"
    r"가정\s*(?:시나리오|예시)|사실관계\s*전달|"
    r"업데이트\s*공지|알려준\s*적|기능도?\s*(?:더\s*)?강화|"
    r"관심(?:을)?\s*가져보는\s*걸\s*추천|시사하는\s*것|"
    r"설계\s*원칙|핵심\s*질문|"
    r"목적\s*:\s*(?:잠재|기존|동일|고객)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_TRADE_GUIDE_PHRASE = re.compile(
    r"지속\s*가능한\s*성장|현실적인\s*(?:방법|대안)|진짜\s*활성화|"
    r"(?:후기|거래).{0,30}위험.{0,20}(?:분석|정리)|"
    r"장점.{0,80}(?:위험|주의)|예상치\s*못한\s*위험|"
    r"(?:꼼꼼하게|자세히)\s*(?:파헤쳐|알아보)|"
    r"계정\s*회수.{0,80}(?:피해|위험|분쟁)|"
    r"정책\s*위반|(?:거래|판매).{0,80}지양해야|"
    r"콘텐츠\s*(?:경쟁력|품질)|운영\s*(?:전략|노하우)|"
    r"구매\s*전.{0,30}(?:점검|확인)|안전한\s*(?:방법|거래)|"
    r"어떻게\s*구해야|판매처\s*선택\s*가이드|"
    r"활용하는\s*노하우|좋은\s*(?:DB|디비)란|"
    r"(?:함께|자세히)\s*파헤쳐",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_PASSPORT_PHOTO_GUIDE = re.compile(
    r"여권\s*사진.{0,220}(?:보정|촬영|규정|발급|반려|"
    r"출입국|얼굴\s*인식)|"
    r"(?:보정|촬영|사진\s*규정).{0,220}여권\s*사진|"
    r"(?:신분증|여권|운전면허증)\s*사진.{0,120}"
    r"(?:제작기|편집기|만들기|자르기|배경|조명|크기|무료|온라인)|"
    r"(?:제작기|편집기|만들기).{0,120}"
    r"(?:신분증|여권|운전면허증)\s*사진",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_LICENSE_REQUIREMENTS_GUIDE = re.compile(
    r"운전면허증에\s*대해|"
    r"국제\s*운전면허증.{0,500}"
    r"(?:발급국|유효\s*기한|거주\s*증명|제네바\s*조약|차량\s*대여)|"
    r"(?:발급국|유효\s*기한|거주\s*증명|제네바\s*조약|차량\s*대여).{0,500}"
    r"국제\s*운전면허증",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_INSURANCE_RECRUITMENT = re.compile(
    r"보험\s*설계사|보험\s*영업|GA\s*보험사|GA\s*대리점|"
    r"전속\s*FC|FC\s*입사|보험\s*대리점",
    re.IGNORECASE,
)
RELEVANCE_RECRUITMENT_OR_ORG_OFFER = re.compile(
    r"채용|입사|이직|위촉|리쿠르팅|소속\s*설계사|"
    r"지점\s*(?:지원|오픈)|본부.{0,40}(?:지원|제공)|"
    r"수수료\s*(?:체계|개편|분급제)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_RECRUITMENT_DB_BENEFIT = re.compile(
    r"(?:DB|디비).{0,80}(?:무료|무한\s*생성|지원|제공|"
    r"고객\s*유입|내방객|영업\s*시스템)|"
    r"(?:무료|지원|제공).{0,80}(?:고객\s*)?(?:DB|디비)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_DB_PURCHASE_ALTERNATIVE = re.compile(
    r"(?:DB|디비)\s*(?:구매|구입)\s*없이|"
    r"(?:DB|디비)를?\s*사지\s*않고|"
    r"(?:DB|디비)\s*사는\s*시대는?\s*끝|"
    r"(?:DB|디비)를?\s*사는\s*것이\s*아니라|"
    r"외부\s*(?:DB|디비)\s*구매.{0,300}(?:대신|벗어나|직접\s*만들)|"
    r"유료\s*(?:DB|디비).{0,200}(?:대신|벗어나|직접\s*만들)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_LONGFORM_REPORT_PHRASE = re.compile(
    r"심층\s*보고서|목차|서론|현황|정책\s*제언|"
    r"제도\s*개선|결론|논리적\s*검토|종합적으로\s*고찰",
    re.IGNORECASE,
)
RELEVANCE_DB_BRAND_OR_STOCK = re.compile(
    r"DB하이텍|DB\s*글로벌칩|DB\s*손해보험|DB손보|"
    r"(?:파운드리|반도체|주가|매수).{0,100}DB하이텍",
    re.IGNORECASE,
)
RELEVANCE_DB_PC_JOB_CONTEXT = re.compile(
    r"(?:DB|디비)\s*(?:PC|피시)\s*(?:카페|방).{0,500}"
    r"(?:시급|알바|구인|매장\s*관리|월\s*[~～-]\s*금)|"
    r"(?:시급|알바|구인|매장\s*관리).{0,500}"
    r"(?:DB|디비)\s*(?:PC|피시)\s*(?:카페|방)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_NORMAL_ID_CARD_MARKET = re.compile(
    r"신분증\s*제작\s*업체.{0,160}(?:트렌드|시장\s*개요|시장)|"
    r"(?:스마트\s*카드|NFC).{0,300}RFID.{0,300}"
    r"(?:보안|접근\s*제어)|"
    r"(?:NFC|RFID).{0,300}(?:사원증|출입\s*카드|스마트\s*카드)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_DOCUMENTARY_CASE_CONTEXT = re.compile(
    r"보더\s*시큐리티|"
    r"(?:이민성|공항)\s*직원.{0,500}(?:승객|인터뷰|통역사)|"
    r"(?:승객|인터뷰|통역사).{0,500}(?:이민성|공항)\s*직원",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_INSURANCE_INDUSTRY_ANALYSIS = re.compile(
    r"(?:1[,.]?200%\s*룰|4년\s*분급제|7년\s*분급제|"
    r"판매\s*수수료\s*개편).{0,500}"
    r"(?:규제|가이드라인|변화|분석|전략|영향)|"
    r"(?:규제|가이드라인|변화|분석|전략|영향).{0,500}"
    r"(?:1[,.]?200%\s*룰|4년\s*분급제|7년\s*분급제|"
    r"판매\s*수수료\s*개편)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_SOCIAL_METRIC_SERVICE = re.compile(
    r"(?:팔로워|구독자|좋아요|조회수|뷰)\s*(?:구매|늘리기|증가|서비스)|"
    r"(?:팔로워|구독자|좋아요|조회수|뷰).{0,30}(?:품질|유지|이탈|활성)",
    re.IGNORECASE,
)
RELEVANCE_ACCOUNT_OFFER = re.compile(
    r"(?:계정|아이디)\s*(?:판매|팝니다|매입|삽니다|구매|대여|임대)|"
    r"(?:판매|팝니다|매입|삽니다|구매|대여|임대)\s*(?:하는|할|합니다|해요|중)?\s*"
    r"(?:계정|아이디)",
    re.IGNORECASE,
)
RELEVANCE_SEO_SPAM_TEMPLATE_PHRASE = re.compile(
    r"예약\s*(?:방법|확정|문의|시)|방문\s*(?:및\s*이용|고객|전\s*문의)|"
    r"(?:원하시는|이용)\s*코스|심야\s*(?:시간대|에도)|"
    r"평일[·ㆍ/\s-]*주말\s*(?:모두\s*)?영업|"
    r"(?:재방문|방문)\s*의사|단골\s*고객|이용\s*후기|"
    r"편안하게\s*(?:모십니다|즐길)|분위기를\s*찾|"
    r"날짜와\s*인원|당일\s*예약|첫\s*방문\s*고객|"
    r"저녁부터\s*심야|가장\s*좋은\s*조건으로\s*예약",
    re.IGNORECASE,
)
RELEVANCE_COHERENT_ACCOUNT_SERVICE = re.compile(
    r"(?:아이디|계정)\s*다량\s*보유.{0,40}즉시\s*거래\s*가능|"
    r"(?:블로그|카페|스마트스토어)용.{0,40}(?:계정\s*)?제공|"
    r"(?:네이버|인스타그램|페이스북|틱톡).{0,80}"
    r"(?:아이디|계정).{0,30}(?:판매|구매|임대).{0,80}"
    r"(?:서비스\s*문의|주문)|"
    r"(?:계정|아이디)\s*(?:판매|구매).{0,20}/.{0,20}"
    r"(?:계정|아이디)\s*(?:판매|구매)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_BODY_BOILERPLATE = re.compile(
    r"개인정보\s*(?:처리|취급)방침|"
    r"통신판매업신고번호|통신판매중개자로서|사업자등록번호|"
    r"통장업로드|전표전송|회계프로그램",
    re.IGNORECASE,
)
RELEVANCE_MARKET_GUIDE_CONTEXT = re.compile(
    r"계정\s*주요\s*기능\s*안내|구매\s*(?:가이드|선택\s*방법)|"
    r"안전하고\s*신뢰할\s*수\s*있는\s*옵션",
    re.IGNORECASE,
)
RELEVANCE_MARKET_GUIDE_WEAK = re.compile(
    r"시장\s*(?:규모|동향|전망)|(?:공급업체|판매업체)\s*(?:선정|선택|비교)|"
    r"구매자\s*리뷰|비교표|품질\s*검증\s*절차|최소\s*주문\s*수량",
    re.IGNORECASE,
)
RELEVANCE_LEGAL_PROP_OR_SECURITY_CONTEXT = re.compile(
    r"(?:소품용|촬영\s*소품|촬영용|연출용).{0,80}"
    r"(?:여권|신분증|운전면허증|면허증|의사면허증)|"
    r"(?:여권|신분증|운전면허증|면허증|의사면허증).{0,80}"
    r"(?:소품용|촬영\s*소품|촬영용|연출용)|"
    r"(?:촬영|영화|드라마|광고).{0,80}(?:여권|신분증|운전면허증).{0,80}소품"
    r".{0,300}(?:VOID|SAMPLE|비식별|무효|실제\s*개인정보\s*사용\s*금지)|"
    r"(?:미술팀|미술\s*제작|아트워크|art\s*work|design\s*service).{0,200}"
    r"(?:여권|신분증|운전면허증).{0,80}(?:제작|의뢰)|"
    r"(?:여권|신분증|운전면허증).{0,80}(?:제작|의뢰).{0,200}"
    r"(?:미술팀|미술\s*제작|아트워크|art\s*work|design\s*service)|"
    r"(?:여권|신분증|운전면허증)\s*위조\s*방지.{0,150}"
    r"(?:보안\s*(?:구조|설계)|안전\s*(?:기준|설계))|"
    r"실제\s*(?:발급|대행|위조).{0,40}(?:제공하지|하지\s*않)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_TELEGRAM_DIRECTORY_CONTEXT = re.compile(
    r"엄선된\s*Telegram\s*채널.{0,80}(?:그룹|봇).{0,80}한\s*곳에서|"
    r"(?:카테고리\s*순위|구독자).{0,100}(?:최근\s*게시물|업데이트)",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_ATTRIBUTABLE_BODY_CONTACT = re.compile(
    r"(?:카카오톡|카톡|텔레그램|텔그|오픈채팅|오픈톡|문의|연락)"
    r".{0,100}(?:\[ACCOUNT\]|\[MESSENGER_ID\]|https?://|@\w{3,})",
    re.IGNORECASE | re.DOTALL,
)
RELEVANCE_LISTING_DETAIL = re.compile(
    r"가격|값|단가|팔로워|좋아요|라이브|매매가|선제시|송금|입금|"
    r"대여|판매|매입|구매|계정|아이디|비밀번호|비번",
    re.IGNORECASE,
)
RELEVANCE_EXCLUDED_TITLE = re.compile(
    r"개인정보\s*처리방침|개인정보\s*보호정책|이용\s*약관|서비스\s*약관|"
    r"법적\s*고지|운영\s*정책|관련\s*(?:팁|주의사항)|"
    r"고객센터|고객지원|도움말|자주\s*묻는\s*질문|FAQ|로그인|회원가입|"
    r"계정\s*만들기|바로가기|사용법|동기화|다운로드|설치|위키|사전|매뉴얼|"
    r"documentation|codelab|고객\s*권리\s*안내|신용정보\s*권리|"
    r"호스팅\s*(?:홍보|서비스)|서버\s*(?:임대|호스팅)",
    re.IGNORECASE,
)
RELEVANCE_EXCLUDED_DOMAINS = {
    "apple.com",
    "citibank.co.kr",
    "claude.com",
    "enuri.com",
    "google.com",
    "google.co.kr",
    "ibk.co.kr",
    "kakao.com",
    "kakaobank.com",
    "kakaocorp.com",
    "kbstar.com",
    "kbanknow.com",
    "kebhana.com",
    "messenger.com",
    "minecraft.wiki",
    "nhbank.com",
    "nonghyup.com",
    "privacy.go.kr",
    "shinhan.com",
    "snuh.org",
    "standardchartered.co.kr",
    "thewiki.kr",
    "tossbank.com",
    "wikimedia.org",
    "wikipedia.org",
    "wiktionary.org",
    "wooribank.com",
    "zeta-ai.io",
}
RELEVANCE_PRESS_DOMAINS = {
    "aagag.com",
    "asiatoday.co.kr",
    "chosun.com",
    "ddaily.co.kr",
    "digitaltoday.co.kr",
    "donga.com",
    "dt.co.kr",
    "edaily.co.kr",
    "etnews.com",
    "fnnews.com",
    "hani.co.kr",
    "hankyung.com",
    "imbc.com",
    "joins.com",
    "joongang.co.kr",
    "jtbc.co.kr",
    "kbs.co.kr",
    "kmib.co.kr",
    "mbn.co.kr",
    "mediatoday.co.kr",
    "mk.co.kr",
    "mt.co.kr",
    "news1.kr",
    "newsis.com",
    "nocutnews.co.kr",
    "ohmynews.com",
    "sbs.co.kr",
    "segye.com",
    "seoul.co.kr",
    "yna.co.kr",
    "ytn.co.kr",
    "zdnet.co.kr",
}
RELEVANCE_DISCOVERY_PRESS_PATH = re.compile(
    r"(?:^|/)(?:news|press|article|articles)(?:/|\.|$)|"
    r"(?:articleview|newsid|arcid)=",
    re.IGNORECASE,
)
RELEVANCE_DISCOVERY_PRESS_CONTEXT = re.compile(
    r"기자|취재진|보도(?:했다|자료)?|언론|적발|검거|체포|기소|송치|구속|"
    r"경찰|검찰|법원|판결|혐의|우려|논란|성행|기승|활개|"
    r"(?:밝혔|전했|알려졌|나타났)|무단전재|재배포\s*금지",
    re.IGNORECASE,
)


def nearby_matches(
    left: re.Match[str] | None,
    right: re.Match[str] | None,
    maximum_gap: int = 120,
) -> bool:
    """Return whether two signals occur in the same compact phrase context."""
    if left is None or right is None:
        return False
    gap = max(left.start() - right.end(), right.start() - left.end(), 0)
    return gap <= maximum_gap


def looks_like_keyword_stuffing(title: str, text: str) -> bool:
    """Detect long phrase dumps that contain no attributable transaction offer."""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) < 40:
        return False
    short_ratio = sum(len(line) <= 28 for line in lines) / len(lines)
    average_length = sum(map(len, lines)) / len(lines)
    punctuation_ratio = sum(
        bool(re.search(r"[.!?。？！]$", line)) for line in lines
    ) / len(lines)
    combined = title + "\n" + text
    target_mentions = len(RELEVANCE_STRICT_TARGET.findall(combined)) + len(
        RELEVANCE_STRICT_SHORT_TARGET.findall(combined)
    )
    trade_mentions = len(RELEVANCE_TRADE.findall(combined))
    normalized_lines = [re.sub(r"\s+", " ", line.lower()) for line in lines]
    repeated_rows = sum(
        count for count in Counter(normalized_lines).values() if count > 1
    )
    topic_families = sum(
        bool(pattern.search(combined))
        for pattern in (
            re.compile(r"상품권|기프티콘", re.I),
            re.compile(r"비트코인|테더|이더리움|리플|암호화폐|코인", re.I),
            re.compile(r"해킹|해커|게시물\s*내리기|공격", re.I),
            re.compile(r"라우터|에그|쌍둥이폰|번호판", re.I),
            re.compile(r"신분증|운전면허증|여권|주민등록증", re.I),
            re.compile(r"계정|아이디|인증", re.I),
            re.compile(r"포커|머니상|환전", re.I),
        )
    )
    # Search-poisoning pages often paste the same block of unrelated illicit
    # keywords twice. This differs from an attributable multi-product listing:
    # there are almost no sentences, prices/forms, or explicit seller claims.
    repeated_multitopic_dump = bool(
        len(lines) >= 45
        and repeated_rows / len(lines) >= 0.55
        and topic_families >= 4
        and punctuation_ratio <= 0.08
        and average_length <= 80
    )
    if repeated_multitopic_dump:
        return True
    if short_ratio < 0.85:
        return False
    phrase_dump = (
        len(lines) >= 40
        and short_ratio >= 0.95
        and average_length <= 12
        and punctuation_ratio <= 0.03
    )
    if phrase_dump:
        # Telegram trade channels also use hundreds of short lines for prices,
        # form fields and repeated offers.  A coherent transaction form plus an
        # attributable offer/contact is evidence of a real listing, not SEO
        # keyword stuffing.
        structured_trade = bool(
            re.search(
                r"거래\s*양식|(?:성함|이름).{0,100}(?:전화번호|연락처|번호)"
                r".{0,100}(?:계좌|신분증)|"
                r"(?:아이디|계정).{0,80}(?:비번|비밀번호).{0,120}"
                r"(?:성함|이름|전화번호|계좌)",
                text,
                re.IGNORECASE | re.DOTALL,
            )
            and RELEVANCE_DIRECT_OFFER.search(text)
            and RELEVANCE_CONTACT.search(text)
        )
        if structured_trade:
            return False
        return target_mentions >= 1 and trade_mentions >= 5
    # Less extreme dumps are excluded only when there is no attributable offer
    # or contact in the body. This preserves structured Telegram price lists.
    return (
        target_mentions >= 2
        and trade_mentions >= 8
        and not RELEVANCE_UNAMBIGUOUS_OFFER.search(text)
        and not RELEVANCE_STRONG_CONTACT.search(text)
    )


def looks_like_service_template_keyword_spam(title: str, text: str) -> bool:
    """Detect incoherent account-keyword injection into local service copy."""
    combined = title + "\n" + text[:5_000]
    template_hits = RELEVANCE_SEO_SPAM_TEMPLATE_PHRASE.findall(combined)
    target_hits = len(RELEVANCE_STRICT_TARGET.findall(combined)) + len(
        RELEVANCE_STRICT_SHORT_TARGET.findall(combined)
    )
    coherent_service_hits = RELEVANCE_COHERENT_ACCOUNT_SERVICE.findall(
        text[:5_000]
    )
    return (
        len(template_hits) >= 4
        and target_hits >= 3
        and len(coherent_service_hits) < 2
    )


def looks_like_search_spam(title: str, text: str) -> bool:
    """Detect SSR/search-poisoning copy without an attributable direct post."""
    combined = title + "\n" + text[:8_000]
    target_hits = len(RELEVANCE_STRICT_TARGET.findall(combined)) + len(
        RELEVANCE_STRICT_SHORT_TARGET.findall(combined)
    )
    trade_hits = len(RELEVANCE_TRADE.findall(combined))

    if re.search(
        r"(?:DB|디비)\s*(?:\([^)]*\))?\s*관련\s*(?:홍보|광고)"
        r".{0,120}(?:이용해|광고\s*대행|문의)",
        combined,
        re.IGNORECASE | re.DOTALL,
    ):
        return True

    quoted_blocks = [
        re.sub(r"\s+", " ", item).strip().lower()
        for item in re.findall(r'["“]([^"”]{18,240})["”]', text[:2_000])
    ]
    if (
        target_hits
        and trade_hits
        and max(Counter(quoted_blocks).values(), default=0) >= 3
    ):
        return True

    sentence_marks = len(re.findall(r"[.!?。？！]", text[:8_000]))
    if (
        len(text) >= 800
        and target_hits >= 20
        and trade_hits >= 15
        and sentence_marks <= 3
    ):
        return True

    seo_guide_hits = len(
        re.findall(
            r"Google\s*기준\s*상위\s*노출|SEO\s*기준|마스터\s*가이드|"
            r"키워드\s*통합\s*정리|관련\s*직접\s*경험한\s*내용을\s*정리|"
            r"검색\s*키워드들을?\s*구조적으로\s*포함|"
            r"정보가\s*너무\s*많아서\s*오히려\s*혼란",
            combined,
            re.IGNORECASE,
        )
    )
    if target_hits >= 6 and trade_hits >= 5 and seo_guide_hits >= 2:
        return True

    seo_agency_hits = len(
        re.findall(
            r"상위\s*노출|최상단\s*고정|상단\s*(?:고정|유지|자리)|"
            r"구글\s*(?:검색|영역|1\s*페이지)|광고비|광고주|"
            r"실행사|대행사|키워드|노출\s*(?:마케팅|서비스)",
            combined,
            re.IGNORECASE,
        )
    )
    # Some compromised boards contain an SEO agency's advert packed with
    # illicit-sale keywords.  The page is selling search placement, not the
    # DB/account named in those keywords, so it must not become a positive.
    if (
        target_hits >= 1
        and seo_agency_hits >= 5
        and re.search(r"광고비|광고주|실행사|대행사", combined, re.I)
        and re.search(
            r"상위\s*노출|최상단\s*고정|구글\s*(?:검색|1\s*페이지)",
            combined,
            re.I,
        )
    ):
        return True

    promotion_hits = len(
        re.findall(
            r"홍보|광고|상단\s*(?:노출|유지|고정|장악)|도배|광고주|"
            r"검색\s*결과\s*첫\s*페이지|문의\s*폭주\s*유도",
            combined,
            re.IGNORECASE,
        )
    )
    return target_hits >= 4 and promotion_hits >= 12


def looks_like_telegram_stub(title: str, text: str, page_type: str) -> bool:
    """Reject Telegram View/Contact shells unless their own copy states an offer."""
    if page_type != "public_messenger_page" or len(text) > 500:
        return False
    if not re.search(
        r"Telegram:\s*(?:View|Contact)|If\s+you\s+have\s+Telegram",
        title + "\n" + text,
        re.IGNORECASE,
    ):
        return False
    own_copy = re.sub(
        r"If\s+you\s+have\s+Telegram,?.*?(?:right\s+away\.)?",
        " ",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    explicit_offer = re.search(
        r"판매\s*(?:합니다|해요|중|업체)|팝니다|"
        r"매입\s*(?:합니다|해요|중|업체)|삽니다|"
        r"(?:DB|디비)\s*전문\s*판매\s*업체|"
        r"(?:계정|아이디|DB|디비|통장|계좌)\s*최저가|"
        r"다양한.{0,40}(?:계정|아이디)\s*판매|"
        r"(?:계좌|통장)\s*매입.{0,160}"
        r"(?:즉시\s*정산|상시\s*(?:모집|구인)|제휴\s*업체)|"
        r"(?:계정|아이디|DB|디비|계좌|통장).{0,40}"
        r"(?:전문으로\s*취급|전문\s*업체|모든\s*품목\s*취급|소량\s*/\s*대량\s*문의)|"
        r"실시간\s*(?:DB|디비).{0,40}문의|"
        r"(?:계좌|통장)\s*매입\s*안내\s*채널",
        own_copy,
        re.IGNORECASE | re.DOTALL,
    )
    target = RELEVANCE_STRICT_TARGET.search(own_copy) or (
        RELEVANCE_STRICT_SHORT_TARGET.search(own_copy)
    )
    return not bool(explicit_offer and target)


def discovery_candidate_relevant(candidate: Candidate) -> bool:
    text = candidate.discovery_text
    if not text:
        return False
    if RELEVANCE_EXCLUDED_TITLE.search(text[:1_000]):
        return False
    target = RELEVANCE_TARGET.search(text) or RELEVANCE_SHORT_TARGET.search(text)
    support = RELEVANCE_TRADE.search(text) or RELEVANCE_DISCOVERY_CONTACT.search(text)
    return bool(target and support)


def discovery_candidate_passes(candidate: Candidate, mode: str) -> bool:
    """Apply the requested search-snippet prefilter without final labeling."""
    if mode == "off":
        return True
    text = candidate.discovery_text
    if not text or RELEVANCE_EXCLUDED_TITLE.search(text[:1_000]):
        return False
    host = (urlsplit(candidate.url).hostname or "").lower()
    domain = registrable_domain(host)
    if domain in RELEVANCE_EXCLUDED_DOMAINS or domain.endswith(".wiki"):
        return False
    if mode in {"intent", "strict"}:
        parts = urlsplit(candidate.url)
        path_query = parts.path + "?" + parts.query
        if (
            domain in RELEVANCE_PRESS_DOMAINS
            or host.endswith((".ac.kr", ".go.kr"))
            or (
                RELEVANCE_DISCOVERY_PRESS_PATH.search(path_query)
                and RELEVANCE_DISCOVERY_PRESS_CONTEXT.search(text[:2_000])
            )
        ):
            return False
        lead = text[:1_500]
        if (
            RELEVANCE_NORMAL_PRODUCT_CONTEXT.search(lead)
            or RELEVANCE_NORMAL_ID_PRODUCT_CONTEXT.search(lead)
            or RELEVANCE_DRIVER_LICENSE_PHOTO_GUIDE.search(lead)
            or RELEVANCE_PUBLIC_BUSINESS_DIRECTORY_CONTEXT.search(lead)
            or RELEVANCE_NORMAL_TELECOM_SERVICE.search(lead)
            or RELEVANCE_SINGLE_ACCOUNT_CONTEXT.search(lead)
            or RELEVANCE_ACCOUNT_TOOL_CONTEXT.search(lead)
            or RELEVANCE_VIRTUAL_NUMBER_SERVICE.search(lead)
            or RELEVANCE_DB_MANAGEMENT_SOFTWARE.search(lead)
            or RELEVANCE_LEGAL_PROP_OR_SECURITY_CONTEXT.search(lead)
            or RELEVANCE_TELEGRAM_DIRECTORY_CONTEXT.search(lead)
        ):
            return False
        if (
            RELEVANCE_GIFT_CARD_ONLY.search(lead)
            and len(re.findall(r"상품권|문화상품권|기프티콘", lead, re.I)) >= 3
            and not RELEVANCE_CORE_PERSONAL_TARGET.search(lead)
        ):
            return False
        if (
            len(RELEVANCE_TRADE_GUIDE_PHRASE.findall(lead)) >= 2
            and not RELEVANCE_UNAMBIGUOUS_OFFER.search(lead)
        ):
            return False
        if (
            len(RELEVANCE_SOCIAL_METRIC_SERVICE.findall(lead)) >= 4
            and not RELEVANCE_ACCOUNT_OFFER.search(lead)
        ):
            return False
        if looks_like_service_template_keyword_spam(text[:500], lead):
            return False
        if (
            RELEVANCE_WARNING_AGAINST_TRADE.search(lead)
            and not RELEVANCE_UNAMBIGUOUS_OFFER.search(lead)
        ):
            return False
    if mode == "labeling":
        # Search snippets can omit the traded object even when the destination
        # title or body contains it. Keep trade-word candidates for annotation,
        # then apply the stricter document-level gate after extraction.
        return bool(LABELING_TARGET.search(text) or RELEVANCE_TRADE.search(text))
    if mode in {"intent", "strict"}:
        masked = mask_text(text[:2_000])
        # Search cards may concatenate unrelated result fragments. Require the
        # traded object and a direct offer in one local window, but defer the
        # concrete-contact requirement until the destination page is fetched.
        for start in range(0, len(masked), 180):
            window = masked[start : start + 360]
            if RELEVANCE_REPORTING_CONTEXT.search(window):
                continue
            target = RELEVANCE_STRICT_TARGET.search(
                window
            ) or RELEVANCE_STRICT_SHORT_TARGET.search(window)
            direct_offer = RELEVANCE_DIRECT_OFFER.search(window)
            negated_offer = RELEVANCE_NEGATED_OFFER.search(window)
            if (
                nearby_matches(target, direct_offer)
                and not nearby_matches(direct_offer, negated_offer)
            ):
                return True
            if mode == "intent":
                continue
            if (
                target
                and RELEVANCE_TRADE.search(window)
                and RELEVANCE_STRONG_CONTACT.search(window)
            ):
                return True
        return False
    return discovery_candidate_relevant(candidate)


def prefilter_seed_candidates(
    candidates: Iterable[Candidate], mode: str
) -> list[Candidate]:
    """Reuse discovery evidence when a prior candidate queue becomes a seed.

    Manually supplied URLs have no discovery text and must still be fetched.
    Prior search queues retain their snippets, so applying the requested gate
    avoids downloading every broad-review candidate again.
    """
    return [
        candidate
        for candidate in candidates
        if mode == "off"
        or (
            not candidate.discovery_text
            and candidate.source_type in {"seed", "manual"}
        )
        or (
            bool(candidate.discovery_text)
            and discovery_candidate_passes(candidate, mode)
        )
    ]


def normalized_expansion_query(query: str) -> str:
    """Normalize a query for deduplication without retaining search operators."""
    terms = [
        term.strip('"').lower()
        for term in query.split()
        if term and not term.startswith("-")
    ]
    return " ".join(terms)


def mine_keyword_expansions(
    candidates: Iterable[Candidate],
    known_specs: Iterable[QuerySpec],
    round_number: int,
    limit: int,
    minimum_domains: int,
    require_contact: bool = False,
) -> list[KeywordExpansion]:
    """Mine target/support pairs repeated across high-relevance search snippets."""
    known = {normalized_expansion_query(spec.query) for spec in known_specs}
    documents: defaultdict[str, set[str]] = defaultdict(set)
    domains: defaultdict[str, set[str]] = defaultdict(set)
    metadata: defaultdict[str, Counter[tuple[str, str]]] = defaultdict(Counter)
    display_queries: dict[str, str] = {}
    contact_queries: set[str] = set()

    for candidate in candidates:
        if not discovery_candidate_relevant(candidate):
            continue
        text = mask_text(candidate.discovery_text[:2_000])
        domain = registrable_domain(urlsplit(candidate.url).hostname or "")
        for start in range(0, len(text), 120):
            window = text[start : start + 240]
            targets = {
                normalize_extracted_text(match.group(0))
                for match in RELEVANCE_SHORT_TARGET.finditer(window)
            }
            contact_terms = {
                normalize_extracted_text(match.group(0))
                for match in EXPANSION_CONTACT_TERM.finditer(window)
            }
            trade_terms = {
                normalize_extracted_text(match.group(0))
                for match in RELEVANCE_TRADE.finditer(window)
            }
            support_terms = (
                contact_terms if require_contact else contact_terms | trade_terms
            )
            if not targets or not support_terms:
                continue
            for target in targets:
                for support in support_terms:
                    query = f"{target} {support}".strip()
                    key = normalized_expansion_query(query)
                    if not key or key in known or len(query) > 80:
                        continue
                    documents[key].add(candidate.url)
                    if domain:
                        domains[key].add(domain)
                    metadata[key][
                        (candidate.query_group, candidate.detection_type)
                    ] += 1
                    display_queries.setdefault(key, query)
                    if support in contact_terms:
                        contact_queries.add(key)

    eligible = [
        key for key in documents if len(domains[key]) >= minimum_domains
    ]
    eligible.sort(
        key=lambda key: (
            key in contact_queries,
            len(domains[key]),
            len(documents[key]),
            -len(display_queries[key]),
            display_queries[key],
        ),
        reverse=True,
    )
    expansions: list[KeywordExpansion] = []
    for key in eligible[:limit]:
        query_group, detection_type = metadata[key].most_common(1)[0][0]
        expansions.append(
            KeywordExpansion(
                round_number=round_number,
                query_group=query_group,
                detection_type=detection_type,
                query=display_queries[key],
                document_frequency=len(documents[key]),
                domain_frequency=len(domains[key]),
            )
        )
    return expansions


def merge_candidates(
    current: Iterable[Candidate], additions: Iterable[Candidate]
) -> list[Candidate]:
    """Merge search candidates while preserving all available snippet evidence."""
    merged: dict[str, Candidate] = {}
    for candidate in [*current, *additions]:
        existing = merged.get(candidate.url)
        if existing is None:
            merged[candidate.url] = candidate
            continue
        if (
            candidate.discovery_text
            and candidate.discovery_text not in existing.discovery_text
        ):
            existing.discovery_text = normalize_extracted_text(
                existing.discovery_text + "\n" + candidate.discovery_text
            )[:2_000]
        existing_providers = {
            name for name in existing.search_provider.split(",") if name
        }
        additional_providers = {
            name for name in candidate.search_provider.split(",") if name
        }
        if additional_providers - existing_providers:
            existing.search_provider = ",".join(
                sorted(existing_providers | additional_providers)
            )
    return list(merged.values())


def discovery_relevance_score(candidate: Candidate) -> int:
    """Rank search-result documents without assigning a final label."""
    text = candidate.discovery_text[:2_000]
    if not text:
        return -100
    score = 0
    target_count = min(4, len(RELEVANCE_TARGET.findall(text)))
    trade_count = min(4, len(RELEVANCE_TRADE.findall(text)))
    contact_count = min(3, len(RELEVANCE_CONTACT.findall(text)))
    score += target_count * 5 + trade_count * 3 + contact_count * 2
    if target_count and (trade_count or contact_count):
        score += 10
    for start in range(0, len(text), 180):
        window = text[start : start + 360]
        target = RELEVANCE_STRICT_TARGET.search(
            window
        ) or RELEVANCE_STRICT_SHORT_TARGET.search(window)
        direct_offer = RELEVANCE_DIRECT_OFFER.search(window)
        if nearby_matches(target, direct_offer):
            score += 30
            break
    if RELEVANCE_REPORTING_CONTEXT.search(text):
        score -= 50
    if RELEVANCE_QUESTION_OR_GUIDE_TITLE.search(text[:500]):
        score -= 30
    if RELEVANCE_NORMAL_PRODUCT_CONTEXT.search(text[:1_000]):
        score -= 30
    if RELEVANCE_SINGLE_ACCOUNT_CONTEXT.search(text[:1_000]):
        score -= 30
    if RELEVANCE_EXCLUDED_TITLE.search(text):
        score -= 20
    host = (urlsplit(candidate.url).hostname or "").lower()
    domain = registrable_domain(host)
    if domain in RELEVANCE_EXCLUDED_DOMAINS or domain.endswith(".wiki"):
        score -= 20
    return score


def relevance_gate_reason(
    title: str,
    text: str,
    url: str,
    page_type: str,
    mode: str,
    discovery_text: str = "",
) -> str:
    """Return an exclusion reason, or an empty string for a retained candidate."""
    if page_type in {
        "search_result_list",
        "search_reflection",
        "deleted_or_inaccessible",
        "empty_container",
        "board_listing",
    }:
        return "excluded_page_type"
    # Structural exclusions are always active. ``off`` disables topical
    # filtering only; it must not turn search-result pages into samples.
    if mode == "off":
        return ""
    precision_mode = mode in {"intent", "strict"}
    if precision_mode and page_type == "news_or_education":
        return "excluded_page_type"
    host = (urlsplit(url).hostname or "").lower()
    domain = registrable_domain(host)
    if domain in RELEVANCE_EXCLUDED_DOMAINS or domain.endswith(".wiki"):
        return "excluded_domain"
    if precision_mode and domain in RELEVANCE_PRESS_DOMAINS:
        return "excluded_press_domain"
    if precision_mode and host.endswith((".ac.kr", ".go.kr")):
        return "excluded_institutional_domain"
    title_and_lead = title + "\n" + text[:800]
    buying_inquiry = bool(
        RELEVANCE_BUYING_INQUIRY.search(title_and_lead)
        and not re.search(
            r"고소|신고|사기|피해|처벌|불법인지|문제되|범죄",
            title_and_lead,
            re.IGNORECASE,
        )
    )
    explicit_buying_inquiry = bool(
        buying_inquiry
        and RELEVANCE_BUYING_INQUIRY_QUESTION.search(title_and_lead)
    )
    # These are title indicators, not arbitrary body stop words. Genuine
    # channels commonly contain "바로가기" or "위키" in nearby link text.
    if RELEVANCE_EXCLUDED_TITLE.search(title):
        return "excluded_document_type"

    if precision_mode and RELEVANCE_EMPTY_CONTAINER_CONTEXT.search(text[:2_000]):
        return "excluded_empty_container"
    if precision_mode and RELEVANCE_EMPTY_LISTING_TEMPLATE.search(text[:2_500]):
        return "excluded_empty_listing_template"
    if precision_mode and looks_like_telegram_stub(title, text, page_type):
        return "excluded_empty_container"
    if precision_mode and RELEVANCE_PUBLIC_BUSINESS_DIRECTORY_CONTEXT.search(
        title + "\n" + text[:1_800]
    ):
        return "excluded_public_business_directory"
    if precision_mode and RELEVANCE_NORMAL_TELECOM_SERVICE.search(
        title + "\n" + text[:2_000]
    ):
        return "excluded_normal_telecom_service"
    if precision_mode and RELEVANCE_NORMAL_ID_PRODUCT_CONTEXT.search(
        title + "\n" + text[:2_500]
    ):
        return "excluded_normal_product_context"
    if precision_mode and RELEVANCE_PASSPORT_PHOTO_GUIDE.search(
        title + "\n" + text[:2_500]
    ):
        return "excluded_identity_photo_guide"
    if precision_mode and RELEVANCE_LICENSE_REQUIREMENTS_GUIDE.search(
        title + "\n" + text[:3_000]
    ):
        return "excluded_identity_document_guide"
    if precision_mode and RELEVANCE_DB_BRAND_OR_STOCK.search(
        title + "\n" + text[:2_500]
    ):
        return "excluded_db_brand_or_stock"
    if precision_mode and RELEVANCE_DB_PC_JOB_CONTEXT.search(
        title + "\n" + text[:2_500]
    ):
        return "excluded_db_job_context"
    if precision_mode and RELEVANCE_NORMAL_ID_CARD_MARKET.search(
        title + "\n" + text[:3_000]
    ):
        return "excluded_normal_product_context"
    if precision_mode and RELEVANCE_DOCUMENTARY_CASE_CONTEXT.search(
        title + "\n" + text[:4_000]
    ):
        return "excluded_reporting_context"
    if precision_mode and RELEVANCE_INSURANCE_INDUSTRY_ANALYSIS.search(
        title + "\n" + text[:4_000]
    ):
        return "excluded_insurance_industry_analysis"
    if precision_mode and RELEVANCE_DB_IT_SYSTEM_CONTEXT.search(
        title + "\n" + text[:3_000]
    ):
        return "excluded_it_database_system"
    if precision_mode and RELEVANCE_DERIVATIVES_TRADING_SERVICE.search(
        title + "\n" + text[:3_000]
    ):
        return "excluded_investment_trading_service"
    if (
        precision_mode
        and len(
            RELEVANCE_LONGFORM_REPORT_PHRASE.findall(
                title + "\n" + text[:4_000]
            )
        )
        >= 4
    ):
        return "excluded_informational_report"
    if (
        precision_mode
        and RELEVANCE_INSURANCE_RECRUITMENT.search(
            title + "\n" + text[:4_000]
        )
        and RELEVANCE_RECRUITMENT_OR_ORG_OFFER.search(
            title + "\n" + text[:4_000]
        )
        and RELEVANCE_RECRUITMENT_DB_BENEFIT.search(
            title + "\n" + text[:4_000]
        )
    ):
        return "excluded_insurance_recruitment"
    if precision_mode and RELEVANCE_DB_PURCHASE_ALTERNATIVE.search(
        title + "\n" + text[:4_000]
    ):
        return "excluded_db_purchase_alternative"
    if (
        precision_mode
        and RELEVANCE_DB_COMPLIANCE_GUIDE_STRUCTURE.search(
            title + "\n" + text[:4_000]
        )
        and len(
            RELEVANCE_DB_COMPLIANCE_GUIDE_PHRASE.findall(
                title + "\n" + text[:4_000]
            )
        )
        >= 1
    ):
        return "excluded_db_compliance_guide"
    if precision_mode and RELEVANCE_EVENT_TICKET_CONTEXT.search(
        title + "\n" + text[:2_500]
    ):
        return "excluded_event_ticket_context"
    if precision_mode and RELEVANCE_DRIVER_LICENSE_PHOTO_GUIDE.search(
        title + "\n" + text[:2_000]
    ):
        return "excluded_question_or_guide"
    if precision_mode and RELEVANCE_LEGAL_PROP_OR_SECURITY_CONTEXT.search(
        title + "\n" + text[:3_000]
    ):
        return "excluded_legal_prop_or_security_guide"
    if precision_mode and RELEVANCE_TELEGRAM_DIRECTORY_CONTEXT.search(
        title + "\n" + text[:2_000]
    ):
        return "excluded_telegram_directory"
    if precision_mode and RELEVANCE_CORPORATE_TRANSFER_CONTEXT.search(
        title_and_lead
    ):
        return "excluded_corporate_transfer"
    if precision_mode and RELEVANCE_AGGREGATION_OR_COMMENTARY_CONTEXT.search(
        title_and_lead
    ):
        return "excluded_aggregation_or_commentary"
    warning_match = RELEVANCE_WARNING_AGAINST_TRADE.search(
        title + "\n" + text[:2_500]
    )
    body_offer_matches = list(
        RELEVANCE_UNAMBIGUOUS_OFFER.finditer(text[:2_500])
    )
    if (
        precision_mode
        and warning_match
        and (
            not body_offer_matches
            or (
                len(body_offer_matches) == 1
                and body_offer_matches[0].start() < 200
            )
        )
    ):
        return "excluded_trade_warning"
    if (
        precision_mode
        and RELEVANCE_TRADE_MOTIVE_QUESTION.search(title_and_lead)
        and not RELEVANCE_UNAMBIGUOUS_OFFER.search(title_and_lead)
    ):
        return "excluded_trade_motive_question"
    if precision_mode and RELEVANCE_ACCOUNT_TOOL_CONTEXT.search(
        title + "\n" + text[:2_000]
    ):
        return "excluded_account_creation_tool"
    if precision_mode and RELEVANCE_VIRTUAL_NUMBER_SERVICE.search(
        title + "\n" + text[:2_000]
    ):
        return "excluded_virtual_number_service"
    if precision_mode and RELEVANCE_DB_MANAGEMENT_SOFTWARE.search(
        title + "\n" + text[:2_500]
    ):
        return "excluded_db_management_software"
    if (
        precision_mode
        and len(
            RELEVANCE_NEGATED_DB_TRADE_CONTEXT.findall(
                title + "\n" + text[:3_000]
            )
        )
        >= 2
        and not RELEVANCE_UNAMBIGUOUS_OFFER.search(text[:3_000])
    ):
        return "excluded_negated_trade"
    if precision_mode and looks_like_keyword_stuffing(title, text[:8_000]):
        return "excluded_keyword_stuffing"
    if precision_mode and looks_like_service_template_keyword_spam(title, text):
        return "excluded_keyword_stuffing"
    if (
        precision_mode
        and page_type != "public_messenger_page"
        and looks_like_search_spam(title, text)
    ):
        return "excluded_search_spam"

    if precision_mode:
        commodity_context = title + "\n" + text[:2_500]
        if (
            RELEVANCE_GIFT_CARD_ONLY.search(commodity_context)
            and len(
                re.findall(
                    r"상품권|문화상품권|기프티콘",
                    commodity_context,
                    re.IGNORECASE,
                )
            )
            >= 3
            and not RELEVANCE_CORE_PERSONAL_TARGET.search(commodity_context)
        ):
            return "excluded_normal_product_context"
        guide_hits = RELEVANCE_TRADE_GUIDE_PHRASE.findall(
            title + "\n" + text[:4_000]
        )
        if (
            len(guide_hits) >= 2
            and not RELEVANCE_UNAMBIGUOUS_OFFER.search(text[:4_000])
        ):
            return "excluded_informational_article"
        metric_hits = RELEVANCE_SOCIAL_METRIC_SERVICE.findall(text[:4_000])
        if len(metric_hits) >= 4 and not RELEVANCE_ACCOUNT_OFFER.search(text[:4_000]):
            return "excluded_normal_product_context"

    if precision_mode:
        explainer_context = title + "\n" + text[:3_000]
        explainer_phrases = RELEVANCE_SEO_EXPLAINER_PHRASE.findall(
            explainer_context
        )
        if (
            (
                len(explainer_phrases) >= 5
                and not RELEVANCE_TITLE_TRADE.search(title)
                and not RELEVANCE_ATTRIBUTABLE_OFFER_VERB.search(text[:3_000])
            )
            or len(explainer_phrases) >= 3
            and not RELEVANCE_UNAMBIGUOUS_OFFER.search(text[:3_000])
        ):
            return "excluded_seo_explainer"
        informational_phrases = RELEVANCE_INFORMATIONAL_ARTICLE_PHRASE.findall(
            explainer_context
        )
        if (
            len(informational_phrases) >= 3
            or (
                len(informational_phrases) >= 2
                and not RELEVANCE_UNAMBIGUOUS_OFFER.search(explainer_context)
            )
        ):
            return "excluded_informational_article"

    body_head = text[:700]
    boilerplate = RELEVANCE_BODY_BOILERPLATE.search(body_head)
    if precision_mode and boilerplate:
        body_target = RELEVANCE_STRICT_TARGET.search(
            body_head
        ) or RELEVANCE_STRICT_SHORT_TARGET.search(body_head)
        body_offer = RELEVANCE_DIRECT_OFFER.search(body_head)
        if boilerplate.start() < 100 or not nearby_matches(body_target, body_offer):
            return "excluded_extraction_boilerplate"
    if precision_mode:
        guide_context = title + "\n" + text[:1_500]
        if (
            RELEVANCE_MARKET_GUIDE_CONTEXT.search(guide_context)
            or len(RELEVANCE_MARKET_GUIDE_WEAK.findall(guide_context)) >= 2
        ):
            return "excluded_market_guide"

    if (
        precision_mode
        and not buying_inquiry
        and RELEVANCE_REPORTING_CONTEXT.search(title_and_lead)
    ):
        return "excluded_reporting_context"
    if (
        precision_mode
        and len(
            RELEVANCE_LEGAL_DECISION_CONTEXT.findall(
                title + "\n" + text[:3_000]
            )
        )
        >= 3
    ):
        return "excluded_legal_decision"
    if (
        precision_mode
        and not buying_inquiry
        and RELEVANCE_QUESTION_OR_GUIDE_TITLE.search(title)
    ):
        return "excluded_question_or_guide"
    if precision_mode and RELEVANCE_AGGREGATION_OR_COMMENTARY_CONTEXT.search(
        title_and_lead
    ):
        return "excluded_aggregation_or_commentary"
    if precision_mode and RELEVANCE_NORMAL_PRODUCT_CONTEXT.search(title_and_lead):
        return "excluded_normal_product_context"
    if precision_mode and RELEVANCE_SINGLE_ACCOUNT_CONTEXT.search(title_and_lead):
        return "excluded_single_account_trade"

    title_target = RELEVANCE_STRICT_TARGET.search(
        title
    ) or RELEVANCE_STRICT_SHORT_TARGET.search(title)
    title_offer = RELEVANCE_DIRECT_OFFER.search(title)
    body_target = RELEVANCE_STRICT_TARGET.search(
        text[:2_000]
    ) or RELEVANCE_STRICT_SHORT_TARGET.search(text[:2_000])
    body_trade = RELEVANCE_TRADE.search(text[:2_000])
    if (
        precision_mode
        and title_target
        and title_offer
        and not body_target
        and not body_trade
        and not RELEVANCE_ATTRIBUTABLE_BODY_CONTACT.search(text[:2_000])
        and len(RELEVANCE_LISTING_DETAIL.findall(text[:2_000])) < 2
    ):
        return "excluded_title_body_mismatch"

    # A person asking what price to pay for a relevant DB/account is expressing
    # purchase intent even when the post contains no seller handle.  Keep this
    # after the news, guide, normal-product and game-account exclusions so that
    # a question mark alone cannot turn those documents into positives.
    if precision_mode and explicit_buying_inquiry:
        inquiry_target = RELEVANCE_STRICT_TARGET.search(
            title_and_lead
        ) or RELEVANCE_STRICT_SHORT_TARGET.search(title_and_lead)
        if inquiry_target:
            return ""

    if mode == "labeling":
        # Keep topical positives and hard negatives for human annotation. The
        # stronger review/strict modes remain available for precision-first runs.
        title_has_target = bool(LABELING_TARGET.search(title))
        title_has_trade = bool(RELEVANCE_TITLE_TRADE.search(title))
        title_has_contact = bool(RELEVANCE_CONTACT.search(title))
        lead_has_target = bool(LABELING_TARGET.search(text[:1_000]))
        if title_has_target or title_has_trade:
            return ""
        if title_has_contact and lead_has_target:
            return ""
        return "missing_relevant_target"

    title_has_target = bool(
        RELEVANCE_TARGET.search(title) or RELEVANCE_SHORT_TARGET.search(title)
    )
    title_has_trade = bool(RELEVANCE_TITLE_TRADE.search(title))
    title_has_contact = bool(RELEVANCE_CONTACT.search(title))
    if not precision_mode and (
        not title_has_target or not (title_has_trade or title_has_contact)
    ):
        return "missing_title_signal"

    combined = title + "\n" + text
    best_signals: set[str] = set()
    weak_complete_window = False
    # Signals must occur in the same local window. This prevents a long generic
    # privacy policy from passing because unrelated words occur far apart.
    for start in range(0, len(combined), 250):
        window = combined[start : start + 500]
        signals = set()
        target_pattern = (
            RELEVANCE_STRICT_TARGET if precision_mode else RELEVANCE_TARGET
        )
        short_target_pattern = (
            RELEVANCE_STRICT_SHORT_TARGET
            if precision_mode
            else RELEVANCE_SHORT_TARGET
        )
        target = target_pattern.search(window) or short_target_pattern.search(window)
        if target:
            signals.add("target")
        if RELEVANCE_TRADE.search(window):
            signals.add("trade")
        contact_pattern = (
            RELEVANCE_STRONG_CONTACT if mode == "strict" else RELEVANCE_CONTACT
        )
        if contact_pattern.search(window):
            signals.add("contact")
        if len(signals) > len(best_signals):
            best_signals = signals
        if mode == "intent" and "target" in signals:
            direct_offer = RELEVANCE_DIRECT_OFFER.search(window)
            negated_offer = RELEVANCE_NEGATED_OFFER.search(window)
            if (
                nearby_matches(target, direct_offer)
                and not nearby_matches(direct_offer, negated_offer)
            ):
                return ""
            weak_offer = RELEVANCE_WEAK_OFFER.search(window)
            strong_contact = RELEVANCE_STRONG_CONTACT.search(window)
            if (
                nearby_matches(target, weak_offer)
                and strong_contact
                and not nearby_matches(weak_offer, negated_offer)
            ):
                return ""
        if mode == "strict" and signals == {"target", "trade", "contact"}:
            direct_offer = RELEVANCE_DIRECT_OFFER.search(window)
            negated_offer = RELEVANCE_NEGATED_OFFER.search(window)
            if (
                nearby_matches(target, direct_offer)
                and not nearby_matches(direct_offer, negated_offer)
            ):
                return ""
            weak_complete_window = True
        if (
            mode == "review"
            and "target" in signals
            and len(signals) >= 2
            and title_has_target
            and (title_has_trade or title_has_contact)
        ):
            return ""
    if "target" not in best_signals:
        return "missing_relevant_target"
    if mode == "intent":
        return "missing_body_offer"
    if mode == "strict":
        # The destination body must independently show the traded object and
        # a direct offer. Search-result text may supply a contact omitted by
        # extraction, but must never make an unrelated landing page pass.
        body_has_offer = False
        for start in range(0, len(combined), 250):
            window = combined[start : start + 500]
            target = RELEVANCE_STRICT_TARGET.search(
                window
            ) or RELEVANCE_STRICT_SHORT_TARGET.search(window)
            direct_offer = RELEVANCE_DIRECT_OFFER.search(window)
            negated_offer = RELEVANCE_NEGATED_OFFER.search(window)
            if (
                target
                and RELEVANCE_TRADE.search(window)
                and nearby_matches(target, direct_offer)
                and not nearby_matches(direct_offer, negated_offer)
            ):
                body_has_offer = True
                break
        if not body_has_offer:
            return "missing_body_offer"
        if discovery_text:
            discovery = mask_text(discovery_text[:2_000])
            for start in range(0, len(discovery), 180):
                window = discovery[start : start + 360]
                target = RELEVANCE_STRICT_TARGET.search(
                    window
                ) or RELEVANCE_STRICT_SHORT_TARGET.search(window)
                if target and RELEVANCE_STRONG_CONTACT.search(window):
                    return ""
        if "contact" not in best_signals:
            return "missing_concrete_contact"
        if weak_complete_window:
            return "missing_direct_offer"
        return "missing_trade_or_contact_signal"
    return "missing_supporting_signal"


def discover_related_internal_links(
    html: str, base_url: str, limit: int
) -> list[str]:
    if limit <= 0:
        return []
    base = urlsplit(base_url)
    base_host = base.hostname or ""
    base_domain = registrable_domain(base_host)
    base_parent = base.path.rsplit("/", 1)[0]
    topic_terms = re.compile(
        r"개인정보|고객|DB|디비|계정|ID|아이디|판매|구매|"
        r"텔레그램|텔그|오픈채팅|문의|회원|여권|통장|실명",
        re.IGNORECASE,
    )
    content_path = re.compile(
        r"/(?:board|boards|bbs|post|posts|article|articles|view|read|detail|entry|blog)/|"
        r"(?:^|[/_-])(?:board|bbs|post|article|view|read|detail|idx)(?:[/_.?=&-]|$)",
        re.IGNORECASE,
    )
    blocked_path = re.compile(
        r"/(?:login|logout|signup|register|account|admin|download|attachment|file)/?",
        re.IGNORECASE,
    )
    scored: dict[str, int] = {}
    soup = BeautifulSoup(html, "lxml")
    for anchor in soup.select("a[href]"):
        raw = str(anchor.get("href") or "").strip()
        if not raw or raw.startswith(("#", "javascript:", "mailto:", "tel:")):
            continue
        url = canonicalize_url(urljoin(base_url, raw))
        if not url or url == base_url:
            continue
        parts = urlsplit(url)
        if registrable_domain(parts.hostname or "") != base_domain:
            continue
        if blocked_path.search(parts.path):
            continue
        params = {name.lower() for name, _ in parse_qsl(parts.query)}
        if params & {"q", "query", "keyword", "search", "searchword"}:
            continue
        anchor_text = anchor.get_text(" ", strip=True)
        score = 0
        if topic_terms.search(anchor_text):
            score += 3
        if topic_terms.search(parts.path):
            score += 2
        if content_path.search(parts.path):
            score += 2
        if base_parent and base_parent != "/" and parts.path.startswith(base_parent):
            score += 1
        # Same-directory and generic content-path matches alone are not topical.
        # Require an explicit topic term in the anchor text or target path.
        if score > 0 and (
            topic_terms.search(anchor_text) or topic_terms.search(parts.path)
        ):
            scored[url] = max(scored.get(url, 0), score)
    return [
        url
        for url, _ in sorted(scored.items(), key=lambda item: (-item[1], item[0]))[
            :limit
        ]
    ]


MESSENGER_CONTACT_PATTERNS = (
    re.compile(
        r"(?i)(?:https?://)?(?:t\.me|telegram\.me)/[A-Za-z0-9_.+/?=&-]{2,}"
    ),
    re.compile(
        r"(?i)(?:https?://)?(?:open\.kakao\.com|pf\.kakao\.com)/"
        r"[A-Za-z0-9_./?=&+-]+"
    ),
    re.compile(r"(?i)(?:https?://)?line\.me/[A-Za-z0-9_./?=&+-]+"),
    re.compile(
        r"(?i)(?:텔레그램|telegram|텔그|텔레|카카오톡|카카오|카톡|오픈채팅|라인|line)"
        r"\s*(?:아이디|id|주소|문의|연락|[:：])?\s*[@:]?\s*"
        r"[A-Za-z0-9_.-]{3,}"
    ),
    re.compile(r"(?:ㅌㄹㄱ?|ㅌ그)\s*[:：]?\s*[A-Za-z0-9_.-]{3,}"),
)


def mask_text(value: str, preserve_messenger_ids: bool = False) -> str:
    if not value:
        return ""
    text = value
    protected_messenger_contacts: list[str] = []
    if preserve_messenger_ids:
        def protect_messenger(match: re.Match[str]) -> str:
            protected_messenger_contacts.append(match.group(0))
            return f"\ufff0MSG{len(protected_messenger_contacts) - 1}\ufff1"

        for pattern in MESSENGER_CONTACT_PATTERNS:
            text = pattern.sub(protect_messenger, text)
    # Preserve the existence and channel of direct messenger contacts while
    # removing the handle itself. Generic URLs are masked separately below.
    text = re.sub(
        r"(?i)(?:https?://)?(?:t\.me|telegram\.me)/[A-Za-z0-9_.+-]{2,}",
        "텔레그램 [MESSENGER_ID]",
        text,
    )
    text = re.sub(
        r"(?i)(?:https?://)?(?:open\.kakao\.com|pf\.kakao\.com)/[A-Za-z0-9_./?=&+-]+",
        "카카오톡 [MESSENGER_ID]",
        text,
    )
    text = re.sub(
        r"(?i)(?:https?://)?line\.me/[A-Za-z0-9_./?=&+-]+",
        "라인 [MESSENGER_ID]",
        text,
    )
    # Order matters: remaining URLs and emails must be removed before generic IDs.
    text = re.sub(
        r"(?i)(?:https?://|www\.)[^\s<>\"']+",
        "[CONTACT_URL]",
        text,
    )
    text = re.sub(r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", "[EMAIL]", text)
    text = re.sub(r"(?<!\d)(?:\d{6}\s*[-–]?\s*[1-4]\d{6})(?!\d)", "[NATIONAL_ID]", text)
    text = re.sub(
        r"(?<!\d)(?:01[016789]|02|0[3-6][1-5])[- .]?\d{3,4}[- .]?\d{4}(?!\d)",
        "[PHONE]",
        text,
    )
    text = re.sub(
        r"(?<!\d)(?:\+?82[- .]?)?1[016789][- .]?\d{3,4}[- .]?\d{4}(?!\d)",
        "[PHONE]",
        text,
    )
    text = re.sub(
        r"(?i)(?<![\w.])(?:\d{1,3}\.){3}\d{1,3}(?![\w.])", "[IP_ADDRESS]", text
    )
    text = re.sub(
        r"(?i)(텔레그램|telegram|텔그|텔레|카카오톡|카카오|카톡|오픈채팅|라인|line|ㅌㄹㄱ?|ㅌ그)"
        r"\s*(?:아이디|id|주소|문의|[:：])?\s*[@:]?\s*[A-Za-z0-9_.-]{3,}",
        lambda m: m.group(1) + " [MESSENGER_ID]",
        text,
    )
    text = re.sub(r"(?<!\w)@[A-Za-z0-9_]{3,}(?!\w)", "[ACCOUNT]", text)
    text = re.sub(
        r"(?i)(아이디|ID|계정명|닉네임|사용자명)\s*[:：=]\s*[A-Za-z0-9_.-]{2,}",
        lambda m: m.group(1) + ": [ACCOUNT]",
        text,
    )
    text = re.sub(
        r"(이름|성명|예금주)\s*[:：=]\s*[가-힣]{2,4}",
        lambda m: m.group(1) + ": [NAME]",
        text,
    )
    text = re.sub(
        r"(계좌(?:번호)?|입금계좌)\s*[:：=]?\s*(?:\d[- ]?){9,16}\d",
        lambda m: m.group(1) + ": [BANK_ACCOUNT]",
        text,
    )
    # Any remaining long digit sequence is not useful research content.
    text = re.sub(r"(?<!\d)\d{10,16}(?!\d)", "[NUMERIC_IDENTIFIER]", text)
    for index, contact in enumerate(protected_messenger_contacts):
        text = text.replace(f"\ufff0MSG{index}\ufff1", contact)
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = text.strip()
    # CSV quoting alone does not stop spreadsheet applications from evaluating cells.
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def make_restricted_review_record(
    record: dict[str, object],
    final_url: str,
    title: str,
    text: str,
) -> dict[str, str]:
    """Keep messenger handles only in the access-controlled review copy."""
    return {
        "sample_id": str(record["sample_id"]),
        "collected_at": str(record["collected_at"]),
        "source_url": final_url,
        "registrable_domain": str(record["registrable_domain"]),
        "title": mask_text(title, preserve_messenger_ids=True),
        "text": mask_text(text, preserve_messenger_ids=True),
    }


def language_mix(text: str) -> str:
    has_ko = bool(re.search(r"[가-힣]", text))
    has_en = bool(re.search(r"[A-Za-z]", text))
    if has_ko and has_en:
        return "ko_en_mixed"
    if has_ko:
        return "ko"
    if has_en:
        return "en"
    return "other"


def obfuscation_type(text: str) -> str:
    flags = []
    if re.search(r"[ㅏ-ㅣㄱ-ㅎ]{2,}", text):
        flags.append("jamo")
    if re.search(r"\w[._*·-]\w[._*·-]\w", text):
        flags.append("symbol_separated")
    if re.search(r"[A-Za-z].*[가-힣]|[가-힣].*[A-Za-z]", text):
        flags.append("mixed_script")
    return "+".join(flags) if flags else "none"


def classify_page_type(url: str, title: str, text: str) -> str:
    parts = urlsplit(url)
    host = (parts.hostname or "").lower()
    path_query = (parts.path + "?" + parts.query).lower()
    combined = (title + "\n" + text[:2_000]).lower()
    domain = registrable_domain(host)
    if re.search(
        r"(?:^|[/_?&=-])(search|query|keyword|find|input|results?)"
        r"(?:[/_?&=-]|$)",
        path_query,
    ):
        params = parse_qs(parts.query)
        reflected_values = [
            value.lower().strip()
            for key, values in params.items()
            if key.lower()
            in {
                "q",
                "query",
                "keyword",
                "search",
                "searchword",
                "search_query",
                "s",
                "i",
                "text",
                "term",
                "wd",
            }
            for value in values
            if len(value.strip()) >= 4
        ]
        if any(value in combined for value in reflected_values):
            return "search_reflection"
        return "search_result_list"
    if re.search(
        r"번호\s*제목\s*작성자\s*작성일(?:\s*(?:추천|조회))*",
        text[:1_500],
        re.IGNORECASE,
    ):
        return "search_result_list"
    compact_listing_rows = re.findall(
        r"(?:^|\n)\s*\d{4,6}[^\n]{0,180}20\d{2}[-./]\d{1,2}[-./]\d{1,2}"
        r"[^\n]{0,80}(?:조회|추천)\s*\d+",
        text[:6_000],
        re.IGNORECASE,
    )
    if len(compact_listing_rows) >= 3:
        return "board_listing"
    generic_board_title = re.fullmatch(
        r"(?:갤러리|자유\s*게시판|게시판|커뮤니티|목록)(?:\s*[-|:]\s*[^\n]+)?",
        normalize_extracted_text(title),
        re.IGNORECASE,
    )
    board_list_controls = re.search(
        r"이미지형\s*리스트형|게시물\s*검색|"
        r"(?:제목|내용)\s*(?:글쓴이|작성자)\s*(?:아이디|별명|작성일)",
        text[:4_000],
        re.IGNORECASE,
    )
    listing_total = re.search(
            r"전체\s*[\d,]+\s*건\s*/\s*[\d,]+\s*페이지",
            text[:5_000],
        )
    marketplace_listing = bool(
        listing_total
        and (
            len(
            re.findall(
                r"쪽지보내기\s*메일보내기\s*자기소개\s*아이디로\s*검색",
                text[:8_000],
                re.IGNORECASE,
            )
            )
            >= 2
            or (
                "게시물 검색" in text[:8_000]
                and len(
                    re.findall(
                        r"판매중|거래완료|(?:^|\n)\s*팝니다\s*(?:\n|$)",
                        text[:8_000],
                        re.IGNORECASE,
                    )
                )
                >= 3
            )
        )
    )
    category_path_segments = [segment for segment in parts.path.split("/") if segment]
    generic_market_category = bool(
        listing_total
        and len(category_path_segments) == 1
        and re.search(
            r"(?:계정|채널|웹사이트)\s*거래\s*(?:-|$)",
            title,
            re.IGNORECASE,
        )
    )
    if marketplace_listing or generic_market_category:
        return "board_listing"
    timestamp_rows = len(
        re.findall(
            r"20\d{2}[-./]\d{1,2}[-./]\d{1,2}\s+\d{1,2}:\d{2}",
            text[:5_000],
        )
    )
    if generic_board_title and (board_list_controls or timestamp_rows >= 3):
        return "board_listing"
    if (
        re.fullmatch(r"네이버\s*인플루언서", normalize_extracted_text(title), re.I)
        and len(re.findall(r"조회수\s*[\d,]+", text[:4_000])) >= 3
    ):
        return "board_listing"
    if RELEVANCE_EMPTY_CONTAINER_CONTEXT.search(text[:2_000]):
        return "empty_container"
    if any(
        term in combined
        for term in (
            "삭제된 게시물",
            "존재하지 않는 게시물",
            "페이지를 찾을 수 없습니다",
        )
    ):
        return "deleted_or_inaccessible"
    document_lead = (title + "\n" + text[:600]).lower()
    if any(
        term in document_lead
        for term in ("보도자료", "교육자료", "예방 수칙", "피해 사례")
    ):
        return "news_or_education"
    hard_press_path = re.search(
        r"(?:^|/)news/(?:articleview(?:\.html)?|view)(?:/|\.|$)|"
        r"(?:^|/)news/articleview\.html(?:\?|$)",
        path_query,
        re.IGNORECASE,
    )
    if hard_press_path:
        return "news_or_education"
    press_path = re.search(
        r"(?:^|/)(?:news(?:[_-](?:list|room))?|press(?:[_-]release)?|"
        r"article|articles|view|read)(?:/|\.|$)|"
        r"(?:articleview|newsid|arcid)=",
        path_query,
    )
    press_lead = re.search(
        r"(?:\[[^\]]+\]\s*)?(?:=\s*)?[가-힣]{2,4}\s*기자|"
        r"(?:기자|특파원)\s*(?:=|·|:)|무단전재|재배포\s*금지|"
        r"취재를\s*종합하면|편집자\s*주|"
        r"(?:관련\s*업계|당국|업계).{0,30}(?:따르면|밝혔|전했)|"
        r"(?:자사|당사).{0,120}(?:사업을?\s*수주|제공한다고|적용할\s*계획)|"
        r"(?:사업을?\s*수주했다|시스템을?\s*구축할\s*계획)",
        title + "\n" + text[:2_000],
        re.IGNORECASE,
    )
    if domain in RELEVANCE_PRESS_DOMAINS:
        return "news_or_education"
    if press_path and (press_lead or re.search(r"news\s*room|뉴스룸", title, re.I)):
        return "news_or_education"
    if host in {"t.me", "telegram.me", "www.telegram.me"}:
        return "public_messenger_page"
    return "unknown"


def near_duplicate_id(masked_title: str, masked_text: str) -> str:
    tokens = re.findall(
        r"[가-힣A-Za-z0-9_]{2,}", (masked_title + " " + masked_text).lower()
    )[:2_000]
    if not tokens:
        return ""
    weights = [0] * 64
    # Repeated sales phrases are common in these pages. Hash each distinct token
    # once and apply its frequency as a weight; this is equivalent to the former
    # per-occurrence loop and materially reduces CPU work on long copied posts.
    for token, count in Counter(tokens).items():
        value = int.from_bytes(
            hashlib.blake2b(token.encode("utf-8"), digest_size=8).digest(), "big"
        )
        for bit in range(64):
            weights[bit] += count if value & (1 << bit) else -count
    fingerprint = sum((1 << bit) for bit, weight in enumerate(weights) if weight >= 0)
    return f"simhash64:{fingerprint:016x}"


def contact_campaign_id(key: bytes, raw_text: str) -> str:
    normalized = unicodedata.normalize("NFKC", raw_text)
    contacts: set[str] = set()
    for match in re.finditer(
        r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b",
        normalized,
    ):
        contacts.add("email:" + match.group(0).lower().rstrip(".,;)"))
    for match in re.finditer(
        r"(?<!\d)(?:01[016789]|02|0[3-6][1-5])[- .]?\d{3,4}[- .]?\d{4}(?!\d)",
        normalized,
    ):
        contacts.add("phone:" + re.sub(r"\D", "", match.group(0)))
    for match in re.finditer(
        r"(?<!\w)@([A-Za-z0-9_]{3,})(?!\w)",
        normalized,
    ):
        contacts.add("account:" + match.group(1).lower())
    messenger_pattern = re.compile(
        r"(?i)(?:텔레그램|telegram|텔그|텔레|텔[ᄀ-ᄒㅏ-ㅣ]?|"
        r"카카오톡|카톡|오픈채팅|라인|line)"
        r"(?:\s|[^\w가-힣]){0,6}"
        r"(?:아이디|id|주소|문의)?"
        r"(?:\s|[^\w가-힣]){0,6}@?([A-Za-z0-9_.-]{3,})",
    )
    for match in messenger_pattern.finditer(normalized):
        contacts.add("account:" + match.group(1).lower().rstrip(".,;)-"))
    if not contacts:
        contacts = {
            re.sub(r"\s+", "", match.group(0)).lower().rstrip(".,;)")
            for match in re.finditer(
                r"(?i)(?:https?://|www\.)[^\s<>\"']+",
                normalized,
            )
        }
    if not contacts:
        return ""
    payload = "\n".join(sorted(contacts))
    return (
        "contact-hmac:"
        + hmac.new(key, payload.encode("utf-8"), hashlib.sha256).hexdigest()[:24]
    )


def campaign_fingerprint_text(title: str, text: str) -> str:
    """Build the reproducible contact view used for campaign grouping.

    The shareable review copy preserves messenger IDs but masks phone numbers
    and email addresses.  Using that same view here prevents a campaign from
    changing when revalidation operates on the review copy rather than the
    original HTTP response.
    """
    return (
        mask_text(title, preserve_messenger_ids=True)
        + "\n"
        + mask_text(text, preserve_messenger_ids=True)
    )


def make_record(
    index: int,
    candidate: Candidate,
    final_url: str,
    http_status: int,
    title: str,
    text: str,
    key: bytes,
) -> dict[str, object]:
    masked_title = mask_text(title)
    masked_text = mask_text(text)
    source_kind, source_identity = source_unit_descriptor(
        final_url,
        title,
        text,
    )
    fingerprint = near_duplicate_id(masked_title, masked_text)
    now = dt.datetime.now(dt.timezone(dt.timedelta(hours=9))).isoformat(
        timespec="seconds"
    )
    record: dict[str, object] = {name: "" for name in SCHEMA}
    record.update(
        {
            "sample_id": f"EG-{index:06d}",
            "collected_at": now,
            "source_type": candidate.source_type,
            "registrable_domain": registrable_domain(
                urlsplit(final_url).hostname or ""
            ),
            "source_unit_kind": source_kind,
            "source_unit_hmac": source_unit_token(
                key,
                (source_kind, source_identity),
            ),
            "url_hmac": url_digest(key, candidate.url),
            "http_status": http_status,
            "final_url_hmac": url_digest(key, final_url),
            "page_type": classify_page_type(final_url, masked_title, masked_text),
            "live_status": "accessible",
            "extraction_status": "success",
            "masked_title": masked_title,
            "masked_text": masked_text,
            "language_mix": language_mix(masked_title + "\n" + masked_text),
            "obfuscation_type": obfuscation_type(masked_title + "\n" + masked_text),
            "intent_label": "",
            "target_label": "",
            "contact_label": "",
            "final_label": "uncertain",
            "evidence_spans": "{}",
            # Keep the old field during the handoff transition. Both values are
            # fingerprints, not precomputed duplicate cluster identifiers.
            "near_duplicate_cluster": fingerprint,
            "near_duplicate_fingerprint": fingerprint,
            "campaign_group": contact_campaign_id(
                key,
                campaign_fingerprint_text(title, text),
            ),
        }
    )
    return record


def existing_hashes(csv_path: Path) -> set[str]:
    if not csv_path.exists():
        return set()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return {
            row["url_hmac"] for row in csv.DictReader(handle) if row.get("url_hmac")
        }


def existing_fingerprints(csv_path: Path) -> set[str]:
    """Load fingerprints already retained, including older handoff files."""
    if not csv_path.exists():
        return set()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return {
            fingerprint
            for row in csv.DictReader(handle)
            if (
                fingerprint := (
                    row.get("near_duplicate_fingerprint")
                    or row.get("near_duplicate_cluster")
                    or ""
                )
            )
        }


def next_sample_index(csv_path: Path) -> int:
    """Return an unused numeric suffix, including after interrupted resumes."""
    if not csv_path.exists():
        return 1
    maximum = 0
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            match = re.fullmatch(r"EG-(\d+)", row.get("sample_id", ""))
            if match:
                maximum = max(maximum, int(match.group(1)))
    return maximum + 1


def existing_success_domains(csv_path: Path) -> set[str]:
    if not csv_path.exists():
        return set()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return {
            row["registrable_domain"]
            for row in csv.DictReader(handle)
            if row.get("registrable_domain")
        }


def existing_success_domain_counts(csv_path: Path) -> Counter[str]:
    if not csv_path.exists():
        return Counter()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return Counter(
            row["registrable_domain"]
            for row in csv.DictReader(handle)
            if row.get("registrable_domain")
        )


def existing_success_type_counts(csv_path: Path) -> Counter[str]:
    if not csv_path.exists():
        return Counter()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return Counter(
            infer_collection_type(
                "",
                row.get("masked_title", ""),
                row.get("masked_text", ""),
            )
            for row in csv.DictReader(handle)
        )


def existing_success_campaign_counts(csv_path: Path) -> Counter[str]:
    if not csv_path.exists():
        return Counter()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return Counter(
            campaign
            for row in csv.DictReader(handle)
            if (campaign := row.get("campaign_group", ""))
        )


def existing_success_source_unit_counts(csv_path: Path) -> Counter[str]:
    if not csv_path.exists():
        return Counter()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        return Counter(
            unit
            for row in csv.DictReader(handle)
            if (unit := row.get("source_unit_hmac", ""))
        )


def existing_source_unit_descriptors(review_path: Path) -> set[tuple[str, str]]:
    """Load raw source descriptors used to suppress same-board rediscovery."""
    if not review_path.exists():
        return set()
    descriptors: set[tuple[str, str]] = set()
    with review_path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            source_url = row.get("source_url", "")
            if not source_url:
                continue
            descriptors.add(
                source_unit_descriptor(
                    source_url,
                    row.get("title", ""),
                    row.get("text", ""),
                )
            )
    return descriptors


def terminal_attempt_hashes(log_path: Path) -> set[str]:
    if not log_path.exists():
        return set()
    terminal: set[str] = set()
    with log_path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            digest = row.get("url_hmac", "")
            outcome = row.get("outcome", "")
            reason = row.get("reason", "")
            status = row.get("http_status", "")
            if not digest or outcome == "success":
                continue
            if outcome == "skipped":
                terminal.add(digest)
                continue
            if reason in {
                "insufficient_text",
                "insufficient_meaningful_tokens",
                "insufficient_korean_text",
                "challenge_or_access_page",
            }:
                terminal.add(digest)
                continue
            if reason == "http_status" and status.isdigit():
                code = int(status)
                if 400 <= code < 500 and code not in {408, 429}:
                    terminal.add(digest)
    return terminal


def append_csv(
    path: Path, records: Iterable[dict[str, object]], fieldnames: list[str]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    exists = path.exists() and path.stat().st_size > 0
    with path.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        if not exists:
            writer.writeheader()
        writer.writerows(records)
    os.chmod(path, 0o600)


def replace_csv(
    path: Path,
    records: Iterable[dict[str, object]],
    fieldnames: list[str],
    mode: int,
) -> None:
    """Atomically replace a generated CSV while preserving its share mode."""
    temp_path = path.with_suffix(path.suffix + ".rewrite.tmp")
    with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fieldnames,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(records)
    os.chmod(temp_path, mode)
    temp_path.replace(path)


def revalidate_existing_records(
    masked_path: Path,
    review_path: Path,
    key: bytes,
    relevance_mode: str,
    source_unit_limit: int,
    campaign_limit: int,
) -> Counter[str]:
    """Reapply current precision and diversity rules to prior shared rows."""
    removed: Counter[str] = Counter()
    if not masked_path.exists() and not review_path.exists():
        return removed
    if not masked_path.exists() or not review_path.exists():
        raise RuntimeError(
            "Revalidation requires both candidates_masked.csv and data.csv"
        )
    with masked_path.open(encoding="utf-8-sig", newline="") as handle:
        masked_rows = list(csv.DictReader(handle))
    with review_path.open(encoding="utf-8-sig", newline="") as handle:
        review_rows = list(csv.DictReader(handle))
    review_by_id = {row.get("sample_id", ""): row for row in review_rows}
    if len(review_by_id) != len(review_rows):
        raise RuntimeError("Shareable review data contains duplicate sample_id values")

    kept_masked: list[dict[str, object]] = []
    kept_review: list[dict[str, object]] = []
    source_counts: Counter[str] = Counter()
    campaign_counts: Counter[str] = Counter()
    for masked in masked_rows:
        sample_id = masked.get("sample_id", "")
        raw = review_by_id.get(sample_id)
        if raw is None:
            raise RuntimeError(f"Shareable review row is missing: {sample_id}")
        source_url = raw.get("source_url", "")
        title = raw.get("title", "")
        text = raw.get("text", "")
        reason = relevance_gate_reason(
            title,
            text,
            source_url,
            masked.get("page_type", "unknown"),
            relevance_mode,
        )
        if reason:
            removed[reason] += 1
            continue
        descriptor = source_unit_descriptor(source_url, title, text)
        source_token = source_unit_token(key, descriptor)
        if source_unit_limit and source_counts[source_token] >= source_unit_limit:
            removed["source_unit_record_limit"] += 1
            continue
        campaign = contact_campaign_id(
            key,
            campaign_fingerprint_text(title, text),
        )
        if campaign and campaign_limit and campaign_counts[campaign] >= campaign_limit:
            removed["campaign_record_limit"] += 1
            continue
        masked["source_unit_kind"] = descriptor[0]
        masked["source_unit_hmac"] = source_token
        masked["campaign_group"] = campaign
        source_counts[source_token] += 1
        if campaign:
            campaign_counts[campaign] += 1
        kept_masked.append(masked)
        kept_review.append(raw)

    if len(kept_review) != len(kept_masked):
        raise RuntimeError("Revalidation produced mismatched output rows")
    replace_csv(masked_path, kept_masked, SCHEMA, 0o600)
    replace_csv(review_path, kept_review, RESTRICTED_REVIEW_SCHEMA, 0o644)
    return removed


def upgrade_existing_csv_schema(path: Path) -> None:
    """Add newly standardized columns before appending to an older run."""
    if not path.exists():
        return
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        old_fields = reader.fieldnames or []
    if old_fields == SCHEMA:
        return
    if not old_fields or not set(old_fields).issubset(SCHEMA):
        raise RuntimeError("Existing output schema is incompatible with this collector")
    temp_path = path.with_suffix(path.suffix + ".upgrade.tmp")
    with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SCHEMA)
        writer.writeheader()
        for row in rows:
            row.setdefault("extraction_status", "success")
            if row.get("live_status") == "true":
                row["live_status"] = "accessible"
            row["source_type"] = {
                "public_web_search": "search",
                "private_seed": "seed",
            }.get(row.get("source_type", ""), row.get("source_type", "") or "search")
            row["page_type"] = {
                "reflected_search_page": "search_reflection",
                "cached_or_deleted": "deleted_or_inaccessible",
            }.get(row.get("page_type", ""), row.get("page_type", ""))
            row.setdefault(
                "near_duplicate_fingerprint", row.get("near_duplicate_cluster", "")
            )
            writer.writerow(row)
    os.chmod(temp_path, 0o600)
    temp_path.replace(path)


def upgrade_collection_log_schema(path: Path) -> None:
    if not path.exists():
        return
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        old_fields = reader.fieldnames or []
    if old_fields == LOG_SCHEMA:
        return
    if not old_fields or not set(old_fields).issubset(LOG_SCHEMA):
        raise RuntimeError("Existing collection log schema is incompatible")
    temp_path = path.with_suffix(path.suffix + ".upgrade.tmp")
    with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=LOG_SCHEMA)
        writer.writeheader()
        writer.writerows(rows)
    os.chmod(temp_path, 0o600)
    temp_path.replace(path)


def extraction_status_for_failure(reason: str, text_chars: int) -> str:
    if reason == "non_html_content":
        return "non_html"
    if reason in {"challenge_or_access_page", "dynamic_render_failure"}:
        return "dynamic_render_failure"
    if reason == "insufficient_meaningful_tokens":
        return "boilerplate_only"
    if reason == "insufficient_text":
        return "partial" if text_chars else "empty"
    if reason == "encoding_failure":
        return "encoding_failure"
    return "other_failure"


def extraction_failure_record(log: CollectionLog) -> dict[str, object] | None:
    extraction_reasons = {
        "non_html_content",
        "content_too_large",
        "insufficient_text",
        "insufficient_meaningful_tokens",
        "insufficient_korean_text",
        "challenge_or_access_page",
        "dynamic_render_failure",
        "encoding_failure",
    }
    if log.outcome == "success" or log.reason not in extraction_reasons:
        return None
    return {
        "url_hmac": log.url_hmac,
        "query_group": log.query_group,
        "attempted_at": log.attempted_at,
        "http_status": log.http_status,
        "reason": log.reason,
        "extraction_status": extraction_status_for_failure(
            log.reason, log.text_chars
        ),
        "text_chars": log.text_chars,
        "extraction_method": log.extraction_method,
    }


RESIDUAL_PATTERNS = {
    "email": r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b",
    "phone": r"(?<!\d)(?:01[016789]|02|0[3-6][1-5])[- .]?\d{3,4}[- .]?\d{4}(?!\d)",
    "national_id": r"(?<!\d)\d{6}\s*[-–]?\s*[1-4]\d{6}(?!\d)",
    "http_url": r"(?i)https?://\S+",
}


def masking_validation(csv_path: Path, dataset_version: str) -> dict[str, object]:
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    joined = "\n".join(
        row.get("masked_title", "") + "\n" + row.get("masked_text", "")
        for row in rows
    )
    hits = {
        name: len(re.findall(pattern, joined))
        for name, pattern in RESIDUAL_PATTERNS.items()
    }
    return {
        "dataset_version": dataset_version,
        "checked_rows": len(rows),
        **{f"residual_{name}_hits": count for name, count in hits.items()},
        "manual_reviewed_rows": 0,
        "manual_review_failures": 0,
        "passed": not any(hits.values()),
        "generated_at": dt.datetime.now(
            dt.timezone(dt.timedelta(hours=9))
        ).isoformat(timespec="seconds"),
    }


def write_restricted_json(path: Path, payload: dict[str, object]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    os.chmod(path, 0o600)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def csv_row_count(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.DictReader(handle))


def write_collector_labeling_workbook(
    review_data_path: Path,
    workbook_path: Path,
) -> int:
    """Create the shareable URL-review workbook directly from collector output."""
    with review_data_path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    source_urls = {row["sample_id"]: row["source_url"] for row in rows}
    workbook_rows = [
        {
            "sample_id": row["sample_id"],
            "registrable_domain": row["registrable_domain"],
            "masked_title": row["title"],
        }
        for row in rows
    ]
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    write_labeling_workbook(workbook_path, workbook_rows, source_urls)
    os.chmod(workbook_path, 0o644)
    return len(rows)


def collection_log_metrics(path: Path) -> dict[str, object]:
    if not path.exists():
        return {
            "attempted_candidates": 0,
            "outcome_counts": {},
            "attempt_reconciliation_passed": True,
        }
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    outcomes = Counter(row.get("outcome", "") for row in rows)
    expected = sum(outcomes.get(name, 0) for name in ("success", "failed", "skipped"))
    return {
        "attempted_candidates": len(rows),
        "outcome_counts": dict(outcomes),
        "attempt_reconciliation_passed": expected == len(rows),
    }


def collector_code_version() -> dict[str, str]:
    script_path = Path(__file__).resolve()
    result = {"collector_sha256": sha256_file(script_path)}
    try:
        root = script_path.parents[1]
        commit = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=root,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
        dirty = subprocess.run(
            ["git", "status", "--porcelain", "--", str(script_path)],
            cwd=root,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
        result["git_commit"] = commit + ("+dirty" if dirty else "")
    except (OSError, subprocess.SubprocessError):
        result["git_commit"] = "unknown"
    return result


def data_manifest(
    out_dir: Path,
    dataset_version: str,
    paths: Iterable[Path],
    settings: dict[str, object],
) -> dict[str, object]:
    files = []
    for path in paths:
        if not path.exists():
            continue
        item: dict[str, object] = {
            "path": str(path.relative_to(out_dir)),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }
        if path.suffix.lower() == ".csv":
            item["rows"] = csv_row_count(path)
        files.append(item)
    return {
        "dataset_version": dataset_version,
        "generated_at": dt.datetime.now(
            dt.timezone(dt.timedelta(hours=9))
        ).isoformat(timespec="seconds"),
        "code_version": collector_code_version(),
        "settings": settings,
        "field_compatibility": {
            "near_duplicate_cluster": (
                "legacy alias of near_duplicate_fingerprint; "
                "value is a SimHash fingerprint"
            ),
            "source_unit_hmac": (
                "HMAC identifier for one SNS account/channel, one board, "
                "or one standalone registrable domain"
            ),
        },
        "files": files,
    }


DETECTION_TYPES = {"개인정보DB", "여권 및 통장", "포털ID", "해킹대행", "기타"}


def infer_detection_type(configured_type: str, title: str, text: str) -> str:
    combined = (title + "\n" + text[:4_000]).lower()
    if configured_type != "기타":
        return configured_type
    if re.search(r"여권|통장|계좌|passport|bank", combined):
        return "여권 및 통장"
    if re.search(r"포털|네이버|다음|구글|portal", combined):
        return "포털ID"
    if re.search(r"해킹\s*(?:대행|의뢰)|hack(?:ing)?\s*(?:service|for hire)", combined):
        return "해킹대행"
    return configured_type


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, 7):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def prepare_detection_workbook(template_path: Path, output_path: Path, resume: bool):
    if resume and output_path.exists():
        workbook = load_workbook(output_path)
    else:
        if not template_path.exists():
            raise FileNotFoundError(f"Detection template not found: {template_path}")
        workbook = load_workbook(template_path)
        sheet = workbook[workbook.sheetnames[0]]
        # Delete only the example values. Formatting and the H4:N5 notice stay intact.
        for column in range(1, 7):
            sheet.cell(4, column).value = None
    sheet = workbook[workbook.sheetnames[0]]
    return workbook, sheet


def existing_detection_urls(sheet) -> set[str]:
    return {
        str(sheet.cell(row, 3).value).strip()
        for row in range(4, sheet.max_row + 1)
        if sheet.cell(row, 3).value
    }


def append_detection_entries(sheet, entries: Iterable[DetectionEntry]) -> None:
    occupied_rows = [
        row for row in range(4, sheet.max_row + 1) if sheet.cell(row, 3).value
    ]
    row = max(occupied_rows, default=3) + 1
    for entry in entries:
        if entry.detection_type not in DETECTION_TYPES:
            raise ValueError(f"Unsupported detection type: {entry.detection_type}")
        if row > 31:
            copy_row_style(sheet, 31, row)
        sheet.cell(row, 1).value = row - 3
        sheet.cell(row, 2).value = entry.detected_on
        sheet.cell(row, 2).number_format = "yyyy-mm-dd"
        sheet.cell(row, 3).value = entry.url
        sheet.cell(row, 3).hyperlink = entry.url
        sheet.cell(row, 3).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row, 4).value = entry.detection_type
        sheet.cell(row, 5).value = safe_spreadsheet_text(entry.registrant)
        sheet.cell(row, 6).value = entry.note
        row += 1


def safe_spreadsheet_text(value: str) -> str:
    value = value.strip()
    return "'" + value if value.startswith(("=", "+", "-", "@")) else value


def save_restricted_workbook(workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    os.chmod(path.parent, 0o700)
    workbook.save(path)
    os.chmod(path, 0o600)


def validate_output(
    csv_path: Path,
    target: int,
    domain_limit: int = 0,
    minimum_domains: int = 0,
    minimum_type_counts: dict[str, int] | None = None,
    campaign_limit: int = 0,
    source_unit_limit: int = 1,
) -> None:
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    source_units = Counter(
        row.get("source_unit_hmac", "")
        for row in rows
        if row.get("source_unit_hmac")
    )
    if len(source_units) < target:
        raise RuntimeError(
            f"Only {len(source_units)} effective source units; target is {target}"
        )
    if source_unit_limit and max(source_units.values(), default=0) > source_unit_limit:
        raise RuntimeError("A source unit exceeds the configured representative limit")
    if list(rows[0].keys()) != SCHEMA:
        raise RuntimeError("Output schema does not match the research plan")
    joined = "\n".join(
        (row["masked_title"] + "\n" + row["masked_text"]) for row in rows
    )
    hits = [
        name
        for name, pattern in RESIDUAL_PATTERNS.items()
        if re.search(pattern, joined)
    ]
    if hits:
        raise RuntimeError("Residual sensitive patterns found: " + ", ".join(hits))
    if any(row["url_hmac"].startswith(("http:", "https:")) for row in rows):
        raise RuntimeError("Raw URL leaked into URL identifier field")
    sample_ids = [row["sample_id"] for row in rows]
    if any(not sample_id for sample_id in sample_ids) or len(sample_ids) != len(
        set(sample_ids)
    ):
        raise RuntimeError("sample_id values must be present and unique")
    allowed_sources = {"search", "seed", "manual"}
    if any(row["source_type"] not in allowed_sources for row in rows):
        raise RuntimeError("Unsupported source_type found")
    domains = Counter(row["registrable_domain"] for row in rows)
    if domain_limit and max(domains.values(), default=0) > domain_limit:
        raise RuntimeError("A domain exceeds the configured record limit")
    if len(domains) < minimum_domains:
        raise RuntimeError(
            f"Only {len(domains)} domains; minimum is {minimum_domains}"
        )
    type_counts = Counter(
        infer_collection_type("", row["masked_title"], row["masked_text"])
        for row in rows
    )
    for collection_type, minimum in (minimum_type_counts or {}).items():
        if type_counts[collection_type] < minimum:
            raise RuntimeError(
                f"Only {type_counts[collection_type]} {collection_type} records; "
                f"minimum is {minimum}"
            )
    campaigns = Counter(
        row["campaign_group"] for row in rows if row.get("campaign_group")
    )
    if campaign_limit and max(campaigns.values(), default=0) > campaign_limit:
        raise RuntimeError("A contact campaign exceeds the configured record limit")


def dataset_metrics(csv_path: Path) -> dict[str, object]:
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    lengths = sorted(len(row["masked_text"]) for row in rows)
    domains = Counter(row["registrable_domain"] for row in rows)
    languages = Counter(row["language_mix"] for row in rows)
    page_types = Counter(row["page_type"] for row in rows)
    collection_types = Counter(
        infer_collection_type("", row["masked_title"], row["masked_text"])
        for row in rows
    )
    campaigns = Counter(
        row["campaign_group"] for row in rows if row.get("campaign_group")
    )
    source_units = Counter(
        row["source_unit_hmac"] for row in rows if row.get("source_unit_hmac")
    )
    source_kinds = Counter(
        row["source_unit_kind"] for row in rows if row.get("source_unit_kind")
    )
    fingerprints = Counter(
        row["near_duplicate_cluster"]
        for row in rows
        if row["near_duplicate_cluster"]
    )
    duplicate_groups = {key: count for key, count in fingerprints.items() if count > 1}
    middle = len(lengths) // 2
    median = (
        (lengths[middle - 1] + lengths[middle]) / 2
        if lengths and len(lengths) % 2 == 0
        else lengths[middle]
        if lengths
        else 0
    )
    return {
        "dataset_rows": len(rows),
        "unique_url_hmacs": len({row["url_hmac"] for row in rows}),
        "unique_domains": len(domains),
        "top_domains": domains.most_common(10),
        "largest_domain_rows": max(domains.values(), default=0),
        "largest_domain_share": (
            round(max(domains.values(), default=0) / len(rows), 4) if rows else 0
        ),
        "effective_source_units": len(source_units),
        "source_unit_kind_counts": dict(source_kinds),
        "largest_source_unit_rows": max(source_units.values(), default=0),
        "source_unit_diversity_ratio": (
            round(len(source_units) / len(rows), 4) if rows else 0
        ),
        "collection_type_counts": dict(collection_types),
        "unique_contact_campaigns": len(campaigns),
        "largest_contact_campaign_rows": max(campaigns.values(), default=0),
        "language_mix_counts": dict(languages),
        "page_type_counts": dict(page_types),
        "text_chars_min": min(lengths, default=0),
        "text_chars_median": median,
        "text_chars_average": round(sum(lengths) / len(lengths), 1) if lengths else 0,
        "text_chars_max": max(lengths, default=0),
        "exact_simhash_duplicate_groups": len(duplicate_groups),
        "exact_simhash_duplicate_rows": sum(duplicate_groups.values()),
    }


def main() -> int:
    args = parse_args()
    validate_args(args)
    domain_record_limit = effective_domain_record_limit(
        args.target,
        args.max_records_per_domain,
        args.max_domain_share,
    )
    minimum_domains = effective_minimum_domains(
        args.target,
        domain_record_limit,
        args.min_domains,
    )
    minimum_type_counts = collection_type_minimums(
        args.target,
        args.min_type_share,
    )
    if minimum_domains > args.target:
        raise ValueError("The minimum domain count cannot exceed --target")
    if sum(minimum_type_counts.values()) > args.target:
        raise ValueError("The configured type minimums exceed --target")
    discovery_relevance_gate = (
        args.discovery_relevance_gate or args.relevance_gate
    )
    excluded_urls = load_excluded_urls(args.exclude_csv)
    excluded_fingerprints = load_excluded_fingerprints(args.exclude_csv)
    args.out.mkdir(parents=True, exist_ok=True)
    os.chmod(args.out, 0o755)
    csv_path = args.out / "candidates_masked.csv"
    review_data_path = args.out / "data.csv"
    detection_path = args.out / "restricted" / "탐지내역_자동수집.xlsx"
    labeling_workbook_path = args.out / "label.xlsx"
    log_path = args.out / "collection_log.csv"
    failure_path = args.out / "extraction_failures.csv"
    summary_path = args.out / "collection_summary.json"
    masking_report_path = args.out / "masking_validation_report.json"
    manifest_path = args.out / "data_manifest.json"
    private_dir = args.out / ".private"
    key = get_or_create_hmac_key(private_dir)
    queue_path = private_dir / "candidate_queue.jsonl"
    keyword_expansion_path = private_dir / "keyword_expansions.csv"
    revalidation_removed: Counter[str] = Counter()

    workbook = None
    detection_sheet = None
    detection_urls: set[str] = set()
    if not args.skip_detection_workbook:
        workbook, detection_sheet = prepare_detection_workbook(
            args.template, detection_path, args.resume
        )
        detection_urls = existing_detection_urls(detection_sheet)

    if csv_path.exists() and not args.resume:
        raise RuntimeError(f"{csv_path} exists; use --resume or move it first")
    if args.resume:
        upgrade_existing_csv_schema(csv_path)
        upgrade_collection_log_schema(log_path)
    if args.revalidate_existing:
        revalidation_removed = revalidate_existing_records(
            csv_path,
            review_data_path,
            key,
            args.relevance_gate,
            args.max_records_per_source_unit,
            args.max_records_per_campaign,
        )
        print(
            "revalidated existing data: removed "
            f"{sum(revalidation_removed.values())} rows "
            f"{dict(revalidation_removed)}",
            flush=True,
        )
    done_hashes = existing_hashes(csv_path)
    retained_fingerprints = existing_fingerprints(csv_path) | excluded_fingerprints
    retained_domain_counts = existing_success_domain_counts(csv_path)
    retained_type_counts = existing_success_type_counts(csv_path)
    retained_campaign_counts = existing_success_campaign_counts(csv_path)
    retained_source_unit_counts = existing_success_source_unit_counts(csv_path)
    retained_source_unit_descriptors = existing_source_unit_descriptors(
        review_data_path
    )
    terminal_hashes = terminal_attempt_hashes(log_path) if args.resume else set()
    existing_record_count = len(done_hashes)
    existing_source_unit_count = len(retained_source_unit_counts)
    next_record_index = next_sample_index(csv_path)

    session = requests.Session()
    session.headers.update(
        {"User-Agent": USER_AGENT, "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.5"}
    )
    limiter = RateLimiter(args.domain_delay)
    robots_cache: dict[str, tuple[bool, str]] = {}
    logs: list[CollectionLog] = []
    successes: list[dict[str, object]] = []
    restricted_successes: list[dict[str, str]] = []
    detection_entries: list[DetectionEntry] = []
    successful_text_lengths: list[int] = []
    successful_provider_counts: Counter[str] = Counter()
    newly_discovered_links = 0
    flushed_successes = 0
    flushed_restricted_successes = 0
    flushed_detection_entries = 0

    def flush_pending(force: bool = False) -> None:
        nonlocal flushed_successes, flushed_restricted_successes
        nonlocal flushed_detection_entries
        if not force and len(logs) < args.checkpoint_every:
            return
        pending_successes = successes[flushed_successes:]
        if pending_successes:
            append_csv(csv_path, pending_successes, SCHEMA)
            flushed_successes = len(successes)
        pending_restricted = restricted_successes[flushed_restricted_successes:]
        if pending_restricted:
            append_csv(
                review_data_path,
                pending_restricted,
                RESTRICTED_REVIEW_SCHEMA,
            )
            os.chmod(review_data_path, 0o644)
            flushed_restricted_successes = len(restricted_successes)
        if logs:
            append_csv(log_path, [asdict(item) for item in logs], LOG_SCHEMA)
            failures = [
                failure
                for item in logs
                if (failure := extraction_failure_record(item)) is not None
            ]
            if failures:
                append_csv(failure_path, failures, EXTRACTION_FAILURE_SCHEMA)
            logs.clear()
        pending_detection = detection_entries[flushed_detection_entries:]
        if (
            pending_detection
            and detection_sheet is not None
            and workbook is not None
        ):
            append_detection_entries(detection_sheet, pending_detection)
            save_restricted_workbook(workbook, detection_path)
            flushed_detection_entries = len(detection_entries)
        save_candidate_queue(queue_path, candidates)

        if pending_successes:
            print(
                f"source units {len(retained_source_unit_counts)}/{args.target}; "
                f"records {existing_record_count + len(successes)}; "
                f"attempted {candidate_index}/{len(candidates)}",
                flush=True,
            )

    if args.seed_file:
        seed_candidates: list[Candidate] = []
        for seed_path in args.seed_file:
            try:
                loaded_seed = load_seed_candidates(seed_path)
            except ValueError as exc:
                if (
                    len(args.seed_file) == 1
                    or "did not contain any usable" not in str(exc)
                ):
                    raise
                print(f"skipped empty seed file: {seed_path}", flush=True)
                continue
            seed_candidates = merge_candidates(seed_candidates, loaded_seed)
        if not seed_candidates:
            raise ValueError("Seed files did not contain any usable public URLs")
        if args.seed_offset >= len(seed_candidates):
            raise ValueError(
                "--seed-offset must leave at least one candidate in the seed"
            )
        candidates = prefilter_seed_candidates(
            seed_candidates[args.seed_offset :], discovery_relevance_gate
        )
        save_candidate_queue(queue_path, candidates)
    elif (
        args.resume
        and not args.refresh_discovery
        and queue_path.exists()
        and queue_path.stat().st_size > 0
    ):
        queued_candidates = load_candidate_queue(queue_path)
        candidates = prefilter_seed_candidates(
            queued_candidates,
            discovery_relevance_gate,
        )
        print(
            f"loaded {len(candidates)}/{len(queued_candidates)} candidates "
            "from private resume queue after discovery filtering",
            flush=True,
        )
    else:
        known_query_specs = expand_query_specs(
            load_query_specs(args.queries), args.query_variants
        )
        query_specs = list(known_query_specs)
        if args.strict_search:
            query_specs = constrain_query_specs(query_specs)
        print(f"prepared {len(query_specs)} search queries", flush=True)
        requested_providers = args.search_provider
        google_api_enabled = bool(
            requested_providers and "google_api" in requested_providers
        )
        browser_providers = (
            [name for name in requested_providers if name != "google_api"]
            if requested_providers is not None
            else None
        )
        driver = (
            connect_browser(args.cdp)
            if browser_providers is None or browser_providers
            else None
        )
        google_api_key = (
            os.environ.get(args.google_api_key_env, "").strip()
            if google_api_enabled
            else ""
        )
        google_cse_id = (
            os.environ.get(args.google_cse_id_env, "").strip()
            if google_api_enabled
            else ""
        )

        def run_search(
            specs: list[QuerySpec], desired: int, soft_target_multiplier: int
        ) -> list[Candidate]:
            discovered: list[Candidate] = []
            soft_target = max(
                desired * soft_target_multiplier,
                desired + 30,
            )
            if google_api_enabled:
                discovered = discover_google_api_candidates(
                    session,
                    query_specs=specs,
                    desired=desired,
                    pages=args.search_pages,
                    delay=args.search_delay,
                    prefilter_mode=discovery_relevance_gate,
                    api_key=google_api_key,
                    cse_id=google_cse_id,
                    soft_target_multiplier=soft_target_multiplier,
                    minimum_domains=minimum_domains,
                    minimum_source_units=desired,
                    max_candidates_per_domain=args.max_candidates_per_domain,
                    known_source_units=retained_source_unit_descriptors,
                    page_offset=args.search_page_offset,
                )
            if driver is not None and len(discovered) < soft_target:
                browser_discovered = discover_candidates(
                    driver,
                    query_specs=specs,
                    desired=max(desired - len(discovered), 1),
                    pages=args.search_pages,
                    delay=args.search_delay,
                    soft_target_multiplier=soft_target_multiplier,
                    prefilter_mode=discovery_relevance_gate,
                    providers_enabled=browser_providers,
                    provider_stale_pages_limit=args.provider_stale_pages,
                    minimum_domains=minimum_domains,
                    minimum_source_units=desired,
                    max_candidates_per_domain=args.max_candidates_per_domain,
                    known_source_units=retained_source_unit_descriptors,
                    page_offset=args.search_page_offset,
                )
                discovered = merge_candidates(discovered, browser_discovered)
            return discovered

        try:
            candidates = run_search(
                query_specs,
                desired=max(args.target - existing_source_unit_count, 1),
                soft_target_multiplier=(
                    4 if discovery_relevance_gate == "labeling" else
                    15 if discovery_relevance_gate != "off" else 3
                ),
            )
            for round_number in range(1, args.keyword_expansion_rounds + 1):
                expansions = mine_keyword_expansions(
                    candidates,
                    known_query_specs,
                    round_number=round_number,
                    limit=args.keyword_expansion_per_round,
                    minimum_domains=args.keyword_expansion_min_domains,
                    require_contact=args.keyword_expansion_require_contact,
                )
                if not expansions:
                    print(
                        f"keyword expansion round {round_number}: no repeated "
                        "high-relevance pairs",
                        flush=True,
                    )
                    break
                append_csv(
                    keyword_expansion_path,
                    [asdict(item) for item in expansions],
                    KEYWORD_EXPANSION_SCHEMA,
                )
                expanded_specs = [
                    QuerySpec(
                        item.query_group,
                        item.detection_type,
                        item.query,
                    )
                    for item in expansions
                ]
                known_query_specs.extend(expanded_specs)
                search_specs = (
                    constrain_query_specs(expanded_specs)
                    if args.strict_search
                    else expanded_specs
                )
                print(
                    f"keyword expansion round {round_number}: searching "
                    f"{len(expanded_specs)} new pairs",
                    flush=True,
                )
                additions = run_search(
                    search_specs,
                    desired=max(args.target - len(candidates), 1),
                    soft_target_multiplier=3,
                )
                before_merge = len(candidates)
                candidates = merge_candidates(candidates, additions)
                print(
                    f"keyword expansion round {round_number}: "
                    f"{len(candidates) - before_merge} new candidates",
                    flush=True,
                )
        finally:
            if driver is not None:
                disconnect_browser(driver)
        save_candidate_queue(queue_path, candidates)

    candidates_before_known_source_filter = len(candidates)
    candidates = exclude_known_source_unit_candidates(
        candidates,
        retained_source_unit_descriptors,
    )
    known_source_candidates_removed = (
        candidates_before_known_source_filter - len(candidates)
    )
    if known_source_candidates_removed:
        print(
            "excluded "
            f"{known_source_candidates_removed} candidates from already "
            "retained boards/accounts/sites",
            flush=True,
        )
        save_candidate_queue(queue_path, candidates)

    if excluded_urls:
        before_exclusion = len(candidates)
        candidates = [
            candidate for candidate in candidates if candidate.url not in excluded_urls
        ]
        removed = before_exclusion - len(candidates)
        print(
            f"excluded {removed} previously sampled URLs from candidate pool",
            flush=True,
        )
        save_candidate_queue(queue_path, candidates)

    # Within each source unit, try the clearest offer first so the single
    # representative is not whichever post happened to be discovered first.
    # In labeling mode the lower-ranked documents remain in the queue as hard
    # negatives; ordering does not discard them.
    candidates.sort(key=discovery_relevance_score, reverse=True)

    candidates = prioritize_candidates_by_type_deficit(
        candidates,
        retained_type_counts,
        minimum_type_counts,
    )
    candidates = interleave_candidates_by_domain(
        candidates,
        args.max_candidates_per_domain,
        args.max_candidates_per_source_unit,
    )
    save_candidate_queue(queue_path, candidates)

    if args.resume and args.follow_links_per_page:
        priority_domains = existing_success_domains(csv_path)
        if priority_domains:
            candidates.sort(
                key=lambda candidate: (
                    registrable_domain(urlsplit(candidate.url).hostname or "")
                    not in priority_domains
                )
            )
            save_candidate_queue(queue_path, candidates)
            print(
                f"prioritized candidates from {len(priority_domains)} successful domains",
                flush=True,
            )

    candidate_pool_limit = args.candidate_pool_limit or args.target * 4
    candidate_urls = {candidate.url for candidate in candidates}
    candidate_domain_counts = defaultdict(int)
    candidate_source_unit_counts: Counter[tuple[str, str]] = Counter()
    for queued_candidate in candidates:
        candidate_domain_counts[
            registrable_domain(urlsplit(queued_candidate.url).hostname or "")
        ] += 1
        candidate_source_unit_counts[
            source_unit_descriptor(
                queued_candidate.url,
                queued_candidate.discovery_text,
                "",
                allow_unresolved_social_post=True,
            )
        ] += 1
    candidate_index = 0

    def enqueue_related_candidates(
        html: str, final_url: str, parent: Candidate
    ) -> None:
        nonlocal newly_discovered_links
        if not args.follow_links_per_page or len(candidates) >= candidate_pool_limit:
            return
        remaining_pool = candidate_pool_limit - len(candidates)
        related_links = discover_related_internal_links(
            html,
            final_url,
            min(args.follow_links_per_page, remaining_pool),
        )
        priority_candidates: list[Candidate] = []
        for related_url in related_links:
            if related_url in candidate_urls:
                continue
            related_domain = registrable_domain(
                urlsplit(related_url).hostname or ""
            )
            if (
                args.max_candidates_per_domain
                and candidate_domain_counts[related_domain]
                >= args.max_candidates_per_domain
            ):
                continue
            related_source_unit = source_unit_descriptor(
                related_url,
                allow_unresolved_social_post=True,
            )
            if (
                args.max_candidates_per_source_unit
                and candidate_source_unit_counts[related_source_unit]
                >= args.max_candidates_per_source_unit
            ):
                continue
            candidate_urls.add(related_url)
            candidate_domain_counts[related_domain] += 1
            candidate_source_unit_counts[related_source_unit] += 1
            priority_candidates.append(
                Candidate(
                    url=related_url,
                    query_group=parent.query_group,
                    detection_type=parent.detection_type,
                    source_type=parent.source_type,
                    search_provider=parent.search_provider,
                )
            )
            newly_discovered_links += 1
        if priority_candidates:
            candidates[candidate_index:candidate_index] = priority_candidates

    print(f"discovered {len(candidates)} candidates; collecting pages", flush=True)
    while candidate_index < len(candidates):
        flush_pending()
        candidate = candidates[candidate_index]
        candidate_index += 1
        if len(retained_source_unit_counts) >= args.target:
            break
        digest = url_digest(key, candidate.url)
        already_retained = digest in done_hashes
        if already_retained and not args.expand_existing_links:
            continue
        if digest in terminal_hashes:
            continue
        if candidate.url in detection_urls:
            continue
        allowed, robots_reason = robots_allowed(
            session, candidate.url, limiter, robots_cache
        )
        if not allowed:
            logs.append(
                CollectionLog(
                    digest, candidate.query_group, "skipped", reason=robots_reason
                )
            )
            continue
        response, final_url, reason = request_once(session, candidate.url, limiter)
        if response is None:
            logs.append(
                CollectionLog(digest, candidate.query_group, "failed", reason=reason)
            )
            continue
        status = response.status_code
        if status != 200:
            response.close()
            logs.append(
                CollectionLog(
                    digest, candidate.query_group, "failed", str(status), "http_status"
                )
            )
            continue
        html, reason = read_html(response)
        if html is None:
            response.close()
            logs.append(
                CollectionLog(
                    digest, candidate.query_group, "skipped", str(status), reason
                )
            )
            continue
        title, text, extraction_method = extract_title_text(html, final_url)
        quality_reason = text_quality_reason(
            text,
            args.min_text_chars,
            minimum_korean_chars=args.min_korean_chars,
            title=title,
        )
        fallback_url = public_content_fallback_url(final_url)
        if quality_reason and fallback_url:
            response.close()
            fallback_allowed, fallback_robots_reason = robots_allowed(
                session, fallback_url, limiter, robots_cache
            )
            if fallback_allowed:
                fallback_response, fallback_final_url, _ = request_once(
                    session, fallback_url, limiter
                )
                if fallback_response is not None and fallback_response.status_code == 200:
                    fallback_html, _ = read_html(fallback_response)
                    if fallback_html is not None:
                        fallback_title, fallback_text, fallback_method = (
                            extract_title_text(fallback_html, fallback_final_url)
                        )
                        fallback_quality_reason = text_quality_reason(
                            fallback_text,
                            args.min_text_chars,
                            minimum_korean_chars=args.min_korean_chars,
                            title=fallback_title,
                        )
                        if not fallback_quality_reason:
                            response = fallback_response
                            final_url = fallback_final_url
                            status = fallback_response.status_code
                            html = fallback_html
                            title = fallback_title
                            text = fallback_text
                            extraction_method = (
                                "public_mobile_fallback:" + fallback_method
                            )
                            quality_reason = ""
                        else:
                            fallback_response.close()
                    else:
                        fallback_response.close()
            elif fallback_robots_reason:
                quality_reason = fallback_robots_reason
        if quality_reason:
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "failed",
                    str(status),
                    quality_reason,
                    len(text),
                    extraction_method,
                )
            )
            continue
        canonical_final_url = canonicalize_url(final_url)
        if canonical_final_url and canonical_final_url in excluded_urls:
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "previously_sampled_url",
                    len(text),
                    extraction_method,
                )
            )
            continue
        masked_title = mask_text(title)
        masked_text = mask_text(text)
        collector_page_type = classify_page_type(
            final_url, masked_title, masked_text
        )
        # A category/list page is useful for traversal but is not a sample.
        # Mine its topical post links before the structural gate drops it.
        if collector_page_type == "board_listing":
            enqueue_related_candidates(html, final_url, candidate)
        relevance_reason = relevance_gate_reason(
            masked_title,
            masked_text,
            final_url,
            collector_page_type,
            args.relevance_gate,
            candidate.discovery_text,
        )
        if relevance_reason:
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    relevance_reason,
                    len(text),
                    extraction_method,
                )
            )
            continue
        if already_retained:
            enqueue_related_candidates(html, final_url, candidate)
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "expanded_existing_links",
                    len(text),
                    extraction_method,
                )
            )
            continue
        record_source_unit = source_unit_token(
            key,
            source_unit_descriptor(final_url, title, text),
        )
        if (
            args.max_records_per_source_unit
            and retained_source_unit_counts[record_source_unit]
            >= args.max_records_per_source_unit
        ):
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "source_unit_record_limit",
                    len(text),
                    extraction_method,
                )
            )
            continue
        record_domain = registrable_domain(urlsplit(final_url).hostname or "")
        if (
            domain_record_limit
            and retained_domain_counts[record_domain]
            >= domain_record_limit
        ):
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "domain_record_limit",
                    len(text),
                    extraction_method,
                )
            )
            continue
        enqueue_related_candidates(html, final_url, candidate)
        record = make_record(
            next_record_index + len(successes),
            candidate,
            final_url,
            status,
            title,
            text,
            key,
        )
        record_type = infer_collection_type(
            candidate.detection_type,
            title,
            text,
        )
        remaining_slots = args.target - len(retained_source_unit_counts)
        remaining_type_deficit = sum(
            max(0, minimum - retained_type_counts[collection_type])
            for collection_type, minimum in minimum_type_counts.items()
        )
        if (
            record_type not in minimum_type_counts
            or retained_type_counts[record_type]
            >= minimum_type_counts[record_type]
        ) and remaining_type_deficit >= remaining_slots:
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "reserved_for_type_diversity",
                    len(text),
                    extraction_method,
                )
            )
            continue
        fingerprint = str(record["near_duplicate_fingerprint"])
        if fingerprint and fingerprint in retained_fingerprints:
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "exact_simhash_duplicate",
                    len(text),
                    extraction_method,
                )
            )
            continue
        campaign = str(record["campaign_group"])
        if (
            campaign
            and args.max_records_per_campaign
            and retained_campaign_counts[campaign]
            >= args.max_records_per_campaign
        ):
            response.close()
            logs.append(
                CollectionLog(
                    digest,
                    candidate.query_group,
                    "skipped",
                    str(status),
                    "campaign_record_limit",
                    len(text),
                    extraction_method,
                )
            )
            continue
        response.close()
        if fingerprint:
            retained_fingerprints.add(fingerprint)
        retained_domain_counts[record_domain] += 1
        retained_source_unit_counts[record_source_unit] += 1
        retained_type_counts[record_type] += 1
        if campaign:
            retained_campaign_counts[campaign] += 1
        successes.append(record)
        restricted_successes.append(
            make_restricted_review_record(record, final_url, title, text)
        )
        for provider_name in candidate.search_provider.split(","):
            if provider_name:
                successful_provider_counts[provider_name] += 1
        successful_text_lengths.append(len(text))
        if detection_sheet is not None:
            detection_entries.append(
                DetectionEntry(
                    detected_on=dt.datetime.now(
                        dt.timezone(dt.timedelta(hours=9))
                    ).date(),
                    url=final_url,
                    detection_type=infer_detection_type(
                        candidate.detection_type, title, text
                    ),
                    registrant=args.registrant,
                )
            )
        logs.append(
            CollectionLog(
                digest,
                candidate.query_group,
                "success",
                str(status),
                robots_reason,
                len(text),
                extraction_method,
            )
        )
    flush_pending(force=True)
    # Required handoff files retain a header even when this run has no rows.
    append_csv(csv_path, [], SCHEMA)
    append_csv(review_data_path, [], RESTRICTED_REVIEW_SCHEMA)
    os.chmod(review_data_path, 0o644)
    append_csv(failure_path, [], EXTRACTION_FAILURE_SCHEMA)
    labeling_workbook_rows = write_collector_labeling_workbook(
        review_data_path,
        labeling_workbook_path,
    )

    total = existing_record_count + len(successes)
    effective_total = len(retained_source_unit_counts)
    dataset_version = (
        dt.datetime.now(dt.timezone(dt.timedelta(hours=9))).date().isoformat()
        + "-"
        + args.out.name
    )
    summary = {
        "generated_at": dt.datetime.now(dt.timezone(dt.timedelta(hours=9))).isoformat(
            timespec="seconds"
        ),
        "target": args.target,
        "target_unit": "distinct_source_unit",
        "successful_records": total,
        "effective_source_units": effective_total,
        "new_records": len(successes),
        "revalidated_removed_records": sum(revalidation_removed.values()),
        "revalidated_removal_reasons": dict(revalidation_removed),
        "discovered_candidates": len(candidates),
        "known_source_candidates_removed": known_source_candidates_removed,
        "processed_queue_positions": candidate_index,
        "remaining_queue_positions": max(len(candidates) - candidate_index, 0),
        "same_site_links_added": newly_discovered_links,
        "candidate_provider_counts": dict(
            Counter(
                provider_name
                for candidate in candidates
                for provider_name in candidate.search_provider.split(",")
                if provider_name
            )
        ),
        "new_successful_provider_counts": dict(successful_provider_counts),
        "minimum_text_chars": args.min_text_chars,
        "minimum_korean_chars": args.min_korean_chars,
        "new_text_chars_min": min(successful_text_lengths, default=0),
        "new_text_chars_max": max(successful_text_lengths, default=0),
        "new_text_chars_average": (
            round(sum(successful_text_lengths) / len(successful_text_lengths), 1)
            if successful_text_lengths
            else 0
        ),
        "raw_urls_in_dataset": False,
        "raw_urls_in_restricted_detection_workbook": not args.skip_detection_workbook,
        "raw_urls_in_shareable_labeling_workbook": True,
        "messenger_ids_preserved_in_shareable_data": True,
        "shareable_review_columns": RESTRICTED_REVIEW_SCHEMA,
        "attachments_downloaded": False,
        "login_or_bypass_used": False,
        "ai_judgement_used": False,
        "schema_source": "CISC-W26 research plan section 5.2 and upstream handoff guide",
        "diversity_requirements": {
            "minimum_domains": minimum_domains,
            "maximum_records_per_domain": domain_record_limit,
            "maximum_records_per_source_unit": args.max_records_per_source_unit,
            "minimum_records_per_type": minimum_type_counts,
            "maximum_records_per_contact_campaign": args.max_records_per_campaign,
        },
    }
    metrics = dataset_metrics(csv_path)
    summary.update(metrics)
    summary.update(collection_log_metrics(log_path))
    type_counts = metrics["collection_type_counts"]
    diversity_checks = {
        "minimum_domains_met": metrics["unique_domains"] >= minimum_domains,
        "domain_limit_met": (
            not domain_record_limit
            or metrics["largest_domain_rows"] <= domain_record_limit
        ),
        "source_unit_target_met": metrics["effective_source_units"] >= args.target,
        "source_unit_limit_met": (
            metrics["largest_source_unit_rows"]
            <= args.max_records_per_source_unit
        ),
        "type_minimums_met": all(
            int(type_counts.get(collection_type, 0)) >= minimum
            for collection_type, minimum in minimum_type_counts.items()
        ),
        "campaign_limit_met": (
            not args.max_records_per_campaign
            or metrics["largest_contact_campaign_rows"]
            <= args.max_records_per_campaign
        ),
    }
    summary["diversity_checks"] = diversity_checks
    summary["completion_criteria_met"] = (
        effective_total >= args.target and all(diversity_checks.values())
    )
    write_restricted_json(summary_path, summary)

    masking_report = masking_validation(csv_path, dataset_version)
    write_restricted_json(masking_report_path, masking_report)

    manifest = data_manifest(
        args.out,
        dataset_version,
        (
            csv_path,
            log_path,
            failure_path,
            summary_path,
            masking_report_path,
        ),
        {
            "target": args.target,
            "source_mode": "seed" if args.seed_file else "search",
            "seed_offset": args.seed_offset,
            "search_pages": args.search_pages,
            "search_page_offset": args.search_page_offset,
            "search_providers": args.search_provider or ["all"],
            "provider_stale_pages": args.provider_stale_pages,
            "query_variants": args.query_variants,
            "keyword_expansion_rounds": args.keyword_expansion_rounds,
            "keyword_expansion_per_round": args.keyword_expansion_per_round,
            "keyword_expansion_min_domains": args.keyword_expansion_min_domains,
            "keyword_expansion_require_contact": (
                args.keyword_expansion_require_contact
            ),
            "strict_search": args.strict_search,
            "discovery_relevance_gate": discovery_relevance_gate,
            "relevance_gate": args.relevance_gate,
            "search_delay_seconds": args.search_delay,
            "domain_delay_seconds": args.domain_delay,
            "minimum_text_chars": args.min_text_chars,
            "minimum_korean_chars": args.min_korean_chars,
            "checkpoint_attempts": args.checkpoint_every,
            "follow_links_per_page": args.follow_links_per_page,
            "candidate_pool_limit": candidate_pool_limit,
            "max_candidates_per_domain": args.max_candidates_per_domain,
            "max_candidates_per_source_unit": args.max_candidates_per_source_unit,
            "max_records_per_domain": args.max_records_per_domain,
            "max_records_per_source_unit": args.max_records_per_source_unit,
            "max_domain_share": args.max_domain_share,
            "effective_max_records_per_domain": domain_record_limit,
            "min_domains": args.min_domains,
            "effective_min_domains": minimum_domains,
            "min_type_share": args.min_type_share,
            "minimum_records_per_type": minimum_type_counts,
            "max_records_per_campaign": args.max_records_per_campaign,
            "refresh_discovery": args.refresh_discovery,
            "expand_existing_links": args.expand_existing_links,
            "revalidate_existing": args.revalidate_existing,
            "ai_judgement_used": False,
        },
    )
    manifest["shareable_files"] = [
        {
            "path": str(review_data_path.relative_to(args.out)),
            "rows": labeling_workbook_rows,
            "columns": RESTRICTED_REVIEW_SCHEMA,
            "mode": "0644",
            "raw_source_urls": True,
            "raw_messenger_ids": True,
            "included_in_public_manifest_hashes": False,
        },
        {
            "path": str(labeling_workbook_path.relative_to(args.out)),
            "rows": labeling_workbook_rows,
            "mode": "0644",
            "raw_source_urls": True,
            "included_in_public_manifest_hashes": False,
        },
    ]
    manifest["restricted_files"] = (
        [
            {
                "path": str(detection_path.relative_to(args.out)),
                "mode": "0600",
                "raw_source_urls": True,
                "included_in_public_manifest_hashes": False,
            }
        ]
        if detection_path.exists()
        else []
    )
    write_restricted_json(manifest_path, manifest)

    if effective_total >= args.target and all(diversity_checks.values()):
        validate_output(
            csv_path,
            args.target,
            domain_limit=domain_record_limit,
            minimum_domains=minimum_domains,
            minimum_type_counts=minimum_type_counts,
            campaign_limit=args.max_records_per_campaign,
            source_unit_limit=args.max_records_per_source_unit,
        )
        if not masking_report["passed"]:
            raise RuntimeError("Masking validation failed; do not hand off this dataset")
        print(
            f"complete: {effective_total} source units, {total} masked records",
            flush=True,
        )
        return 0
    unmet = [name for name, passed in diversity_checks.items() if not passed]
    suffix = f"; unmet: {', '.join(unmet)}" if unmet else ""
    print(
        f"incomplete: {effective_total}/{args.target} source units "
        f"({total} records){suffix}; rerun with --resume",
        file=sys.stderr,
    )
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
