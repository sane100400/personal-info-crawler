from __future__ import annotations

import base64
import csv
import datetime as dt
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from selenium.common.exceptions import WebDriverException

from collector.build_labeling_pilot import (
    SourceRow,
    assign_near_duplicate_clusters,
    intent_bucket,
    prioritize_rows,
    select_rows,
    strict_bucket,
    write_labeling_sheet,
)
from collector.labeling_workbook import write_labeling_workbook

from collector.collect_candidates import (
    Candidate,
    CollectionLog,
    DetectionEntry,
    QuerySpec,
    RESTRICTED_REVIEW_SCHEMA,
    SCHEMA,
    append_detection_entries,
    campaign_fingerprint_text,
    canonicalize_url,
    classify_page_type,
    constrain_query_specs,
    contact_campaign_id,
    data_manifest,
    effective_domain_record_limit,
    effective_minimum_domains,
    discover_candidates,
    discover_google_api_candidates,
    discover_related_internal_links,
    discovery_candidate_relevant,
    discovery_candidate_passes,
    discovery_relevance_score,
    expand_query_specs,
    extract_title_text,
    extraction_failure_record,
    exclude_known_source_unit_candidates,
    existing_fingerprints,
    infer_collection_type,
    interleave_candidates_by_domain,
    load_candidate_queue,
    load_excluded_fingerprints,
    load_excluded_urls,
    load_query_specs,
    load_seed_candidates,
    mask_text,
    masking_validation,
    make_record,
    merge_candidates,
    mine_keyword_expansions,
    near_duplicate_id,
    ordered_provider_names,
    prefilter_seed_candidates,
    prepare_detection_workbook,
    public_content_fallback_url,
    relevance_gate_reason,
    safe_spreadsheet_text,
    save_candidate_queue,
    save_restricted_workbook,
    source_unit_descriptor,
    source_unit_token,
    text_quality_reason,
    terminal_attempt_hashes,
    unwrap_search_result_url,
    upgrade_collection_log_schema,
    upgrade_existing_csv_schema,
    write_collector_labeling_workbook,
)

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "(양식) 탐지내역.xlsx"


class CollectorTests(unittest.TestCase):
    def test_five_percent_domain_share_requires_twenty_domains_for_500(self) -> None:
        limit = effective_domain_record_limit(500, 0, 0.05)
        self.assertEqual(limit, 25)
        self.assertEqual(effective_minimum_domains(500, limit, 0), 20)

    def test_absolute_domain_limit_uses_the_stricter_setting(self) -> None:
        self.assertEqual(effective_domain_record_limit(500, 40, 0.05), 25)
        self.assertEqual(effective_domain_record_limit(500, 10, 0.05), 10)

    def test_candidates_are_interleaved_and_capped_by_domain(self) -> None:
        candidates = [
            Candidate(f"https://one.example/{index}", "g", "기타")
            for index in range(4)
        ] + [
            Candidate(f"https://two.example/{index}", "g", "기타")
            for index in range(2)
        ]
        ordered = interleave_candidates_by_domain(candidates, max_per_domain=2)
        self.assertEqual(
            [row.url for row in ordered],
            [
                "https://one.example/0",
                "https://two.example/0",
                "https://one.example/1",
                "https://two.example/1",
            ],
        )

    def test_search_switches_provider_after_repeated_navigation_errors(self) -> None:
        class BrokenDriver:
            def __init__(self) -> None:
                self.calls = 0

            def get(self, _url: str) -> None:
                self.calls += 1
                raise WebDriverException("connection closed")

        driver = BrokenDriver()
        candidates = discover_candidates(
            driver,  # type: ignore[arg-type]
            [
                QuerySpec("group", "기타", f"검색어 {index}")
                for index in range(10)
            ],
            desired=20,
            pages=1,
            delay=0,
            providers_enabled=["bing"],
        )
        self.assertEqual(candidates, [])
        self.assertEqual(driver.calls, 3)

    def test_social_accounts_are_distinct_but_posts_share_the_account_unit(self) -> None:
        first = source_unit_descriptor("https://t.me/s/channel_a/10")
        second = source_unit_descriptor("https://t.me/s/channel_a?before=20")
        other = source_unit_descriptor("https://t.me/s/channel_b/1")
        self.assertEqual(first, second)
        self.assertNotEqual(first, other)
        self.assertEqual(first[0], "social_account")

    def test_instagram_and_twitter_count_each_account_as_a_source(self) -> None:
        instagram_a = source_unit_descriptor("https://www.instagram.com/seller_a/")
        instagram_b = source_unit_descriptor("https://instagram.com/seller_b/")
        twitter_first = source_unit_descriptor(
            "https://x.com/seller_c/status/100"
        )
        twitter_second = source_unit_descriptor(
            "https://twitter.com/seller_c/status/999"
        )
        self.assertNotEqual(instagram_a, instagram_b)
        self.assertEqual(twitter_first, twitter_second)
        self.assertEqual(twitter_first[1], "x-twitter:seller_c")
        self.assertTrue(
            all(
                item[0] == "social_account"
                for item in (
                    instagram_a,
                    instagram_b,
                    twitter_first,
                    twitter_second,
                )
            )
        )

    def test_unresolved_instagram_posts_are_not_counted_as_distinct_accounts(self) -> None:
        first = source_unit_descriptor("https://instagram.com/p/POST_A/")
        second = source_unit_descriptor("https://instagram.com/reel/POST_B/")
        self.assertEqual(first, second)

    def test_board_posts_share_one_unit_but_separate_boards_do_not(self) -> None:
        first = source_unit_descriptor("https://creativebox.kr/igtrade/100")
        second = source_unit_descriptor("https://creativebox.kr/igtrade/999")
        other = source_unit_descriptor("https://creativebox.kr/ttmarket/100")
        self.assertEqual(first, second)
        self.assertNotEqual(first, other)
        generic_first = source_unit_descriptor(
            "https://forum.example/bbs/board.php?bo_table=free&wr_id=1"
        )
        generic_second = source_unit_descriptor(
            "https://forum.example/bbs/board.php?bo_table=free&wr_id=900"
        )
        self.assertEqual(generic_first, generic_second)
        cafe24_first = source_unit_descriptor(
            "https://shop.example/article/product-qa/6/225083/"
        )
        cafe24_second = source_unit_descriptor(
            "https://shop.example/article/product-qa/6/224969/"
        )
        self.assertEqual(cafe24_first, cafe24_second)

    def test_naver_and_daum_cafes_count_each_cafe_as_one_board(self) -> None:
        naver_first = source_unit_descriptor(
            "https://cafe.naver.com/ca-fe/cafes/12345/articles/10"
        )
        naver_second = source_unit_descriptor(
            "https://cafe.naver.com/ca-fe/cafes/12345/articles/999"
        )
        naver_other = source_unit_descriptor(
            "https://cafe.naver.com/ca-fe/cafes/67890/articles/10"
        )
        legacy = source_unit_descriptor(
            "https://cafe.naver.com/ArticleRead.nhn?clubid=12345&articleid=77"
        )
        daum_first = source_unit_descriptor(
            "https://cafe.daum.net/cafe_a/AbCd/1"
        )
        daum_second = source_unit_descriptor(
            "https://cafe.daum.net/cafe_a/AbCd/200"
        )
        daum_other = source_unit_descriptor(
            "https://cafe.daum.net/cafe_b/AbCd/1"
        )
        self.assertEqual(naver_first, naver_second)
        self.assertEqual(naver_first, legacy)
        self.assertNotEqual(naver_first, naver_other)
        self.assertEqual(daum_first, daum_second)
        self.assertNotEqual(daum_first, daum_other)

    def test_standalone_site_pages_count_as_one_source_unit(self) -> None:
        self.assertEqual(
            source_unit_descriptor("https://seller.example/service/a"),
            source_unit_descriptor("https://seller.example/service/b"),
        )

    def test_source_unit_token_is_hmac_and_does_not_reveal_account(self) -> None:
        descriptor = source_unit_descriptor("https://t.me/s/private_channel/1")
        token = source_unit_token(b"test-key", descriptor)
        self.assertTrue(token.startswith("social_account-hmac:"))
        self.assertNotIn("private_channel", token)

    def test_candidate_interleave_caps_repeated_board_posts(self) -> None:
        candidates = [
            Candidate(f"https://creativebox.kr/igtrade/{index}", "g", "기타")
            for index in range(5)
        ] + [
            Candidate("https://creativebox.kr/ttmarket/1", "g", "기타")
        ]
        ordered = interleave_candidates_by_domain(
            candidates,
            max_per_source_unit=2,
        )
        self.assertEqual(len(ordered), 3)
        self.assertEqual(
            [source_unit_descriptor(item.url) for item in ordered],
            [
                ("board", "creativebox:igtrade"),
                ("board", "creativebox:ttmarket"),
                ("board", "creativebox:igtrade"),
            ],
        )

    def test_known_board_and_social_account_candidates_are_excluded(self) -> None:
        known = {
            source_unit_descriptor("https://forum.example/bbs/board.php?bo_table=free&wr_id=1"),
            source_unit_descriptor("https://instagram.com/seller_a/p/POST1"),
        }
        candidates = [
            Candidate(
                "https://forum.example/bbs/board.php?bo_table=free&wr_id=999",
                "g",
                "기타",
            ),
            Candidate(
                "https://forum.example/bbs/board.php?bo_table=trade&wr_id=2",
                "g",
                "기타",
            ),
            Candidate("https://instagram.com/seller_a/p/POST2", "g", "기타"),
            Candidate("https://instagram.com/seller_b/p/POST3", "g", "기타"),
        ]
        kept = exclude_known_source_unit_candidates(candidates, known)
        self.assertEqual(
            [source_unit_descriptor(candidate.url) for candidate in kept],
            [
                ("board", "forum.example:bo_table=trade"),
                ("social_account", "instagram.com:seller_b"),
            ],
        )

    def test_collection_type_strata_are_mutually_exclusive(self) -> None:
        self.assertEqual(
            infer_collection_type("기타", "위조여권 제작", "판매 문의"),
            "신분증·여권 위조/제작",
        )
        self.assertEqual(
            infer_collection_type("기타", "법인통장 매입", "텔레그램 문의"),
            "통장·계좌",
        )
        self.assertEqual(
            infer_collection_type("기타", "대출DB 판매", "실시간 자료"),
            "개인정보DB",
        )
        self.assertEqual(
            infer_collection_type("기타", "네이버 아이디 매입", "대량 문의"),
            "계정·아이디·가입인증",
        )

    def test_intent_gate_rejects_platform_policy_and_bank_product_pages(self) -> None:
        self.assertEqual(
            relevance_gate_reason(
                "법적 고지 - Apple 미디어 서비스 이용 약관",
                "계정과 콘텐츠를 구매하거나 판매하는 거래에 관한 서비스 약관",
                "https://www.apple.com/kr/legal/terms.html",
                "unknown",
                "intent",
            ),
            "excluded_domain",
        )
        self.assertEqual(
            relevance_gate_reason(
                "계정 탈취, 거래, 양도, 교환 등",
                "운영정책상 계정 판매와 구매 행위는 허용되지 않습니다.",
                "https://talksafety.kakao.com/policy/account",
                "unknown",
                "intent",
            ),
            "excluded_domain",
        )
        self.assertEqual(
            relevance_gate_reason(
                "자유입출금 예금 상품",
                "사업자 간 금 거래를 위한 정상 은행 계좌를 제공합니다.",
                "https://www.kebhana.com/product/1479724",
                "unknown",
                "intent",
            ),
            "excluded_domain",
        )

    def test_intent_gate_rejects_warning_post_titled_as_related_tips(self) -> None:
        self.assertEqual(
            relevance_gate_reason(
                "청소년보호법 위반 관련 팁",
                "텔레그램에서 위조신분증을 제작해 준다니 조심하세요.",
                "https://forum.example/post/1",
                "unknown",
                "intent",
            ),
            "excluded_document_type",
        )

    def test_labeling_workbook_has_links_dropdown_and_notes(self) -> None:
        row = {name: "" for name in SCHEMA}
        row.update(
            {
                "sample_id": "LP-000001",
                "registrable_domain": "example.com",
                "masked_title": "고객 DB 판매",
                "masked_text": "판매 의사 확인용 본문",
                "page_type": "unknown",
                "live_status": "accessible",
            }
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "label.xlsx"
            write_labeling_workbook(
                path,
                [row],
                {"LP-000001": "https://example.com/public/post/1"},
            )
            workbook = load_workbook(path)
            sheet = workbook["라벨링"]
            self.assertEqual(sheet["C2"].hyperlink.target, "https://example.com/public/post/1")
            self.assertEqual(sheet["F1"].value, "판정")
            self.assertEqual(sheet["G1"].value, "메모")
            self.assertEqual(len(sheet.data_validations.dataValidation), 1)
            self.assertIn("안내", workbook.sheetnames)
            self.assertNotIn("본문 전체", workbook.sheetnames)

    def test_labeling_workbook_supports_an_empty_result(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "label.xlsx"
            write_labeling_workbook(path, [], {})
            workbook = load_workbook(path)
            sheet = workbook["라벨링"]
            self.assertEqual(sheet.max_row, 1)
            self.assertEqual(len(sheet.data_validations.dataValidation), 0)
            self.assertIn("안내", workbook.sheetnames)

    def test_collector_writes_shareable_labeling_workbook(self) -> None:
        row = {
            "sample_id": "EG-000001",
            "collected_at": "2026-08-28T12:00:00+09:00",
            "source_url": "https://example.com/public/post/2",
            "registrable_domain": "example.com",
            "title": "고객 DB 판매",
            "text": "텔레그램 raw_handle 문의",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            csv_path = root / "data.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=RESTRICTED_REVIEW_SCHEMA)
                writer.writeheader()
                writer.writerow(row)
            workbook_path = root / "label.xlsx"
            count = write_collector_labeling_workbook(
                csv_path, workbook_path
            )
            workbook = load_workbook(workbook_path)
            sheet = workbook["라벨링"]
            self.assertEqual(count, 1)
            self.assertEqual(
                sheet["C2"].hyperlink.target,
                "https://example.com/public/post/2",
            )
            self.assertEqual(workbook_path.stat().st_mode & 0o777, 0o644)

    def test_restricted_labeling_sheet_includes_source_url(self) -> None:
        row = {name: "" for name in SCHEMA}
        row.update(
            {
                "sample_id": "LP-000001",
                "masked_title": "고객 DB 판매",
                "masked_text": "판매 의사 확인용 본문",
                "page_type": "unknown",
                "live_status": "accessible",
            }
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "labeling_with_urls.csv"
            write_labeling_sheet(
                path,
                [row],
                {"LP-000001": "https://example.com/public/post/1"},
            )
            with path.open(encoding="utf-8-sig", newline="") as handle:
                written = list(csv.DictReader(handle))
        self.assertEqual(
            written[0]["source_url"],
            "https://example.com/public/post/1",
        )
        self.assertEqual(written[0]["final_label"], "")

    def test_intent_priority_does_not_require_contact_details(self) -> None:
        self.assertEqual(intent_bucket(""), "intent_priority")
        self.assertEqual(
            intent_bucket("missing_concrete_contact"), "intent_priority"
        )
        self.assertEqual(
            intent_bucket("missing_body_offer"), "boundary_review"
        )
        self.assertEqual(
            intent_bucket("excluded_reporting_context"), "hard_negative"
        )

    def test_labeling_pilot_prioritizes_strict_and_balances_reasons(self) -> None:
        def item(index: int, reason: str) -> SourceRow:
            return SourceRow(
                row={"registrable_domain": f"d{index}.example"},
                candidate=Candidate(f"https://d{index}.example/post", "g", "기타"),
                source=Path("source"),
                source_tier="primary",
                gate_reason=reason,
                selection_bucket=strict_bucket(reason),
            )

        rows = [
            item(1, "excluded_reporting_context"),
            item(2, ""),
            item(3, "excluded_reporting_context"),
            item(4, "excluded_page_type"),
            item(5, "missing_concrete_contact"),
        ]
        ordered = prioritize_rows(rows, priority_enabled=True)
        self.assertEqual(
            [row.gate_reason for row in ordered],
            [
                "",
                "missing_concrete_contact",
                "excluded_reporting_context",
                "excluded_page_type",
                "excluded_reporting_context",
            ],
        )

    def test_labeling_pilot_honors_domain_cap(self) -> None:
        rows = [
            SourceRow(
                row={"registrable_domain": "same.example"},
                candidate=Candidate(f"https://same.example/{index}", "g", "기타"),
                source=Path("source"),
                source_tier="primary",
            )
            for index in range(3)
        ]
        rows.append(
            SourceRow(
                row={"registrable_domain": "other.example"},
                candidate=Candidate("https://other.example/1", "g", "기타"),
                source=Path("source"),
                source_tier="primary",
            )
        )
        selected = select_rows(rows, target=2, max_per_domain=1)
        self.assertEqual(
            [row.row["registrable_domain"] for row in selected],
            ["same.example", "other.example"],
        )

    def test_labeling_pilot_assigns_duplicate_clusters(self) -> None:
        rows = [
            {"masked_title": "계정 판매", "masked_text": "대량 계정 판매 문의"},
            {"masked_title": "계정 판매", "masked_text": "대량 계정 판매 문의"},
            {"masked_title": "다른 글", "masked_text": "전혀 다른 정상 문맥"},
        ]
        assign_near_duplicate_clusters(rows)
        self.assertEqual(
            rows[0]["near_duplicate_cluster"],
            rows[1]["near_duplicate_cluster"],
        )
        self.assertNotEqual(
            rows[0]["near_duplicate_cluster"],
            rows[2]["near_duplicate_cluster"],
        )

    def test_bing_redirect_is_unwrapped_without_request(self) -> None:
        target = "https://example.com/public/post/1"
        encoded = base64.urlsafe_b64encode(target.encode()).decode().rstrip("=")
        wrapped = f"https://www.bing.com/ck/a?u=a1{encoded}&ntb=1"
        self.assertEqual(unwrap_search_result_url(wrapped), target)

    def test_duckduckgo_redirect_is_unwrapped_without_request(self) -> None:
        target = "https://example.com/public/post/2"
        wrapped = f"https://duckduckgo.com/l/?uddg={target}"
        self.assertEqual(unwrap_search_result_url(wrapped), target)

    def test_contact_data_is_masked(self) -> None:
        text = (
            "문의 test@example.com, 010-1234-5678, 텔레그램 sample_id "
            "홈페이지https://example.com/path"
        )
        masked = mask_text(text)
        self.assertIn("[EMAIL]", masked)
        self.assertIn("[PHONE]", masked)
        self.assertIn("[MESSENGER_ID]", masked)
        self.assertIn("홈페이지[CONTACT_URL]", masked)
        self.assertNotIn("sample_id", masked)
        self.assertNotIn("https://", masked)

    def test_restricted_review_keeps_messenger_id_only(self) -> None:
        raw = (
            "문의 텔레그램 raw_handle, https://t.me/raw_channel, "
            "카카오 아이디: kakao_raw, ㅌㄹ short_raw, "
            "test@example.com, 010-1234-5678"
        )
        restricted = mask_text(raw, preserve_messenger_ids=True)
        self.assertIn("raw_handle", restricted)
        self.assertIn("https://t.me/raw_channel", restricted)
        self.assertIn("kakao_raw", restricted)
        self.assertIn("short_raw", restricted)
        self.assertIn("[EMAIL]", restricted)
        self.assertIn("[PHONE]", restricted)
        self.assertNotIn("test@example.com", restricted)
        self.assertNotIn("010-1234-5678", restricted)

    def test_telegram_shorthand_contact_is_masked(self) -> None:
        masked = mask_text("디비 텔그 sample_id 문의")
        self.assertIn("텔그 [MESSENGER_ID]", masked)
        self.assertNotIn("sample_id", masked)

    def test_messenger_contact_url_keeps_channel_but_masks_handle(self) -> None:
        masked = mask_text(
            "상담 https://t.me/private_handle 또는 https://open.kakao.com/o/secret"
        )
        self.assertIn("텔레그램 [MESSENGER_ID]", masked)
        self.assertIn("카카오톡 [MESSENGER_ID]", masked)
        self.assertNotIn("private_handle", masked)
        self.assertNotIn("/secret", masked)

    def test_spreadsheet_formula_prefix_is_neutralized(self) -> None:
        self.assertEqual(mask_text('=HYPERLINK("bad")'), '\'=HYPERLINK("bad")')
        self.assertEqual(safe_spreadsheet_text("+cmd"), "'+cmd")

    def test_tracking_parameters_and_fragment_are_removed(self) -> None:
        url = canonicalize_url("https://Example.com/post?id=7&utm_source=test#part")
        self.assertEqual(url, "https://example.com/post?id=7")

    def test_public_naver_blog_frame_has_mobile_fallback(self) -> None:
        self.assertEqual(
            public_content_fallback_url(
                "https://blog.naver.com/public_writer/223456789012"
            ),
            "https://m.blog.naver.com/public_writer/223456789012",
        )
        self.assertIsNone(
            public_content_fallback_url("https://example.com/public/post/1")
        )

    def test_contact_campaign_uses_hmac_without_plaintext(self) -> None:
        campaign = contact_campaign_id(
            b"test-key", "문의 test@example.com 010-1234-5678"
        )
        self.assertTrue(campaign.startswith("contact-hmac:"))
        self.assertNotIn("example.com", campaign)
        self.assertEqual(
            campaign,
            contact_campaign_id(b"test-key", "010-1234-5678 / test@example.com"),
        )

    def test_direct_contact_campaign_ignores_varying_page_urls(self) -> None:
        first = contact_campaign_id(
            b"test-key",
            "텔레그램 @same_seller https://example.com/post/1",
        )
        second = contact_campaign_id(
            b"test-key",
            "텔레그램 @same_seller https://other.example/post/9",
        )
        self.assertEqual(first, second)

    def test_obfuscated_messenger_label_keeps_same_campaign(self) -> None:
        first = contact_campaign_id(
            b"test-key",
            "문의 텔레•Run55 계정 판매",
        )
        second = contact_campaign_id(
            b"test-key",
            "문의텔ﾩ RUN55 아이디 임대",
        )
        self.assertEqual(first, second)

    def test_campaign_fingerprint_is_stable_after_review_masking(self) -> None:
        title = "문의 010-1234-5678 텔레그램 @same_seller"
        text = "고객 DB 판매 test@example.com"
        first = campaign_fingerprint_text(title, text)
        second = campaign_fingerprint_text(
            mask_text(title, preserve_messenger_ids=True),
            mask_text(text, preserve_messenger_ids=True),
        )
        self.assertEqual(first, second)
        self.assertNotIn("010-1234-5678", first)
        self.assertIn("same_seller", first)

    def test_tree_db_brand_is_not_a_personal_database_target(self) -> None:
        reason = relevance_gate_reason(
            "왕대추나무 100주 삽니다 : 트리디비",
            "조경수는 수종과 규격으로 검색합니다. 왕대추나무를 삽니다.",
            "https://treedb.example/tree/100",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "missing_relevant_target")

    def test_niche_personal_database_offer_is_kept_and_stratified(self) -> None:
        title = "병원DB 판매 텔레그램 문의"
        text = "성형 상담 고객 병원디비를 판매합니다. 대량 공급 가능합니다."
        self.assertEqual(
            relevance_gate_reason(
                title,
                text,
                "https://seller.example/db/1",
                "unknown",
                "intent",
            ),
            "",
        )
        self.assertEqual(
            infer_collection_type("기타", title, text),
            "개인정보DB",
        )

    def test_art_team_identity_card_request_is_not_illicit_forgery(self) -> None:
        reason = relevance_gate_reason(
            "미술팀 한그림 - 신분증 제작의뢰합니다",
            "PROJECT WORKS 제작의뢰 Professional Art work and Design service",
            "https://art.example/project/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_legal_prop_or_security_guide")

    def test_account_purchase_warning_is_not_a_buying_inquiry(self) -> None:
        reason = relevance_gate_reason(
            "네이버카페 계정 구매하시려는 사람들은 꼭 보세요",
            "신규계정은 절대 구매하지 마세요. 판매자가 AS도 해주지 않습니다.",
            "https://forum.example/post/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_trade_warning")

    def test_legal_decision_about_borrowed_account_is_not_trade_intent(self) -> None:
        reason = relevance_gate_reason(
            "차명계좌 대여에 따른 사례금으로 보아 과세한 처분의 당부",
            "조세심판원 주문 심판청구를 기각한다. 처분개요 청구인은 "
            "쟁점금액에 대한 청구주장을 제기했고 처분청은 경정·고지했다.",
            "https://law.example/case/100",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_legal_decision")

    def test_repeated_ssr_keyword_block_is_not_a_direct_post(self) -> None:
        block = "법인장파는곳 법인통장구매 법인통장판매 법인통장매입"
        reason = relevance_gate_reason(
            "상품 Q&A",
            f'"{block}" "{block}" "{block}" 기록을 정리한 일반 안내입니다.',
            "https://shop.example/article/qa/6/100",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_search_spam")

    def test_illicit_keyword_advertising_service_is_not_the_trade_post(self) -> None:
        text = " ".join(
            [
                "법인통장삽니다홍보 상단 노출 광고 도배 문의 폭주 유도 광고주",
                "법인장구매광고 상단 유지 키워드 광고 홍보",
            ]
            * 8
        )
        reason = relevance_gate_reason(
            "마케팅 홍보 상품 Q&A",
            text,
            "https://agency.example/article/qa/6/200",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_search_spam")

    def test_db_promotion_vendor_copy_is_not_a_db_sales_post(self) -> None:
        reason = relevance_gate_reason(
            "농산물 쇼핑몰",
            "DB 디비 관련 홍보, 광고는 트리플마케팅 이용해보세요. "
            "대출DB판매 광고 캠페인과 상단 노출을 대행합니다.",
            "https://farm.example/article/data/7/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_search_spam")

    def test_ssr_db_master_guide_is_not_a_direct_sales_post(self) -> None:
        text = (
            "Google 기준 상위 노출 SEO 기준 DB판매 마스터 가이드 "
            "검색 키워드들을 구조적으로 포함 키워드 통합 정리 "
            + "대출DB 판매 코인DB 구매 병원DB 매입 디비 판매 " * 8
        )
        reason = relevance_gate_reason(
            "사업자DB 완전 가이드",
            text,
            "https://shop.example/article/free/5/300",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_search_spam")

    def test_unsolicited_db_sales_complaint_is_not_purchase_intent(self) -> None:
        reason = relevance_gate_reason(
            "사업자 DB 판매한다는 연락 다들 받으시나요?",
            "광고 전화가 폭주합니다. 개인정보보호법 위반 아닌가요? "
            "이 전화 어떻게 차단하나요?",
            "https://community.example/post/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_reporting_context")

    def test_fanclub_ticket_page_is_not_identity_document_trade(self) -> None:
        reason = relevance_gate_reason(
            "팬클럽 추첨제 예매 안내",
            "공연 정보 관객 입장 17시 공연 시작 18시 티켓 금액 지정석 "
            "예매 시 본인 확인을 위해 여권을 지참해 주세요.",
            "https://fan.example/event/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_event_ticket_context")

    def test_repeated_no_db_purchase_copy_is_not_a_buying_post(self) -> None:
        reason = relevance_gate_reason(
            "보험DB 구매 없이 고객을 만나는 방법",
            "보험DB를 구매하지 않습니다. 광고나 보험DB 구매를 하지 않고 "
            "고객이 먼저 찾아오도록 온라인 영업을 합니다. 상담은 연락주세요.",
            "https://blog.example/post/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_db_purchase_alternative")

    def test_db_numbered_scale_model_is_not_personal_database_trade(self) -> None:
        reason = relevance_gate_reason(
            "카스 벤치형저울 DB-1 구매문의",
            "DB-1 저울은 목욕탕용 제품입니다. 모델 사양과 견적을 안내합니다.",
            "https://shop.example/product/db-1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_normal_product_context")

    def test_it_database_procurement_is_not_personal_data_trade(self) -> None:
        reason = relevance_gate_reason(
            "채널계 전용 DB 분리 구축을 위한 eXperDB 구매 입찰공고",
            "시스템 안정성 강화를 위한 DBMS 환경 구축 프로젝트입니다. "
            "eXperDB 소프트웨어 구매 및 설치, 기술지원과 교육을 "
            "제공할 입찰 참가업체를 모집합니다.",
            "https://association.example/notice/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_it_database_system")

    def test_privacy_compliance_keyword_article_is_not_a_db_offer(self) -> None:
        reason = relevance_gate_reason(
            "대출DB판매 대신 알아보는 상담 정보",
            "대출DB판매라는 검색어를 볼 때 확인할 내용입니다. "
            "개인정보 수집과 이용 목적, 이용자 동의 여부를 확인해야 "
            "합니다. 개인정보 제3자 제공 절차와 정보의 출처도 "
            "신중하게 확인해야 합니다. 자주 묻는 질문을 정리합니다.",
            "https://cleaning.example/gallery/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_db_compliance_guide")

    def test_overseas_futures_rental_account_is_not_bank_account_trade(self) -> None:
        reason = relevance_gate_reason(
            "해외선물 대여계좌 대여업체",
            "나스닥 실시간 미국 선물지수 거래를 지원합니다. "
            "전문가 교육과 실시간 담보금 예치, 실체결 거래 서비스를 "
            "상담하세요.",
            "https://futures.example/rental-account",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_investment_trading_service")

    def test_blog_false_positive_contexts_are_excluded(self) -> None:
        cases = [
            (
                "여권사진 과도한 보정, 공항에서 발목 잡히는 이유",
                "여권사진 촬영 규정과 보정 기준, 발급 반려 사유를 "
                "알아봅니다. 여권 위조로 오해받지 않도록 주의하세요.",
                "excluded_identity_photo_guide",
            ),
            (
                "외국인 불법체류와 여권 위조 심층 보고서",
                "목차, 서론, 현황, 사례 비교, 정책 제언, 제도 개선, "
                "결론을 통해 여권 위조 문제를 종합적으로 고찰합니다.",
                "excluded_informational_report",
            ),
            (
                "[채용] 보험설계사 이직 추천 DB 무한생성",
                "GA보험사 본부 채용 안내입니다. 소속 설계사에게 "
                "고객 DB를 무료로 제공하고 영업 시스템을 지원합니다.",
                "excluded_insurance_recruitment",
            ),
            (
                "DB 구매 없이 고객이 찾아오는 방법",
                "유료 DB를 사지 않고 콘텐츠로 고객 유입 구조를 "
                "직접 만드는 인바운드 영업을 소개합니다.",
                "excluded_db_purchase_alternative",
            ),
            (
                "DB하이텍, 지금 매수할까?",
                "DB하이텍은 8인치 파운드리 반도체 업체입니다. "
                "실적과 주가, 매수 시점을 분석합니다.",
                "excluded_db_brand_or_stock",
            ),
            (
                "보험 판매수수료 개편 이후 GA 변화",
                "1,200% 룰과 4년 분급제 규제가 보험 설계사 조직에 "
                "미치는 영향과 전략을 분석합니다.",
                "excluded_insurance_industry_analysis",
            ),
            (
                "네이버 아이디삽니다, 팔지 마세요",
                "본인 명의를 넘기는 것 자체가 잘못된 생각입니다. "
                "계정을 판매하게 되면 사기 범죄에 악용될 수 있으니 무시하세요.",
                "excluded_trade_warning",
            ),
        ]
        for title, text, expected in cases:
            with self.subTest(expected=expected):
                self.assertEqual(
                    relevance_gate_reason(
                        title,
                        text,
                        "https://blog.example/post/1",
                        "unknown",
                        "intent",
                    ),
                    expected,
                )

    def test_simhash_is_stable_for_equivalent_token_order(self) -> None:
        first = near_duplicate_id("제목", "반복 문구 반복 문구")
        second = near_duplicate_id("제목", "반복 문구 반복 문구")
        self.assertEqual(first, second)
        self.assertRegex(first, r"^simhash64:[0-9a-f]{16}$")

    def test_existing_fingerprints_supports_current_and_legacy_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "candidates_masked.csv"
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "near_duplicate_fingerprint",
                        "near_duplicate_cluster",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "near_duplicate_fingerprint": "simhash64:current",
                        "near_duplicate_cluster": "simhash64:legacy-alias",
                    }
                )
                writer.writerow(
                    {
                        "near_duplicate_fingerprint": "",
                        "near_duplicate_cluster": "simhash64:legacy",
                    }
                )
            self.assertEqual(
                existing_fingerprints(path),
                {"simhash64:current", "simhash64:legacy"},
            )

    def test_private_query_yaml_is_validated(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "queries.local.yaml"
            path.write_text(
                "groups:\n"
                "  - name: internal_example\n"
                "    detection_type: 개인정보DB\n"
                "    queries:\n"
                "      - private query\n",
                encoding="utf-8",
            )
            specs = load_query_specs(path)
            self.assertEqual(len(specs), 1)
            self.assertEqual(specs[0].group, "internal_example")
            self.assertEqual(specs[0].detection_type, "개인정보DB")

    def test_query_variants_are_created_in_stable_order(self) -> None:
        specs = [QuerySpec("group", "기타", "테스트 문의")]
        expanded = expand_query_specs(specs, 4)
        self.assertEqual(len(expanded), 4)
        self.assertEqual(expanded[0].query, "테스트 문의")
        self.assertEqual(expanded[-1].query, "테스트 문의 문의")

    def test_requested_search_provider_order_is_preserved(self) -> None:
        available = ["naver", "daum", "bing", "google"]
        self.assertEqual(
            ordered_provider_names(
                available,
                ["daum", "google", "naver", "daum"],
            ),
            ["daum", "google", "naver"],
        )
        self.assertEqual(ordered_provider_names(available, None), available)

    def test_google_api_discovery_uses_snippet_relevance_without_key_output(self) -> None:
        class FakeResponse:
            status_code = 200

            def json(self):
                return {
                    "items": [
                        {
                            "link": "https://example.com/public/post/1",
                            "title": "디비 텔그",
                            "snippet": "판매 관련 연락 안내",
                        }
                    ]
                }

            def close(self) -> None:
                pass

        class FakeSession:
            def __init__(self) -> None:
                self.params = {}

            def get(self, _url, **kwargs):
                self.params = kwargs["params"]
                return FakeResponse()

        session = FakeSession()
        candidates = discover_google_api_candidates(
            session,
            [QuerySpec("group", "개인정보DB", "디비 텔그")],
            desired=1,
            pages=1,
            delay=0,
            prefilter_mode="review",
            api_key="secret-key",
            cse_id="engine-id",
        )
        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0].url, "https://example.com/public/post/1")
        self.assertEqual(session.params["key"], "secret-key")

    def test_strict_queries_use_phrases_and_negative_filters(self) -> None:
        specs = [QuerySpec("group", "기타", "고객 DB 판매")]
        constrained = constrain_query_specs(specs)
        self.assertEqual(len(constrained), 4)
        self.assertTrue(constrained[0].query.startswith("고객 DB 판매"))
        self.assertTrue(constrained[1].query.startswith('"고객 DB 판매"'))
        self.assertTrue(all("-개인정보처리방침" in item.query for item in constrained))

    def test_search_snippet_prefilter_requires_target_and_trade_context(self) -> None:
        relevant = Candidate(
            "https://example.com/post",
            "group",
            "기타",
            discovery_text="고객 DB를 대량 판매한다는 게시물과 텔레그램 문의 안내",
        )
        generic = Candidate(
            "https://example.com/help",
            "group",
            "기타",
            discovery_text="고객센터에서 개인정보 처리방침을 확인하세요",
        )
        self.assertTrue(discovery_candidate_relevant(relevant))
        self.assertFalse(discovery_candidate_relevant(generic))
        target_only = Candidate(
            "https://example.com/news",
            "group",
            "기타",
            discovery_text="개인정보 유출 사고를 다룬 국내 기사",
        )
        self.assertTrue(discovery_candidate_passes(target_only, "labeling"))
        self.assertFalse(discovery_candidate_passes(target_only, "review"))
        hard_negative = Candidate(
            "https://example.com/database-guide",
            "group",
            "기타",
            discovery_text="DB 계정 관리 방법을 설명하는 기술 문서",
        )
        self.assertTrue(discovery_candidate_passes(hard_negative, "labeling"))
        self.assertFalse(discovery_candidate_passes(hard_negative, "review"))
        trade_only = Candidate(
            "https://example.com/classified/7",
            "group",
            "기타",
            discovery_text="대량 판매합니다. 자세한 품목은 본문을 확인하세요.",
        )
        self.assertTrue(discovery_candidate_passes(trade_only, "labeling"))
        self.assertFalse(discovery_candidate_passes(trade_only, "review"))
        self.assertGreater(
            discovery_relevance_score(relevant),
            discovery_relevance_score(generic),
        )

    def test_strict_search_prefilter_requires_local_direct_offer(self) -> None:
        direct_offer = Candidate(
            "https://example.com/post/positive",
            "group",
            "개인정보DB",
            discovery_text=(
                "고객 DB 판매합니다. 건당 단가 문의 "
                "https://t.me/private_handle"
            ),
        )
        reporting = Candidate(
            "https://example.com/news/negative",
            "group",
            "개인정보DB",
            discovery_text="고객 DB 판매 사건을 경찰이 적발한 텔레그램 관련 기사",
        )
        destination_contact_deferred = Candidate(
            "https://example.com/post/weak",
            "group",
            "개인정보DB",
            discovery_text="고객 DB 판매합니다. 텔레그램에서 안내합니다.",
        )
        fused_card = Candidate(
            "https://example.com/post/fused",
            "group",
            "개인정보DB",
            discovery_text=(
                "고객 DB 관련 보안 안내 "
                + "일반 설명 " * 80
                + "중고 자동차 부품 판매합니다. 연락 [PHONE]"
            ),
        )
        self.assertTrue(discovery_candidate_passes(direct_offer, "strict"))
        self.assertFalse(discovery_candidate_passes(reporting, "strict"))
        self.assertTrue(
            discovery_candidate_passes(destination_contact_deferred, "strict")
        )
        self.assertFalse(discovery_candidate_passes(fused_card, "strict"))

    def test_intent_search_prefilter_rejects_press_and_known_normal_contexts(self) -> None:
        quoted_news = Candidate(
            "https://regional.example/news/articleView.html?idxno=10",
            "group",
            "기타",
            discovery_text=(
                "SNS 신분증 위조 판매 기승. 취재진이 판매자에게 제작을 "
                "문의했으며 경찰은 범죄 악용이 우려된다고 밝혔다."
            ),
        )
        game_trade = Candidate(
            "https://market.example/post/20",
            "group",
            "포털ID",
            discovery_text="쿠키런 카카오 계정 구매합니다. 희망 가격 문의",
        )
        real_listing = Candidate(
            "https://community.example/post/21",
            "group",
            "개인정보DB",
            discovery_text=(
                "대출DB 판매합니다. 건당 단가 상담은 텔레그램 "
                "raw_handle 로 문의주세요."
            ),
        )
        normal_id_shop = Candidate(
            "https://shop.example/category/id",
            "group",
            "신분증",
            discovery_text=(
                "사원증 학생증 방문증 신분증 제작 상품목록 60 items "
                "장바구니 배송조회"
            ),
        )
        trade_guide = Candidate(
            "https://blog.example/account-risk",
            "group",
            "포털ID",
            discovery_text=(
                "텔레그램 계정 거래 후기 위험 분석. 계정 회수 피해와 "
                "정책 위반 위험을 자세히 알아봅니다."
            ),
        )
        self.assertFalse(discovery_candidate_passes(quoted_news, "intent"))
        self.assertFalse(discovery_candidate_passes(game_trade, "intent"))
        self.assertFalse(discovery_candidate_passes(normal_id_shop, "intent"))
        self.assertFalse(discovery_candidate_passes(trade_guide, "intent"))
        self.assertTrue(discovery_candidate_passes(real_listing, "intent"))

    def test_shorthand_target_and_contact_pass_review_prefilter(self) -> None:
        shorthand = Candidate(
            "https://example.com/post/8",
            "group",
            "개인정보DB",
            discovery_text="디비 텔그 문의",
        )
        brand_noise = Candidate(
            "https://example.com/news/8",
            "group",
            "개인정보DB",
            discovery_text="DB손해보험 농구단 소식",
        )
        self.assertTrue(discovery_candidate_relevant(shorthand))
        self.assertFalse(discovery_candidate_relevant(brand_noise))
        self.assertTrue(discovery_candidate_passes(shorthand, "review"))
        self.assertEqual(
            relevance_gate_reason(
                "디비 텔그",
                "보유 자료 관련 연락 안내입니다.",
                "https://example.com/post/8",
                "unknown",
                "review",
            ),
            "",
        )

    def test_review_prefilter_does_not_treat_display_url_as_contact(self) -> None:
        normal_result = Candidate(
            "https://support.example/contacts",
            "group",
            "기타",
            discovery_text=(
                "갤럭시 연락처 가져오기 방법 "
                "https://support.example/contacts 연락처 앱 사용 안내"
            ),
        )
        self.assertFalse(discovery_candidate_passes(normal_result, "review"))

    def test_keyword_expansion_requires_repetition_across_domains(self) -> None:
        candidates = [
            Candidate(
                "https://one.example/post/1",
                "personal_info_db",
                "개인정보DB",
                discovery_text="고객디비 텔그 문의",
            ),
            Candidate(
                "https://two.example/post/2",
                "personal_info_db",
                "개인정보DB",
                discovery_text="고객디비 텔그 연락",
            ),
            Candidate(
                "https://one.example/post/3",
                "portal_accounts",
                "포털ID",
                discovery_text="희귀아이디 텔그 문의",
            ),
        ]
        expansions = mine_keyword_expansions(
            candidates,
            [QuerySpec("seed", "개인정보DB", "디비 텔그")],
            round_number=1,
            limit=10,
            minimum_domains=2,
        )
        self.assertEqual([item.query for item in expansions], ["고객디비 텔그"])
        self.assertEqual(expansions[0].domain_frequency, 2)

    def test_candidate_merge_combines_snippet_evidence(self) -> None:
        current = [
            Candidate(
                "https://example.com/post",
                "group",
                "기타",
                discovery_text="디비 판매",
                search_provider="bing",
            )
        ]
        additions = [
            Candidate(
                "https://example.com/post",
                "group",
                "기타",
                discovery_text="텔그 문의",
                search_provider="google_api",
            )
        ]
        merged = merge_candidates(current, additions)
        self.assertEqual(len(merged), 1)
        self.assertIn("디비 판매", merged[0].discovery_text)
        self.assertIn("텔그 문의", merged[0].discovery_text)
        self.assertEqual(merged[0].search_provider, "bing,google_api")

    def test_contact_required_expansion_rejects_trade_only_pair(self) -> None:
        candidates = [
            Candidate(
                "https://one.example/post/1",
                "portal_accounts",
                "포털ID",
                discovery_text="계정 거래 안내",
            ),
            Candidate(
                "https://two.example/post/2",
                "portal_accounts",
                "포털ID",
                discovery_text="계정 거래 게시물",
            ),
        ]
        expansions = mine_keyword_expansions(
            candidates,
            [],
            round_number=1,
            limit=10,
            minimum_domains=2,
            require_contact=True,
        )
        self.assertEqual(expansions, [])

    def test_identity_document_forgery_passes_review_gate(self) -> None:
        self.assertEqual(
            relevance_gate_reason(
                "신분증 위조 텔그",
                "제작 관련 연락 안내입니다.",
                "https://example.com/post/9",
                "unknown",
                "review",
            ),
            "",
        )

    def test_relevance_gate_keeps_local_trade_and_contact_signals(self) -> None:
        reason = relevance_gate_reason(
            "고객 DB 판매합니다",
            "최신 자료 대량 보유, 건당 단가 문의 텔레그램 [MESSENGER_ID]",
            "https://board.example/post/7",
            "unknown",
            "strict",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_prioritizes_direct_offer_without_contact(self) -> None:
        candidate = Candidate(
            "https://board.example/post/intent",
            "group",
            "개인정보DB",
            discovery_text="고객 DB를 대량 보유하고 판매합니다. 건당 단가 안내",
        )
        self.assertTrue(discovery_candidate_passes(candidate, "intent"))
        self.assertEqual(
            relevance_gate_reason(
                "고객 DB 판매합니다",
                "최신 고객 명단을 대량 보유하고 건당 판매합니다.",
                candidate.url,
                "unknown",
                "intent",
            ),
            "",
        )
        self.assertEqual(
            relevance_gate_reason(
                "고객 DB 판매합니다",
                "최신 고객 명단을 대량 보유하고 건당 판매합니다.",
                candidate.url,
                "unknown",
                "strict",
            ),
            "missing_concrete_contact",
        )

    def test_intent_gate_rejects_reporting_and_missing_offer(self) -> None:
        reporting = relevance_gate_reason(
            "고객 DB 판매 게시물 적발",
            "경찰이 개인정보 명단 거래 사건을 수사하고 있습니다.",
            "https://news.example/article/intent",
            "news_or_education",
            "intent",
        )
        no_offer = relevance_gate_reason(
            "고객 DB 안내",
            "고객정보 데이터베이스의 보관 방식을 설명합니다.",
            "https://board.example/post/no-offer",
            "unknown",
            "intent",
        )
        self.assertEqual(reporting, "excluded_page_type")
        self.assertEqual(no_offer, "missing_body_offer")

    def test_intent_gate_rejects_press_copy_without_news_word(self) -> None:
        reason = relevance_gate_reason(
            '양홍원 "인스타그램 아이디 1천만원에 팝니다"…무슨 일?',
            (
                "[마이데일리 = 이승길 기자] SNS 계정을 판매하겠다는 "
                "글을 올려 이목이 쏠리고 있다. 무단전재 및 재배포 금지."
            ),
            "https://media.example/page/view/123",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_reporting_context")

    def test_v6_press_paths_and_reposted_reporting_are_excluded(self) -> None:
        press_text = (
            "편집자 주 쿠팡 구매자는 판매업체에 계좌번호 등 개인정보를 "
            "제공해야 환불받을 수 있는 것으로 확인됐다."
        )
        self.assertEqual(
            classify_page_type(
                "https://regional.example/news/articleView.html?idxno=10",
                "계좌 등 개인정보 제공해야 환불",
                press_text,
            ),
            "news_or_education",
        )
        repost = relevance_gate_reason(
            "홀로그램 가짜 주민등록증 5만원에 판매",
            (
                "중앙일보 취재를 종합하면 SNS에서 신분증을 만들어드립니다라는 "
                "글을 쉽게 찾을 수 있다. 판매자에게 제작을 문의하자 답했다."
            ),
            "https://community.example/view/10",
            "unknown",
            "intent",
        )
        self.assertEqual(repost, "excluded_reporting_context")

    def test_v6_telegram_shell_requires_offer_copy(self) -> None:
        empty_view = relevance_gate_reason(
            "Telegram: View coffee",
            (
                "커피 바이럴 마케팅 계정 매입\n"
                "If you have Telegram, you can view post and join right away."
            ),
            "https://t.me/channel/1",
            "public_messenger_page",
            "intent",
        )
        ambiguous_contact = relevance_gate_reason(
            "Telegram: Contact account",
            (
                "Download\n계좌매입\n제보센터입니다\nSend Message\n"
                "If you have Telegram, you can contact 계좌매입 right away."
            ),
            "https://t.me/account",
            "public_messenger_page",
            "intent",
        )
        explicit_contact = relevance_gate_reason(
            "Telegram: Contact seller",
            (
                "카톡계정 및 텔레그램 계정 최저가 및 각종 DB전문판매업체 "
                "24시간 상담가능\nIf you have Telegram, you can contact seller right away."
            ),
            "https://t.me/seller",
            "public_messenger_page",
            "intent",
        )
        self.assertEqual(empty_view, "excluded_empty_container")
        self.assertEqual(ambiguous_contact, "excluded_empty_container")
        self.assertEqual(explicit_contact, "")

    def test_empty_marketplace_form_does_not_inherit_intent_from_title(self) -> None:
        reason = relevance_gate_reason(
            "페이스북 실계정 다 삽니다",
            (
                "판매글 등록 유의사항\n"
                "1. 페이지명 :\n2. 팔로워 수 :\n3. 매매가 :\n"
                "4. 안전거래 가능 여부 :\n"
                "5. 거래 문의 연락처 (이메일 주소/오픈카톡방 링크 등) :\n"
                "추가로 작성해주시고 싶은 내용이 있으면 적어주세요."
            ),
            "https://market.example/post/empty",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_empty_listing_template")

    def test_v6_normal_id_products_props_and_photo_guide_are_excluded(self) -> None:
        normal_shop = relevance_gate_reason(
            "신분증 제작",
            (
                "사원증 협회신분증 종교신분증 학생증 방문증 Total 60 items "
                "상품명: 학생증01 상품명: 방문증1 장바구니 배송조회"
            ),
            "https://shop.example/category/id",
            "unknown",
            "intent",
        )
        prop = relevance_gate_reason(
            "의사면허증 제작 촬영소품",
            "민감정보를 비식별화한 촬영용 의사면허증 소품을 제작합니다.",
            "https://props.example/medical-license",
            "unknown",
            "intent",
        )
        photo = relevance_gate_reason(
            "자동차운전전문학원",
            "운전면허증 제작용 사진의 표준 규격을 참고해 사진을 제출하세요.",
            "https://academy.example/license-photo",
            "unknown",
            "intent",
        )
        self.assertEqual(normal_shop, "excluded_normal_product_context")
        self.assertEqual(prop, "excluded_legal_prop_or_security_guide")
        self.assertEqual(photo, "excluded_question_or_guide")

    def test_v6_keyword_spam_guides_and_unrelated_commodities_are_excluded(self) -> None:
        seo_template = relevance_gate_reason(
            "010인증 네이버아이디판매 계정대여",
            (
                "네이버계정판매 네이버아이디구매 네이버계정대여. "
                "예약 방법을 안내합니다. 평일 주말 모두 영업합니다. "
                "심야 시간대에도 가능합니다. 날짜와 인원을 알려주세요. "
                "첫 방문 고객 혜택과 이용 후기가 있습니다. 분위기를 찾는 분께 추천합니다."
            ),
            "https://unrelated.example/post/1",
            "unknown",
            "intent",
        )
        guide = relevance_gate_reason(
            "텔레그램 아이디 거래 후기 위험 완벽 분석",
            (
                "계정 회수 피해와 예상치 못한 위험을 자세히 알아봅니다. "
                "비공식 거래는 정책 위반이며 안전한 방법을 확인해야 합니다."
            ),
            "https://blog.example/guide",
            "unknown",
            "intent",
        )
        gift_card = relevance_gate_reason(
            "각종 상품권 판매 및 구매합니다",
            (
                "상품권 도소매 매입 판매 문의. 상품권 현금화와 상품권 구매를 "
                "도와드립니다. 카톡아이디 seller 텔레 아이디 seller2"
            ),
            "https://blog.example/gift-card",
            "unknown",
            "intent",
        )
        self.assertEqual(seo_template, "excluded_keyword_stuffing")
        self.assertEqual(guide, "excluded_informational_article")
        self.assertEqual(gift_card, "excluded_normal_product_context")

    def test_v6_clear_account_sale_on_unrelated_board_is_still_kept(self) -> None:
        reason = relevance_gate_reason(
            "대외활동 게시판 인스타 계정 판매",
            (
                "마케팅용으로 사용하던 인스타그램 계정 판매합니다. "
                "12년 계정부터 다양하게 있으며 입금 가능합니다. "
                "오픈카톡과 텔레그램 seller 로 문의주세요."
            ),
            "https://campus.example/board/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_marketplace_result_page_is_not_retained_as_a_post(self) -> None:
        repeated_controls = (
            "쪽지보내기 메일보내기 자기소개 아이디로 검색 전체게시물 " * 3
        )
        page_type = classify_page_type(
            "https://market.example/account-trade",
            "틱톡 계정 거래",
            (
                "전체 3,057건 / 1 페이지 팝니다 삽니다 거래완료 "
                "틱톡 계정 판매합니다. "
                + repeated_controls
                + "게시물 검색 검색대상 제목 내용 글쓴이"
            ),
        )
        self.assertEqual(page_type, "board_listing")
        self.assertEqual(
            relevance_gate_reason(
                "틱톡 계정 거래",
                "전체 5건 / 1 페이지 계정 판매합니다 " + repeated_controls,
                "https://market.example/account-trade",
                page_type,
                "intent",
            ),
            "excluded_page_type",
        )
        card_layout = classify_page_type(
            "https://market.example/igtrade",
            "인스타 계정 거래 - 마켓",
            (
                "전체 3,057건 / 1 페이지 판매중 팔로워 250 계정 판매 "
                "판매중 인스타 계정 매입 판매중 계정 팝니다"
            ),
        )
        self.assertEqual(card_layout, "board_listing")

    def test_intent_gate_rejects_reposted_warning_channel(self) -> None:
        reason = relevance_gate_reason(
            "저승사자 박제채널",
            (
                "사건내용: 계정 매입 업자에게 피해를 입어 제보합니다. "
                "이전 홍보글에는 계정 매입합니다라는 문구가 있습니다."
            ),
            "https://t.me/s/report_channel",
            "public_messenger_page",
            "intent",
        )
        self.assertEqual(reason, "excluded_aggregation_or_commentary")

    def test_intent_gate_rejects_legal_question_and_news_repost(self) -> None:
        legal_question = relevance_gate_reason(
            "계정 판매 후 명의 정지 고소 됨?",
            "계정을 넘긴 뒤 고소될 수 있는지 묻는 글입니다.",
            "https://community.example/question/3",
            "unknown",
            "intent",
        )
        repost = relevance_gate_reason(
            "이슈/유머 - 개인정보 DB 팝니다",
            "기사본문을 옮긴 게시물입니다.",
            "https://community.example/issue/4",
            "unknown",
            "intent",
        )
        self.assertEqual(legal_question, "excluded_question_or_guide")
        self.assertEqual(repost, "excluded_aggregation_or_commentary")

    def test_intent_gate_keeps_illicit_passport_issue_offer(self) -> None:
        reason = relevance_gate_reason(
            "여권발급이나 새 신분이 필요한 분",
            "위조 여권 제작 가능합니다. 다크웹 판매업체로 연락하세요.",
            "https://board.example/post/passport",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_rejects_normal_products_and_guides(self) -> None:
        passport_case = relevance_gate_reason(
            "여권 케이스 제작 방법",
            "맞춤형 디자인과 소재를 선택해 주문 제작합니다.",
            "https://shop.example/passport-case",
            "unknown",
            "intent",
        )
        bank_product = relevance_gate_reason(
            "농협 어린이 캐릭터 통장 판매",
            "은행에서 어린이 입출금 통장 상품을 출시해 판매합니다.",
            "https://bank.example/product/1",
            "unknown",
            "intent",
        )
        account_guide = relevance_gate_reason(
            "위탁판매 계정 정지 조건과 스마트스토어 차이",
            "정상 판매자가 계정 정지를 피하는 방법을 설명합니다.",
            "https://guide.example/post/1",
            "unknown",
            "intent",
        )
        self.assertEqual(passport_case, "excluded_question_or_guide")
        self.assertEqual(bank_product, "excluded_normal_product_context")
        self.assertEqual(account_guide, "excluded_question_or_guide")

        bankbook_case = relevance_gate_reason(
            "2026년 통장·카드케이스 구매 단가",
            "은행용 통장 케이스와 카드 비닐의 물품 입찰 공고입니다.",
            "https://bid.example/product/2",
            "unknown",
            "intent",
        )
        self.assertEqual(bankbook_case, "excluded_normal_product_context")

    def test_intent_gate_rejects_extraction_boilerplate(self) -> None:
        privacy_policy = relevance_gate_reason(
            "인스타 계정 판매",
            (
                "개인정보 처리방침 회사는 이용자의 개인정보를 보호합니다. "
                "수집 목적과 보유 기간을 안내합니다."
            ),
            "https://board.example/post/12",
            "unknown",
            "intent",
        )
        marketplace_footer = relevance_gate_reason(
            "인스타 계정 판매합니다",
            (
                "사업자등록번호 123-45-67890 통신판매업신고번호 안내. "
                "회사는 통신판매중개자로서 거래 당사자가 아닙니다."
            ),
            "https://market.example/item/12",
            "unknown",
            "intent",
        )
        self.assertEqual(privacy_policy, "excluded_extraction_boilerplate")
        self.assertEqual(marketplace_footer, "excluded_extraction_boilerplate")

    def test_intent_gate_rejects_account_market_guide(self) -> None:
        reason = relevance_gate_reason(
            "구글 계정 판매: 안전하고 신뢰할 수 있는 옵션",
            (
                "구글 계정 판매 시장 동향과 규모를 설명합니다. "
                "공급업체 선정 시 구매자 리뷰와 비교표를 확인하세요."
            ),
            "https://guide.example/account-market",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_market_guide")

    def test_intent_gate_rejects_empty_sales_keyword_container(self) -> None:
        text = "跳至主要内容 博文 此处没有可显示的博文！"
        page_type = classify_page_type(
            "https://sales-keywords.example/",
            "네이버아이디판매 텔레그램 문의",
            text,
        )
        self.assertEqual(page_type, "empty_container")
        self.assertEqual(
            relevance_gate_reason(
                "네이버아이디판매 텔레그램 문의",
                text,
                "https://sales-keywords.example/",
                page_type,
                "intent",
            ),
            "excluded_page_type",
        )

    def test_board_index_is_not_retained_as_a_post(self) -> None:
        text = (
            "이미지형 리스트형 게시물 검색 제목 글쓴이 아이디 "
            "고객DB 판매 2026-08-01 10:01 "
            "계정 매입 2026-08-01 09:40 "
            "통장 대여 2026-08-01 08:30"
        )
        page_type = classify_page_type(
            "https://shop.example/board/gallery/8/",
            "갤러리 - 정상 쇼핑몰",
            text,
        )
        self.assertEqual(page_type, "board_listing")
        self.assertEqual(
            relevance_gate_reason(
                "갤러리 - 정상 쇼핑몰",
                text,
                "https://shop.example/board/gallery/8/",
                page_type,
                "intent",
            ),
            "excluded_page_type",
        )

    def test_naver_influencer_content_index_is_not_retained_as_post(self) -> None:
        text = (
            "04:38 웍 런칭 안내 조회수 5,853\n"
            "02:36 소테팬 재출시 조회수 7,016\n"
            "04:57 틱톡 계정을 판매합니다 조회수 3,478\n"
            "04:32 스텐팬 관리법 조회수 5,711"
        )
        page_type = classify_page_type(
            "https://in.naver.com/creator/contents/external/1",
            "네이버 인플루언서",
            text,
        )
        self.assertEqual(page_type, "board_listing")

    def test_intent_gate_rejects_corporate_transfer_with_incidental_account(self) -> None:
        reason = relevance_gate_reason(
            "법인양도양수",
            (
                "법인 매입합니다. 자본금은 상관없고 한도제한 없는 은행통장이 "
                "있어야 합니다. 법무사에서 대표자 변경 서류를 작성합니다."
            ),
            "https://community.example/company-transfer/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_corporate_transfer")

    def test_intent_gate_rejects_public_business_contact_directory(self) -> None:
        reason = relevance_gate_reason(
            "전국 학원 주소록 연락처 DB 제공합니다",
            (
                "포털 등록 학원 DB 20만 건을 판매합니다. 자료 내역은 "
                "업장명, 구주소, 신주소, 우편번호, 팩스, 홈페이지, "
                "업종이며 결제 후 엑셀로 제공합니다."
            ),
            "https://market.example/business-directory",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_public_business_directory")

    def test_intent_gate_rejects_seo_explainer_without_direct_offer(self) -> None:
        reason = relevance_gate_reason(
            "보험퍼미션DB와 실시간DB 안내",
            (
                "보험퍼미션DB는 상담 현장에서 자주 언급됩니다. 운영 방향을 "
                "잡는 데 도움이 될 수 있습니다. 확인 항목을 살펴보는 것이 "
                "좋습니다. 실시간DB 페이지에 함께 담으면 좋은 내용과 구성 "
                "포인트를 설명합니다. 상담 문의를 남겨주시면 안내합니다."
            ),
            "https://guide.example/insurance-db",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_seo_explainer")

    def test_intent_gate_rejects_account_creation_software(self) -> None:
        reason = relevance_gate_reason(
            "텔레그램 계정 생성프로그램",
            (
                "카카오톡 다중 접속 프로그램과 아이디 생성 프로그램을 "
                "판매합니다. 보유 연락처 DB 자동 친구 추가도 지원합니다."
            ),
            "https://software.example/account-generator",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_account_creation_tool")

    def test_intent_gate_rejects_keyword_phrase_dump(self) -> None:
        phrases = [
            "카톡아이디판매",
            "운전면허증제작",
            "해외카톡구매",
            "인스타해킹가격",
            "각종해킹의뢰",
            "모바일신분증제작",
            "DB판매",
            "계정매입",
        ] * 8
        reason = relevance_gate_reason(
            "텔레그램 [ACCOUNT] 인스타해킹의뢰",
            "\n".join(phrases),
            "https://normal-office.example/qna/spam",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_keyword_stuffing")

        fifty_line_dump = [
            "대출DB판매",
            "주식디비구매",
            "최신DB문의",
            "계정매입",
            "통장판매",
        ] * 10
        reason = relevance_gate_reason(
            "대출DB 최신디비문의 [ACCOUNT]",
            "\n".join(fifty_line_dump),
            "https://hijacked-board.example/qna/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_keyword_stuffing")

    def test_intent_gate_keeps_structured_telegram_trade_list(self) -> None:
        block = """네이버 계정 매입가능
단가 15000
거래 양식
아이디
비밀번호
성함
전화번호
계좌번호
신분증 앞면
문의 [ACCOUNT]"""
        reason = relevance_gate_reason(
            "커뮤니티 계정 매입 채널 – Telegram",
            "\n".join([block] * 12),
            "https://t.me/s/account_buyer",
            "public_messenger_page",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_does_not_treat_transaction_form_example_as_article(self) -> None:
        reason = relevance_gate_reason(
            "네이버 계정 매입 – Telegram",
            (
                "네이버 계정 대여 비용은 당일 6000원 정산합니다. "
                "매입업무 마감 후 입금드립니다. 거래 양식 예시 홍길동 "
                "전화번호 계좌번호를 보내주세요. 카톡 raw_handle 문의"
            ),
            "https://t.me/s/account_rental",
            "public_messenger_page",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_rejects_market_update_article(self) -> None:
        reason = relevance_gate_reason(
            "텔레그램 업데이트 소식",
            (
                "업데이트 공지를 알려준 적이 있죠. 텔레그램 핸들을 거래 "
                "가능하게 하는 기능도 더 강화됐습니다. 시장을 관심 가져보는 "
                "걸 추천하며 개인 브랜딩 시대를 시사하는 것 같습니다."
            ),
            "https://t.me/s/marketing-news",
            "public_messenger_page",
            "intent",
        )
        self.assertEqual(reason, "excluded_informational_article")

        operations_guide = relevance_gate_reason(
            "보험DB 납품 시각이 AS 신청률을 바꿉니다",
            (
                "이 글은 공급사 관점에서 납품 시간을 정리한 것입니다. "
                "관련 안내서에 따르면 야간 전송은 별도 동의가 필요합니다. "
                "이 내용은 사실관계 전달이며 가정 시나리오로 계산합니다. "
                "AS가 발생하면 대체 DB를 납품합니다."
            ),
            "https://blog.example/db-delivery-guide",
            "unknown",
            "intent",
        )
        self.assertEqual(operations_guide, "excluded_informational_article")

    def test_intent_gate_rejects_trade_motive_question(self) -> None:
        reason = relevance_gate_reason(
            "커넥트 계정 사는 애들은 뭐야?",
            "30만원에 계정 매입한다는 사람들은 왜 매입하는 거임? 사기인가?",
            "https://community.example/post/motive-question",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_trade_motive_question")

    def test_intent_gate_rejects_legal_prop_and_security_guide(self) -> None:
        prop = relevance_gate_reason(
            "여권 제작 촬영소품",
            (
                "영화 촬영용 여권 소품이며 VOID와 SAMPLE을 표기하고 "
                "실제 개인정보는 사용하지 않습니다. 실제 발급과 위조는 "
                "제공하지 않습니다."
            ),
            "https://props.example/passport",
            "unknown",
            "intent",
        )
        security = relevance_gate_reason(
            "운전면허증 위조 방지 기술",
            "운전면허증 위조 방지를 위한 카드 보안 구조와 안전 설계 기준입니다.",
            "https://props.example/license-security",
            "unknown",
            "intent",
        )
        self.assertEqual(prop, "excluded_legal_prop_or_security_guide")
        self.assertEqual(security, "excluded_legal_prop_or_security_guide")

    def test_intent_gate_rejects_telegram_directory_and_normal_telecom_page(self) -> None:
        directory = relevance_gate_reason(
            "계좌매입 텔레그램 채널",
            (
                "엄선된 Telegram 채널, 그룹, 봇을 한 곳에서 모두 봅니다. "
                "구독자와 카테고리 순위, 최근 게시물 업데이트를 제공합니다."
            ),
            "https://directory.example/channel/account-buying",
            "unknown",
            "intent",
        )
        telecom = relevance_gate_reason(
            "인터넷디비 | 인터넷3사DB",
            (
                "인터넷 3사 비교상담으로 SK KT LG 요금제와 IPTV 결합, "
                "설치 가능 지역 및 약정 조건을 안내합니다."
            ),
            "https://isp.example/compare",
            "unknown",
            "intent",
        )
        self.assertEqual(directory, "excluded_telegram_directory")
        self.assertEqual(telecom, "excluded_normal_telecom_service")

    def test_intent_gate_rejects_unrelated_body_despite_sales_keyword_title(self) -> None:
        mismatch = relevance_gate_reason(
            "#여권제작 #여권위조",
            "가수의 공연과 팬들의 응원에 관한 오래된 일기입니다.",
            "https://blog.example/unrelated-post",
            "unknown",
            "intent",
        )
        title_only_listing = relevance_gate_reason(
            "모든 인스타 계정 매입합니다",
            "카카오톡 오픈채팅 문의 https://open.kakao.com/o/example",
            "https://market.example/account-post",
            "unknown",
            "intent",
        )
        self.assertEqual(mismatch, "excluded_title_body_mismatch")
        self.assertEqual(title_only_listing, "")

    def test_intent_gate_rejects_page_copywriting_seo_filler(self) -> None:
        reason = relevance_gate_reason(
            "자동차보험DB | 디비판매업체",
            (
                "이 안내형 페이지는 페이지 성격과 문단별 역할을 나누어 "
                "설명형 문장으로 구성합니다. 키워드를 반복 나열하지 않고 "
                "페이지 흐름이 자연스럽고 안정적으로 보이게 정리합니다."
            ),
            "https://seo.example/auto-insurance-db",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_seo_explainer")

    def test_intent_gate_rejects_ai_roleplay_plot_as_an_illicit_post(self) -> None:
        reason = relevance_gate_reason(
            "대포통장 판매 조직 - 제타",
            (
                "불법 도박사이트에 쓰이는 대포통장을 판매하는 조직. "
                "당신의 친누나이자 조직의 사장이라는 캐릭터 설정이다."
            ),
            "https://zeta-ai.io/ko/plots/example/profile",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_domain")

    def test_intent_gate_rejects_illicit_keyword_seo_agency_ad(self) -> None:
        reason = relevance_gate_reason(
            "수원꽃집 새아침식물원",
            (
                "네이버아이디판매 아이디 판매 유입 극대화의 핵심 비법. "
                "실행사의 상위 노출 유지력에 달려 있습니다. 최상단 고정이 "
                "안 되면 광고비를 받지 않겠습니다. 광고주 정보는 암호화하고 "
                "대행사의 키워드 노출 마케팅으로 구글 1페이지를 보장합니다."
            ),
            "https://m.sacgarden.net/board/gallery/read.html?no=2242&board_no=8",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_search_spam")

    def test_intent_gate_rejects_license_requirements_and_id_photo_tools(self) -> None:
        license_guide = relevance_gate_reason(
            "운전면허증에 대해 | 렌터카 안내",
            (
                "국제운전면허증 위조가 증가하여 발급국과 거주 증명을 "
                "확인합니다. 유효 기한과 제네바 조약 기준을 충족해야 "
                "차량 대여가 가능합니다."
            ),
            "https://rental.example/license.html",
            "unknown",
            "intent",
        )
        photo_tool = relevance_gate_reason(
            "온라인 신분증 사진 제작기",
            (
                "무료 신분증 사진 편집기로 여권과 면허증 사진을 "
                "자르고 배경, 조명, 크기를 조정해 만들어 보세요."
            ),
            "https://editor.example/id-photo-maker",
            "unknown",
            "intent",
        )
        self.assertEqual(license_guide, "excluded_identity_document_guide")
        self.assertEqual(photo_tool, "excluded_identity_photo_guide")

    def test_intent_gate_rejects_generic_article_with_sale_keyword_link(self) -> None:
        reason = relevance_gate_reason(
            "경제적인 비실명 ID로",
            (
                "온라인 입지를 확보하는 열쇠는 강력한 마케팅입니다. "
                "비실명 ID를 활용하는 것이 유리합니다. 이러한 ID는 "
                "소규모 회사에도 매력적인 옵션이며 다양한 접근을 가능하게 "
                "합니다. 네이버 아이디 판매 자세한 내용은 웹사이트를 참고하세요."
            ),
            "https://notes.example/article",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_seo_explainer")

    def test_intent_gate_rejects_news_aggregator_even_when_title_has_trade_terms(self) -> None:
        reason = relevance_gate_reason(
            "19년간 실종 한국인 행세…여권 위조 입국 집행유예",
            (
                "총 9개의 출처 보기. 여권을 위조해 입국한 중국인들에게 "
                "법원이 집행유예를 선고했다."
            ),
            "https://aagag.com/issue/?idx=1649889",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_press_domain")

    def test_intent_gate_rejects_documentary_id_card_market_and_db_pc_job(self) -> None:
        documentary = relevance_gate_reason(
            "한국여권 위조해 불법체류하는 중국인들",
            (
                "보더 시큐리티 프로그램입니다. 시드니 공항 직원이 "
                "승객을 인터뷰하고 통역사를 통해 위조 여권을 구했다는 "
                "사실을 확인합니다."
            ),
            "https://forum.example/documentary",
            "unknown",
            "intent",
        )
        id_card_market = relevance_gate_reason(
            "2025 신분증 제작 업체 트렌드",
            (
                "신분증 제작 업체 시장 개요와 제조사를 비교합니다. "
                "스마트 카드, NFC, RFID 기술은 보안과 접근 제어를 "
                "강화하는 정상적인 기업용 제품입니다."
            ),
            "https://market.example/id-card-trend",
            "unknown",
            "intent",
        )
        db_pc_job = relevance_gate_reason(
            "송내DB PC 평일오전 알바님을 찾습니다",
            "DB PC카페 시급 10,320원, 월~금 매장관리 알바 구인 공고입니다.",
            "https://jobs.example/job/1",
            "unknown",
            "intent",
        )
        self.assertEqual(documentary, "excluded_reporting_context")
        self.assertEqual(id_card_market, "excluded_normal_product_context")
        self.assertEqual(db_pc_job, "excluded_db_job_context")

    def test_intent_gate_keeps_weak_price_signal_with_concrete_contact(self) -> None:
        reason = relevance_gate_reason(
            "고객 DB 안내",
            "고객 DB 건당 단가는 텔레그램 [MESSENGER_ID]로 문의주세요.",
            "https://board.example/db-price",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_keeps_db_needed_offer_with_raw_messenger_handle(self) -> None:
        reason = relevance_gate_reason(
            "실시간 대출 디비 전문 실행사입니다",
            (
                "실시간 대출DB 필요하신 대표님들이라면 책임지고 "
                "진행해드립니다. 텔레그램 thomnews 로 문의주세요."
            ),
            "https://marketing.example/db-service/1",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_keeps_messenger_account_sales_offer(self) -> None:
        reason = relevance_gate_reason(
            "국내외 텔레그램·카카오톡 계정 유통점",
            (
                "카톡계정과 텔레그램 계정을 한 달에 10000개 이상 거래 "
                "가능합니다. 계정 판매 취급점은 구매상담 텔레그램 "
                "@talkseller1 로 연락주세요."
            ),
            "https://t.me/s/account_shop",
            "public_messenger_page",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_keeps_phone_and_email_data_sales_offer(self) -> None:
        reason = relevance_gate_reason(
            "텔레그램 전화번호 데이터",
            (
                "휴대전화 번호 데이터와 이메일 데이터를 제공합니다. "
                "최신 연락처 데이터베이스 패키지 구매는 Telegram "
                "@data_seller 로 문의주세요."
            ),
            "https://data.example/telegram-numbers",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "")

    def test_intent_gate_keeps_db_purchase_inquiry(self) -> None:
        reason = relevance_gate_reason(
            "대출DB 구매 문의드립니다",
            (
                "대부업 대출DB 단가는 어느 정도인가요? 거절난 대부업DB는 "
                "얼마에 구매해야 맞는지 문의드립니다."
            ),
            "https://community.example/question/buy-db",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "")

        no_contact_inquiry = relevance_gate_reason(
            "대출DB 문의드립니다",
            (
                "일수 관련 대부업DB의 단가는 어느 정도인가요? 일반 "
                "신용대출보다 다 거절난 대부업DB는 얼마에 구매해야 "
                "맞는 건지 모르겠습니다."
            ),
            "https://community.example/question/db-price",
            "unknown",
            "intent",
        )
        self.assertEqual(no_contact_inquiry, "")

    def test_intent_gate_rejects_normal_foreign_exchange_account(self) -> None:
        reason = relevance_gate_reason(
            "씨티은행 인터넷 뱅킹",
            (
                "계좌간 환전 서비스는 외화통장에 외환을 매입해 두거나 "
                "보유 외환을 매각하는 정상 은행 서비스입니다."
            ),
            "https://bank.example/foreign-exchange",
            "unknown",
            "intent",
        )
        self.assertNotEqual(reason, "")

    def test_intent_gate_rejects_additional_single_game_accounts(self) -> None:
        reason = relevance_gate_reason(
            "배틀그라운드 카카오 계정 판매",
            "제가 사용하던 계정을 10만원에 팝니다.",
            "https://market.example/post/7",
            "unknown",
            "intent",
        )
        self.assertEqual(reason, "excluded_single_account_trade")

        linked_game_account = relevance_gate_reason(
            "구글 계정 팝니다",
            "아이온 캐릭터 레벨 45인 계정 하나를 판매합니다.",
            "https://market.example/post/8",
            "unknown",
            "intent",
        )
        self.assertEqual(linked_game_account, "excluded_single_account_trade")

        league_account = relevance_gate_reason(
            "롤계정 대량 판매합니다",
            "리그오브레전드 계정을 판매합니다. 카톡 문의 가능합니다.",
            "https://game.example/lol-account",
            "unknown",
            "intent",
        )
        self.assertEqual(league_account, "excluded_single_account_trade")

        cookie_run_account = relevance_gate_reason(
            "카카오 쿠키런 계정 구매합니다",
            (
                "크리스탈 수급량이 높은 쿠키런 게임 계정을 구합니다. "
                "희망 가격과 계정 상세 정보를 보내주세요."
            ),
            "https://game.example/cookie-run-account",
            "unknown",
            "intent",
        )
        self.assertEqual(cookie_run_account, "excluded_single_account_trade")

    def test_press_page_classification_uses_path_and_byline(self) -> None:
        page_type = classify_page_type(
            "https://media.example/news/articleView.html?idxno=10",
            '"ID 삽니다" 국내외 계정 매매 성행',
            "홍길동 기자 = 관련 업계에 따르면 계정 거래가 늘고 있다.",
        )
        self.assertEqual(page_type, "news_or_education")

    def test_newsroom_page_classification_does_not_require_a_byline(self) -> None:
        page_type = classify_page_type(
            "https://security.example/about/news_list/view/360",
            "모바일 주민등록증 시스템 구축 - Security Newsroom",
            "자사 기술을 적용해 관련 사업을 수주했으며 시스템을 구축할 계획이다.",
        )
        self.assertEqual(page_type, "news_or_education")

    def test_strict_gate_rejects_reporting_and_generic_links(self) -> None:
        reporting = relevance_gate_reason(
            "고객 DB 판매 게시물 적발",
            "경찰이 텔레그램을 통해 거래한 사건을 검거했다는 기사입니다.",
            "https://news.example/article/1",
            "news_or_education",
            "strict",
        )
        generic_link = relevance_gate_reason(
            "고객 DB 판매합니다",
            "대량 보유 중이며 주문 문의는 홈페이지 [CONTACT_URL]에서 받습니다.",
            "https://board.example/post/1",
            "unknown",
            "strict",
        )
        self.assertEqual(reporting, "excluded_page_type")
        self.assertEqual(generic_link, "missing_concrete_contact")

    def test_strict_gate_rejects_scam_warning_disguised_by_sales_title(self) -> None:
        reason = relevance_gate_reason(
            "SNS 계정 싸게 팝니다",
            "사기입니다. 돈을 보낸 뒤 피해를 입어 사이버수사대에 신고했습니다.",
            "https://community.example/post/2",
            "unknown",
            "strict",
            "SNS 계정 판매 연락 [PHONE]",
        )
        self.assertEqual(reason, "excluded_reporting_context")

    def test_strict_gate_rejects_bulk_messaging_service_context(self) -> None:
        reason = relevance_gate_reason(
            "대량문자 발송 서비스",
            (
                "연락처를 주소록에 저장해 대량문자를 발송합니다. "
                "회원 할인 단가를 제공하며 상담은 텔그 [MESSENGER_ID]"
            ),
            "https://blog.example/post/3",
            "unknown",
            "strict",
        )
        self.assertEqual(reason, "missing_relevant_target")

    def test_strict_gate_keeps_account_verification_buying_post(self) -> None:
        reason = relevance_gate_reason(
            "가입인증 삽니다",
            "성인 본인인증 자료를 매입합니다. 문의 텔레그램 [MESSENGER_ID]",
            "https://board.example/post/4",
            "unknown",
            "strict",
        )
        self.assertEqual(reason, "")

    def test_strict_gate_rejects_single_game_account_trade(self) -> None:
        reason = relevance_gate_reason(
            "게임 계정 하나 팝니다",
            "실사용하던 계정을 10만원에 판매합니다. 문의 [PHONE]",
            "https://market.example/post/5",
            "unknown",
            "strict",
        )
        self.assertEqual(reason, "excluded_single_account_trade")

        linked_platform_account = relevance_gate_reason(
            "구글 계정 팝니다",
            "로드 모바일에서 실사용하던 구글 연동 계정을 판매합니다. 문의 [PHONE]",
            "https://market.example/post/6",
            "unknown",
            "strict",
        )
        self.assertEqual(linked_platform_account, "excluded_single_account_trade")

    def test_strict_gate_keeps_forgery_service_offer(self) -> None:
        reason = relevance_gate_reason(
            "각종 신분증 위조 전문",
            "주민등록증과 운전면허증 제작 가능. 의뢰 문의 텔레그램 [MESSENGER_ID]",
            "https://board.example/post/7",
            "unknown",
            "strict",
        )
        self.assertEqual(reason, "")

    def test_strict_gate_does_not_treat_body_link_words_as_document_type(self) -> None:
        reason = relevance_gate_reason(
            "스토리 계좌매입 채널",
            (
                "계좌매입 후 즉시 정산합니다. 편하게 문의주십쇼 "
                "텔레그램 [MESSENGER_ID]. 채널 바로가기와 AV위키 제휴 안내."
            ),
            "https://t.me/public_channel",
            "unknown",
            "strict",
        )
        self.assertEqual(reason, "")

    def test_strict_gate_rejects_explicitly_negated_offer(self) -> None:
        reason = relevance_gate_reason(
            "종토방 제휴",
            (
                "유심, 통장대여, 코인이체, 계정매입 등 피싱과 관련된 "
                "제휴는 받지 않습니다. 텔레그램 [MESSENGER_ID]"
            ),
            "https://t.me/public_channel",
            "public_messenger_page",
            "strict",
        )
        self.assertNotEqual(reason, "")

    def test_strict_gate_can_use_strong_discovery_evidence(self) -> None:
        reason = relevance_gate_reason(
            "서비스 홍보",
            "주식 고객 DB를 대량 보유하고 판매합니다. 건당 단가 안내 가능합니다.",
            "https://community.example/service/1",
            "unknown",
            "strict",
            "주식 디비 판매합니다. 텔레그램 https://t.me/private_handle 문의",
        )
        self.assertEqual(reason, "")

    def test_strict_gate_rejects_unrelated_destination_despite_search_snippet(self) -> None:
        reason = relevance_gate_reason(
            "기업 홈페이지",
            "ICT 인프라 구축과 기술 컨설팅 서비스를 제공합니다.",
            "https://company.example/",
            "unknown",
            "strict",
            "고객 DB 판매합니다. 텔레그램 https://t.me/private_handle 문의",
        )
        self.assertEqual(reason, "missing_relevant_target")

    def test_input_parameter_reflection_is_classified_as_search_reflection(self) -> None:
        page_type = classify_page_type(
            "https://calculator.example/input?i=customer+db+sale",
            "customer db sale - calculator",
            "customer db sale natural language input",
        )
        self.assertEqual(page_type, "search_reflection")

    def test_forum_index_is_classified_as_search_result_list(self) -> None:
        page_type = classify_page_type(
            "https://board.example/pds",
            "자료실",
            "번호 제목 작성자 작성일 추천 조회 3010 개인통장 매입 문의",
        )
        self.assertEqual(page_type, "search_result_list")

    def test_public_telegram_page_has_distinct_page_type(self) -> None:
        page_type = classify_page_type(
            "https://t.me/public_channel",
            "계좌매입 채널",
            "계좌 매입 문의 텔레그램 [MESSENGER_ID]",
        )
        self.assertEqual(page_type, "public_messenger_page")

    def test_labeling_gate_keeps_topical_hard_negative(self) -> None:
        reason = relevance_gate_reason(
            "개인정보 유출 사고 안내",
            "피해 확인 방법을 설명하며 판매나 구매 의사는 없는 기사입니다.",
            "https://news.example/article/7",
            "news_or_education",
            "labeling",
        )
        self.assertEqual(reason, "")
        sidebar_noise = relevance_gate_reason(
            "여름맞이 경품 이벤트",
            "이벤트 안내입니다. 인기글: 고객 DB 판매 관련 문의",
            "https://community.example/event/9",
            "unknown",
            "labeling",
        )
        self.assertEqual(sidebar_noise, "missing_relevant_target")
        trade_title = relevance_gate_reason(
            "대량 판매합니다",
            "자세한 거래 대상은 게시물 본문을 확인하세요.",
            "https://community.example/post/10",
            "unknown",
            "labeling",
        )
        contact_with_target_lead = relevance_gate_reason(
            "텔레그램 문의",
            "고객 DB와 계정 관련 내용을 안내합니다.",
            "https://community.example/post/11",
            "unknown",
            "labeling",
        )
        generic_contact = relevance_gate_reason(
            "문의하기",
            "행사 참여 방법을 안내합니다.",
            "https://community.example/event/12",
            "unknown",
            "labeling",
        )
        self.assertEqual(trade_title, "")
        self.assertEqual(contact_with_target_lead, "")
        self.assertEqual(generic_contact, "missing_relevant_target")

    def test_relevance_gate_rejects_policy_and_search_pages(self) -> None:
        policy_reason = relevance_gate_reason(
            "개인정보 처리방침",
            "고객정보를 보유하며 상품 구매 문의는 고객센터로 연락하세요.",
            "https://shop.example/privacy",
            "unknown",
            "review",
        )
        search_reason = relevance_gate_reason(
            "고객 DB 판매 검색",
            "검색 결과입니다. 텔레그램 문의",
            "https://board.example/search?q=test",
            "search_reflection",
            "review",
        )
        self.assertEqual(policy_reason, "excluded_document_type")
        self.assertEqual(search_reason, "excluded_page_type")
        self.assertEqual(
            relevance_gate_reason(
                "검색 결과",
                "공개 게시물 목록",
                "https://board.example/search?q=test",
                "search_result_list",
                "off",
            ),
            "excluded_page_type",
        )

    def test_review_gate_requires_title_signal_but_keeps_topical_news(self) -> None:
        generic = relevance_gate_reason(
            "전자계약 서비스 주요 기능",
            "고객 개인정보를 보유하며 계약 거래 문의는 [EMAIL]로 받습니다.",
            "https://service.example/features",
            "unknown",
            "review",
        )
        generic_account_trade = relevance_gate_reason(
            "거래중지 계좌 해지 방법",
            "오래 사용하지 않은 통장의 거래를 다시 시작하고 싶습니다.",
            "https://qna.example/question/2",
            "unknown",
            "review",
        )
        topical_news = relevance_gate_reason(
            "고객 DB 판매 게시물 적발",
            "개인정보 명단을 대량으로 거래한 사례가 확인됐다는 보도입니다.",
            "https://news.example/article/1",
            "news_or_education",
            "review",
        )
        self.assertEqual(generic, "missing_title_signal")
        self.assertEqual(generic_account_trade, "missing_title_signal")
        self.assertEqual(topical_news, "")

    def test_seed_csv_deduplicates_urls(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "seeds.local.csv"
            path.write_text(
                "url,detection_type,query_group\n"
                "https://example.com/post,기타,seed\n"
                "https://example.com/post#fragment,기타,seed\n",
                encoding="utf-8",
            )
            candidates = load_seed_candidates(path)
            self.assertEqual(len(candidates), 1)
            self.assertEqual(candidates[0].url, "https://example.com/post")

    def test_seed_csv_accepts_restricted_raw_url_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "urls.csv"
            path.write_text(
                "sample_id,raw_url\n"
                "LP-000001,https://example.com/public/post\n",
                encoding="utf-8",
            )
            candidates = load_seed_candidates(path)
            self.assertEqual(len(candidates), 1)
            self.assertEqual(candidates[0].url, "https://example.com/public/post")

    def test_seed_csv_accepts_shareable_source_url_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "data.csv"
            path.write_text(
                "sample_id,source_url\n"
                "LP-000001,https://example.com/public/post\n",
                encoding="utf-8",
            )
            candidates = load_seed_candidates(path)
            self.assertEqual(len(candidates), 1)
            self.assertEqual(candidates[0].url, "https://example.com/public/post")

    def test_prior_search_queue_is_prefiltered_but_manual_seed_is_retained(self) -> None:
        direct = Candidate(
            "https://example.com/direct",
            "group",
            "개인정보DB",
            discovery_text=(
                "고객 DB 판매합니다. 텔레그램 https://t.me/direct 문의"
            ),
        )
        reporting = Candidate(
            "https://example.com/report",
            "group",
            "개인정보DB",
            discovery_text="고객 DB 판매 사건을 경찰이 적발했다는 기사",
        )
        manual = Candidate(
            "https://example.com/manual",
            "seed",
            "기타",
            source_type="seed",
        )
        empty_search_artifact = Candidate(
            "https://search.example/navigation",
            "group",
            "기타",
            source_type="search",
        )
        filtered = prefilter_seed_candidates(
            [direct, reporting, manual, empty_search_artifact], "strict"
        )
        self.assertEqual(filtered, [direct, manual])

    def test_resume_prefilter_can_tighten_a_broad_search_queue(self) -> None:
        broad_false_positive = Candidate(
            "https://news.example/report",
            "group",
            "기타",
            discovery_text="불법 신분증 위조 판매 사건을 경찰이 적발했다",
            search_provider="naver",
        )
        direct_offer = Candidate(
            "https://seller.example/post",
            "group",
            "기타",
            discovery_text="신분증 위조 제작합니다 텔레그램 문의",
            search_provider="naver",
        )
        self.assertEqual(
            prefilter_seed_candidates(
                [broad_false_positive, direct_offer],
                "intent",
            ),
            [direct_offer],
        )

    def test_private_candidate_queue_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ".private" / "candidate_queue.jsonl"
            expected = [Candidate("https://example.com/post", "group", "기타")]
            save_candidate_queue(path, expected)
            loaded = load_candidate_queue(path)
            self.assertEqual(loaded, expected)
            self.assertEqual(load_seed_candidates(path), expected)
            self.assertEqual(path.stat().st_mode & 0o777, 0o600)

    def test_prior_sample_urls_can_be_loaded_for_holdout_exclusion(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "prior.csv"
            path.write_text(
                "source_url,title\n"
                "https://example.com/post?a=1&utm_source=test,first\n"
                "https://example.com/other,second\n",
                encoding="utf-8",
            )
            self.assertEqual(
                load_excluded_urls([path]),
                {
                    "https://example.com/post?a=1",
                    "https://example.com/other",
                },
            )

    def test_naver_blog_desktop_mobile_urls_share_document_identity(self) -> None:
        desktop = canonicalize_url(
            "https://blog.naver.com/qnaj2330/223196767893?from=postList"
        )
        mobile = canonicalize_url(
            "https://m.blog.naver.com/qnaj2330/223196767893"
        )
        self.assertEqual(desktop, mobile)

    def test_marketplace_post_search_parameters_do_not_create_new_documents(self) -> None:
        plain = canonicalize_url("https://creativebox.kr/igtrade/5742")
        filtered = canonicalize_url(
            "https://www.creativebox.kr/igtrade/5742?"
            "sfl=mb_id%2C1&stx=example&page=3"
        )
        self.assertEqual(plain, filtered)

    def test_url_canonicalization_preserves_ipv6_brackets(self) -> None:
        self.assertEqual(
            canonicalize_url("https://[2001:4860:4860::8888]/post#fragment"),
            "https://[2001:4860:4860::8888]/post",
        )

    def test_prior_sample_fingerprints_are_loaded_for_holdout_exclusion(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "prior.csv"
            path.write_text(
                "source_url,near_duplicate_fingerprint\n"
                "https://example.com/one,simhash64:1111\n"
                "https://example.com/two,simhash64:2222\n",
                encoding="utf-8",
            )
            self.assertEqual(
                load_excluded_fingerprints([path]),
                {"simhash64:1111", "simhash64:2222"},
            )

    def test_terminal_failures_are_skipped_on_resume(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "collection_log.csv"
            path.write_text(
                "url_hmac,query_group,outcome,http_status,reason,text_chars,extraction_method\n"
                "done,g,skipped,,robots_disallowed,0,\n"
                "short,g,failed,200,insufficient_text,20,visible_body_fallback\n"
                "retry,g,failed,,ReadTimeout,0,\n",
                encoding="utf-8",
            )
            self.assertEqual(terminal_attempt_hashes(path), {"done", "short"})

    def test_collection_log_upgrade_adds_attempt_time_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "collection_log.csv"
            path.write_text(
                "url_hmac,query_group,outcome,http_status,reason,text_chars,extraction_method\n"
                "abc,g,failed,200,insufficient_text,20,visible_body_fallback\n",
                encoding="utf-8",
            )
            upgrade_collection_log_schema(path)
            with path.open(encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                rows = list(reader)
            self.assertIn("attempted_at", reader.fieldnames)
            self.assertEqual(rows[0]["text_chars"], "20")

    def test_existing_dataset_upgrade_preserves_legacy_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "candidates_masked.csv"
            legacy_fields = [
                field
                for field in SCHEMA
                if field not in {"extraction_status", "near_duplicate_fingerprint"}
            ]
            row = {field: "" for field in legacy_fields}
            row.update(
                {
                    "sample_id": "EG-0001",
                    "source_type": "public_web_search",
                    "live_status": "true",
                    "page_type": "reflected_search_page",
                    "near_duplicate_cluster": "simhash64:1234567890abcdef",
                }
            )
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=legacy_fields)
                writer.writeheader()
                writer.writerow(row)
            upgrade_existing_csv_schema(path)
            with path.open(encoding="utf-8-sig", newline="") as handle:
                upgraded = next(csv.DictReader(handle))
            self.assertEqual(upgraded["sample_id"], "EG-0001")
            self.assertEqual(upgraded["source_type"], "search")
            self.assertEqual(upgraded["live_status"], "accessible")
            self.assertEqual(upgraded["page_type"], "search_reflection")
            self.assertEqual(
                upgraded["near_duplicate_fingerprint"],
                upgraded["near_duplicate_cluster"],
            )

    def test_extraction_failure_has_standard_status_and_no_raw_url(self) -> None:
        log = CollectionLog(
            "digest-only",
            "group",
            "failed",
            "200",
            "insufficient_text",
            25,
            "main_container_short",
        )
        failure = extraction_failure_record(log)
        self.assertIsNotNone(failure)
        self.assertEqual(failure["extraction_status"], "partial")
        self.assertRegex(failure["attempted_at"], r"^\d{4}-\d{2}-\d{2}T")

    def test_record_uses_standard_handoff_values(self) -> None:
        candidate = Candidate(
            "https://example.com/post/1", "seed-group", "기타", source_type="seed"
        )
        record = make_record(
            7,
            candidate,
            candidate.url,
            200,
            "테스트 제목",
            "충분한 공개 예시 본문입니다. " * 10,
            b"test-key",
        )
        self.assertEqual(record["sample_id"], "EG-000007")
        self.assertEqual(record["source_type"], "seed")
        self.assertEqual(record["live_status"], "accessible")
        self.assertEqual(record["extraction_status"], "success")
        self.assertEqual(
            record["near_duplicate_cluster"],
            record["near_duplicate_fingerprint"],
        )

    def test_masking_report_and_manifest_include_integrity_data(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            out = Path(temp_dir)
            csv_path = out / "candidates_masked.csv"
            csv_path.write_text(
                "masked_title,masked_text\n제목,연락처 [PHONE]\n",
                encoding="utf-8",
            )
            report = masking_validation(csv_path, "pilot-v1")
            self.assertTrue(report["passed"])
            manifest = data_manifest(out, "pilot-v1", [csv_path], {"target": 1})
            self.assertEqual(manifest["files"][0]["rows"], 1)
            self.assertRegex(manifest["files"][0]["sha256"], r"^[0-9a-f]{64}$")
            self.assertEqual(manifest["settings"]["target"], 1)

    def test_extracts_short_post_from_main_container(self) -> None:
        html = """
        <html><head><title>테스트 게시물</title></head><body>
        <nav>메뉴 메뉴 메뉴</nav>
        <main><p>이것은 본문 추출 테스트를 위한 공개 예시 문장입니다.</p>
        <p>게시물의 핵심 내용이 메뉴보다 우선해서 저장되어야 합니다.</p></main>
        </body></html>
        """
        title, text, method = extract_title_text(html, "https://example.com/post")
        self.assertEqual(title, "테스트 게시물")
        self.assertIn("본문 추출 테스트", text)
        self.assertNotIn("메뉴 메뉴", text)
        self.assertIn(method, {"trafilatura_precision", "main_container"})

    def test_legacy_board_post_beats_longer_footer(self) -> None:
        html = """
        <html><head><title>여권발급이나 새신분이 필요하신분</title></head>
        <body>
          <table><tr><td class="con_f">
            여권과 신분증 위조 제작 가능합니다. 판매 문의는
            텔레그램 sample_handle 또는 카톡 sample_chat으로 주세요.
            신청 대상과 제작 종류를 확인한 뒤 신속하게 안내한다는 게시물입니다.
          </td></tr></table>
          <footer>고객센터 문의와 저작권 안내입니다. """ + "일반 안내 " * 80 + """</footer>
        </body></html>
        """
        title, text, method = extract_title_text(
            html, "https://board.example/public/post/1"
        )
        self.assertIn("여권과 신분증 위조 제작", text)
        self.assertNotIn("저작권 안내", text)
        self.assertEqual(method, "strong_post_container")

    def test_challenge_page_fails_text_quality_gate(self) -> None:
        text = "Checking your browser before accessing the requested public page."
        self.assertEqual(
            text_quality_reason(text, 40), "challenge_or_access_page"
        )

    def test_korean_text_gate_rejects_english_only_page(self) -> None:
        text = "This is a sufficiently long English page with meaningful words."
        self.assertEqual(
            text_quality_reason(text, 40, minimum_korean_chars=5),
            "insufficient_korean_text",
        )

    def test_related_internal_links_are_bounded_and_same_site(self) -> None:
        html = """
        <a href="/board/view?id=2">고객 DB 관련 게시물</a>
        <a href="/login">로그인</a>
        <a href="https://outside.example/post/3">개인정보</a>
        <a href="/search?q=개인정보">검색</a>
        """
        links = discover_related_internal_links(
            html, "https://example.com/board/view?id=1", 5
        )
        self.assertEqual(links, ["https://example.com/board/view?id=2"])

    def test_template_is_preserved_and_rows_expand(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "restricted" / "result.xlsx"
            workbook, sheet = prepare_detection_workbook(TEMPLATE, output, resume=False)
            entries = [
                DetectionEntry(
                    detected_on=dt.date(2026, 8, 17),
                    url=f"https://example.com/post/{index}",
                    detection_type="개인정보DB",
                    registrant="테스트",
                )
                for index in range(1, 35)
            ]
            append_detection_entries(sheet, entries)
            save_restricted_workbook(workbook, output)

            saved = load_workbook(output)["8월"]
            self.assertEqual(saved["A4"].value, 1)
            self.assertEqual(saved["A37"].value, 34)
            self.assertEqual(saved["D4"].value, "개인정보DB")
            self.assertEqual(saved["E4"].value, "테스트")
            self.assertEqual(saved["C4"].hyperlink.target, "https://example.com/post/1")
            self.assertIn("예시 내용", saved["H4"].value)
            self.assertEqual(output.stat().st_mode & 0o777, 0o600)


if __name__ == "__main__":
    unittest.main()
